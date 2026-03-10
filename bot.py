"""
📋 CHANGELOG - bot.py v4.7.7

✅ OHP, KOKSARAY, DATA CENTR VE MMP ŞANTİYESİ OPSİYONEL MOD
- OHP şantiyesi artık opsiyonel rapor modunda çalışıyor
- Rapor gönderilirse işlenir, gönderilmezse eksik listesine dahil edilmez
- Hatırlatma mesajlarında görünmez
- Admin kontrol listelerinde görünmez
- Excel eksik rapor analizinde yer almaz
- Yalnızca OHP için bu özel durum geçerlidir

✅ ÇALIŞMA YOK RAPORLARI DÜZELTMESİ
- "Çalışma yok", "iş yok", "faaliyet yok" gibi raporlar artık doğru şekilde işleniyor
- Tüm personel kategorileri 0 olarak kaydediliyor
- GENEL TOPLAM: 0 olarak hesaplanıyor
- Şantiye bazlı sistemde eksik rapor listesinden çıkarılıyor

✅ HAFTALIK RAPOR TARİH DÜZELTMESİ
- Haftalık rapor artık Pazar 09:00'da doğru tarih aralığı ile gönderiliyor
- Haftalık eksik raport artık Pazar 10:00'da doğru tarih aralığı ile gönderiliyor
- Aylık rapor artık ayın 1'i 08:30'da doğru tarih aralığı ile gönderiliyor
- Aylık eksik rapor artık ayın 1'i 08:45'de doğru tarih aralığı ile gönderiliyor

✅ 7/24 ÇALIŞMA SİSTEMİNE GEÇİŞ
- Hafta sonları (Cumartesi-Pazar) artık tatil değil, çalışma günü
- Tüm raporlarda hafta sonları dahil ediliyor
- Eksik rapor analizinde hafta sonları da kontrol ediliyor
- Haftalık ve aylık raporlarda tüm günler dahil

✅ KRİTİK DÜZELTMELER: TOPLAMA VE YÜZDE HESAPLAMA
- GENEL TOPLAM hesaplaması düzeltildi: Tüm kategorilerin toplamı alınır
- Yüzde hesaplama düzeltildi: (kategori_toplamı / genel_toplam) * 100
- MOS şantiyesi eklendi: Sorumlu @OrhanCeylan
- Haftalık ve aylık raporlarda personel dağılımı yüzdeleri doğru hesaplanıyor
- EKSİK RAPOR ANALİZİ eklendi: Excel ve detaylı raporlama

✅ ZAMANLAMA DÜZELTMELERİ
- HAFTALIK NORMAL RAPOR: Her Pazar 09:00 (7 günlük periyot: Pazartesi 00:00 - Pazar 00:00)
- HAFTALIK EKSİK RAPOR: Her Pazar 10:00 (Haftalık normal raporla aynı tarih aralığı)
- AYLIK NORMAL RAPOR: Her ayın 1'i 08:30 (Bir önceki ayın tamamı)
- AYLIK EKSİK RAPOR: Her ayın 1'i 08:45 (Aylık normal raporla aynı tarih aralığı)
"""

import os
import re
import psycopg2
import pandas as pd
import json
import datetime as dt
import logging
import asyncio
import functools
import tempfile
import requests
import html
import base64
import time as time_module
import hashlib
import subprocess
import shlex
from unicodedata import normalize
from dotenv import load_dotenv
from typing import Dict, List, Tuple

# Çevre değişkenlerini en başta yükle
load_dotenv()

# Loglama ayarı - Railway için konsol çıktısı (EN ÜSTE)
logging.basicConfig(
    format="%(asctime)s %(levelname)s [%(filename)s:%(lineno)d] %(message)s",
    level=logging.INFO,
    handlers=[logging.StreamHandler()]
)

# Railway için PORT ayarı
PORT = int(os.environ.get('PORT', 8443))
logging.info(f"🚀 Railway PORT: {PORT}")

try:
    from telegram import Update, BotCommand, BotCommandScopeAllPrivateChats
    HAS_PRIVATE_SCOPE = True
except Exception as e:
    HAS_PRIVATE_SCOPE = False
    logging.warning(f"BotCommandScopeAllPrivateChats yüklenemedi: {e}")

from telegram.ext import (
    Application, MessageHandler, CommandHandler, ContextTypes, filters
)
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from psycopg2 import pool
from bs4 import BeautifulSoup
from openai import OpenAI

# Çevre değişkeni doğrulama
def validate_environment():
    """Gerekli tüm çevre değişkenlerini doğrula"""
    required_vars = {
        'BOT_TOKEN': 'Telegram Bot Token',
        'DATABASE_URL': 'PostgreSQL Veritabanı URL',
        'OPENAI_API_KEY': 'OpenAI API Anahtarı'
    }
    
    missing_vars = []
    for var, description in required_vars.items():
        if not os.getenv(var):
            missing_vars.append(f"{var} ({description})")
    
    if missing_vars:
        error_msg = f"❌ Eksik çevre değişkenleri: {', '.join(missing_vars)}"
        logging.error(error_msg)
        raise RuntimeError(error_msg)
    
    logging.info("✅ Tüm gerekli çevre değişkenleri ayarlanmış")

# İçe aktarımda çevre değişkenlerini doğrula
validate_environment()

# Veritabanı bağlantı havuzu
DB_POOL = None

def init_db_pool():
    """Hata yönetimi ile veritabanı bağlantı havuzunu başlat"""
    global DB_POOL
    try:
        if DB_POOL is None:
            DB_POOL = pool.ThreadedConnectionPool(
                minconn=1, 
                maxconn=10, 
                dsn=os.environ['DATABASE_URL'], 
                sslmode='require'
            )
            logging.info("✅ Veritabanı bağlantı havuzu başlatıldı")
    except Exception as e:
        logging.error(f"❌ Veritabanı havuzu başlatma hatası: {e}")
        raise

def get_conn_from_pool():
    """Doğrulama ile havuzdan bağlantı al"""
    if DB_POOL is None:
        init_db_pool()
    
    try:
        conn = DB_POOL.getconn()
        if conn.closed:
            logging.warning("⚠️ Bağlantı kapalıydı, yeni oluşturuluyor")
            DB_POOL.putconn(conn)
            conn = DB_POOL.getconn()
        return conn
    except Exception as e:
        logging.error(f"❌ Havuzdan bağlantı alma hatası: {e}")
        raise

def put_conn_back(conn):
    """Bağlantıyı havuza güvenli şekilde geri ver"""
    try:
        if DB_POOL and conn and not conn.closed:
            DB_POOL.putconn(conn)
    except Exception as e:
        logging.error(f"❌ Bağlantıyı havuz iade etme hatası: {e}")

# Güvenli veritabanı yardımcı fonksiyonları
def _sync_fetchall_safe(query, params=()):
    """Güvenli sorgu çalıştır ve tuple index koruması ile tüm sonuçları döndür"""
    conn = get_conn_from_pool()
    cur = None
    try:
        cur = conn.cursor()
        cur.execute(query, params)
        rows = cur.fetchall()
        return rows if rows else []
    except Exception as e:
        logging.error(f"Veritabanı fetchall hatası: {e}")
        return []
    finally:
        if cur:
            cur.close()
        put_conn_back(conn)

def _sync_execute_safe(query, params=()):
    """Güvenli sorgu çalıştır ve satır sayısını döndür"""
    conn = get_conn_from_pool()
    cur = None
    try:
        cur = conn.cursor()
        cur.execute(query, params)
        conn.commit()
        return cur.rowcount
    except Exception as e:
        conn.rollback()
        logging.error(f"Veritabanı execute hatası: {e}")
        return 0
    finally:
        if cur:
            cur.close()
        put_conn_back(conn)

def _sync_fetchone_safe(query, params=()):
    """Güvenli sorgu çalıştır ve tuple index koruması ile tek sonuç döndür"""
    conn = get_conn_from_pool()
    cur = None
    try:
        cur = conn.cursor()
        cur.execute(query, params)
        row = cur.fetchone()
        return row if row else None
    except Exception as e:
        logging.error(f"Veritabanı fetchone hatası: {e}")
        return None
    finally:
        if cur:
            cur.close()
        put_conn_back(conn)

# Async veritabanı operasyonları
async def async_db_query(func, *args, **kwargs):
    """Executor içinde veritabanı sorgusu çalıştır"""
    loop = asyncio.get_running_loop()
    try:
        return await loop.run_in_executor(None, functools.partial(func, *args, **kwargs))
    except Exception as e:
        logging.error(f"Async DB sorgu hatası: {e}")
        raise

async def async_fetchall(query, params=()):
    """Güvenli tuple işleme ile async fetchall"""
    try:
        result = await async_db_query(_sync_fetchall_safe, query, params)
        return result if result else []
    except Exception as e:
        logging.error(f"Async fetchall hatası - Sorgu: {query}, Parametreler: {params}, Hata: {e}")
        return []

async def async_execute(query, params=()):
    """Güvenli işleme ile async execute"""
    return await async_db_query(_sync_execute_safe, query, params)

async def async_fetchone(query, params=()):
    """Güvenli tuple işleme ile async fetchone"""
    try:
        result = await async_db_query(_sync_fetchone_safe, query, params)
        return result
    except Exception as e:
        logging.error(f"Async fetchone hatası - Sorgu: {query}, Parametreler: {params}, Hata: {e}")
        return None

def safe_get_tuple_value(tuple_data, index, default=None):
    """Index sınır kontrolü ile tuple'dan güvenli değer alma"""
    if tuple_data is None:
        return default
    
    if isinstance(tuple_data, (tuple, list)) and len(tuple_data) > index:
        value = tuple_data[index]
        return value if value is not None else default
    
    return default

# Gelişmiş JSON parsing ile doğrulama
def safe_json_loads(json_string, default=None):
    """Kapsamlı hata yönetimi ile güvenli JSON string parsing"""
    if json_string is None:
        return default
    
    # Eğer zaten dict ise, doğrudan döndür
    if isinstance(json_string, dict):
        return json_string
    
    try:
        return json.loads(json_string)
    except json.JSONDecodeError as e:
        logging.error(f"JSON decode hatası: {e}, Girdi: {json_string[:100]}...")
        return default
    except Exception as e:
        logging.error(f"Beklenmeyen JSON parsing hatası: {e}")
        return default

# GELİŞMİŞ EXCEL OKUMA - YENİ FORMAT DESTEĞİ
def safe_read_excel(file_path, required_columns=None):
    """
    GELİŞTİRİLDİ: Yeni Excel formatını destekler
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel dosyası bulunamadı: {file_path}")
    
    try:
        df = pd.read_excel(file_path)
        
        # Gerekli kolonları doğrula (esnek)
        if required_columns:
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                logging.warning(f"⚠️ Eksik kolonlar: {missing_columns}. Mevcut kolonlar: {list(df.columns)}")
                
                # YENİ FORMAT EŞLEŞTİRME - GÜNCELLENDİ
                column_mapping = {
                    'Rol': 'Botdaki Statusu / Rol',
                    'Botdaki Statusu': 'Botdaki Statusu / Rol',
                    'Kullanici Adi Soyadi': 'Kullanici Adi Soyadi',
                    'Telegram ID': 'Telegram ID', 
                    'Proje / Şantiye': 'Proje / Şantiye',
                    'Aktif / Pasif': 'Aktif / Pasif',
                    # YENİ EŞLEŞMELER
                    'Username': 'Username',
                    'Telefon Numarası': 'Telefon Numarası',
                    'Pozisyon Kodu': 'Pozisyon Kodu',
                    'Özel Rapor': 'Özel Rapor'
                }
                
                for required_col in missing_columns:
                    if required_col in column_mapping and column_mapping[required_col] in df.columns:
                        # Mevcut kolonu kullan
                        df[required_col] = df[column_mapping[required_col]]
                        logging.info(f"✅ {required_col} için {column_mapping[required_col]} kolonu kullanıldı")
                    else:
                        # Varsayılan değerlerle ekle
                        if required_col == "Rol":
                            df[required_col] = "KULLANICI"
                        elif required_col == "Botdaki Statusu":
                            df[required_col] = "Aktif"
                        elif required_col == "Aktif / Pasif":
                            df[required_col] = "E"
                        else:
                            df[required_col] = ""
        
        return df
    except Exception as e:
        logging.error(f"Excel okuma hatası: {e}")
        raise

# Timeout ile gelişmiş HTTP istekleri
def safe_http_request(url, method='GET', timeout=30, **kwargs):
    """Timeout ve hata yönetimi ile HTTP isteği yap"""
    try:
        response = requests.request(method, url, timeout=timeout, **kwargs)
        response.raise_for_status()
        return response
    except requests.exceptions.Timeout:
        logging.error(f"HTTP istek timeout: {url}")
        return None
    except requests.exceptions.RequestException as e:
        logging.error(f"HTTP istek hatası: {e}")
        return None

# Helper function for integer conversion - YENİ TELEGRAM ID PARSING
def _to_int_or_none(x):
    """Güvenli şekilde integer'a çevir veya None döndür - YENİ: 8-10 digit Telegram ID"""
    if x is None or pd.isna(x):
        return None
    
    s = str(x).strip()
    if not s:
        return None
    
    # Bilimsel gösterim kontrolü
    if "e+" in s.lower():
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return None
    
    # Sadece rakamları al
    s_clean = re.sub(r'[^\d]', '', s)
    
    if not s_clean:
        return None
    
    # YENİ: 8-10 digit Telegram ID kontrolü
    if len(s_clean) < 8 or len(s_clean) > 10:
        # Özel durum: 10 digit ID'ler kabul edilir
        if len(s_clean) == 10:
            pass
        else:
            return None
    
    try:
        return int(s_clean)
    except (ValueError, TypeError):
        return None

def get_file_hash(filename):
    """Değişiklik tespiti için dosya hash'ini al"""
    try:
        if os.path.exists(filename):
            with open(filename, 'rb') as f:
                return hashlib.md5(f.read()).hexdigest()
        return None
    except Exception as e:
        logging.error(f"Dosya hash hatası: {e}")
        return None

# Konfigürasyon
BOT_TOKEN = os.getenv("BOT_TOKEN")
CHAT_ID = os.getenv("CHAT_ID")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
try:
    GROUP_ID = int(CHAT_ID) if CHAT_ID else None
    logging.info(f"✅ GROUP_ID başarıyla ayarlandı: {GROUP_ID}")
except (ValueError, TypeError) as e:
    GROUP_ID = None
    logging.error(f"❌ GROUP_ID ayarlanamadı: {e}")
TZ = ZoneInfo("Asia/Tashkent")

SUPER_ADMIN_ID = 1000157326

# Fallback kullanıcı veri yapısı
FALLBACK_USERS = [
    {
        "Telegram ID": 1000157326,
        "Kullanici Adi Soyadi": "Atamurat Kamalov", 
        "Aktif / Pasif": "E",
        "Rol": "SÜPER ADMIN",
        "Botdaki Statusu": "Aktif",
        "Proje / Şantiye": "TYM"
    },
    {
        "Telegram ID": 709746899,
        "Kullanici Adi Soyadi": "Eren Boz",
        "Aktif / Pasif": "E", 
        "Rol": "ADMIN",
        "Botdaki Statusu": "Aktif",
        "Proje / Şantiye": "TYM"
    }
]

USERS_FILE = "Kullanicilar.xlsx"

# Global değişkenler başlatma
df = None
rapor_sorumlulari = []
id_to_name = {}
id_to_projects = {}
id_to_status = {}
id_to_rol = {}
ADMINS = []
IZLEYICILER = []
TUM_KULLANICILAR = []
santiye_sorumlulari = {}
santiye_rapor_durumu = {}
last_excel_update = 0
excel_file_hash = None
excel_last_modified = 0

user_role_cache = {}
user_role_cache_time = 0

# Sabit şantiye listesi - TÜM raporlarda kullanılacak
SABIT_SANTIYELER = ['BWC', 'DMC', 'STADYUM', 'LOT13', 'LOT71', 'SKP', 'YHP', 'TYM', 'RMC', 'PİRAMİT', 'MOS',]

# OPSİYONEL ŞANTİYELER (rapor gönderilirse işlenir, gönderilmezse eksik sayılmaz)
OPSIYONEL_SANTIYELER = ['OHP', 'DATA CENTR', 'MMP', 'KÖKSARAY']  # OHP, MMP,KÖKSARAY ve DATA CENTR opsiyonel oldu

# Şantiye bazlı kullanıcı adı (username) eşlemesi - HATIRLATMA MESAJLARI İÇİN
SANTIYE_USERNAME_MAPPING = {
    'BWC': ['YsF1434'],
    'SKP': ['uzyusufmutlu'],
    'DMC': ['umut61x'],
    'STADYUM': ['Adnan_Keleş'],
    'LOT13': ['Adnan_Keleş'],
    'LOT71': ['Adnan_Keleş'],
    'YHP': ['Orhan_Ceylan'],
    'RMC': ['Orhan_Ceylan'],
    'TYM': ['Orhan_Ceylan'],
    'PİRAMİT': ['ON5428'],
    'MOS': ['Orhan_Ceylan'],
    # OHP, MMP, KOKSARAY ve DATA CENTR opsiyonel olduğu için kaldırıldı
}

# Giriş doğrulama fonksiyonları
def validate_user_input(text, max_length=5000):
    """Kullanıcı giriş metnini doğrula"""
    if not text or not isinstance(text, str):
        return False, "Giriş boş olmayan string olmalı"
    
    if len(text) > max_length:
        return False, f"Giriş çok uzun (maksimum {max_length} karakter)"
    
    # Temizleme
    text = html.escape(text.strip())
    
    return True, text

def validate_date_string(date_str):
    """Tarih string formatını doğrula"""
    try:
        dt.datetime.strptime(date_str, '%Y-%m-%d')
        return True
    except ValueError:
        return False

def normalize_site_name(site_name):
    """Şantiye isimlerini standartlaştır"""
    if not site_name:
        return "BELİRSİZ"
        
    site_name = site_name.upper().strip()
    
    mappings = {
        'LOT 13': 'LOT13',
        'LOT-13': 'LOT13', 
        'LOT13': 'LOT13',
        'LOT 71': 'LOT71',
        'LOT-71': 'LOT71',
        'LOT71': 'LOT71',
        'SKP DAHO': 'SKP',
        'SKP': 'SKP',
        '📍 SKP Elektrik Grubu': 'SKP',
        'SKP Elektrik Grubu': 'SKP',
        ' SKP Elektrik Grubu': 'SKP',
        'PİRAMİT TOWER': 'PİRAMİT',
        'PİRAMİT': 'PİRAMİT',
        'PRAMİT': 'PİRAMİT',
        'PRAMIT': 'PİRAMİT',
        'PİRAMİT TOWEr': 'PİRAMİT',
        'PİRAMİT TOWAR': 'PİRAMİT',
        'PIRAMIT': 'PİRAMİT',
        'PIRAMIT TOWER': 'PİRAMİT',
        'PİRAMİD': 'PİRAMİT',
        'PIRAMID': 'PİRAMİT',
        'PYRAMIT': 'PİRAMİT',
        'PYRAMID': 'PİRAMİT',
        'BWC': 'BWC',
        'STADYUM': 'STADYUM',
        'DMC ELLIPSE GARDEN': 'DMC',
        'DMC ELLIPSE': 'DMC',
        'DMC GARDEN': 'DMC',
        'DMC Ellipse Garden Elektrik Grubu': 'DMC',
        'DMC ELLIPSE GARDEN ELEKTRIK GRUBU': 'DMC',
        'DMC ELLIPSE GARDEN ELEKTRIK GRUBU': 'DMC',
        'DMC ELLIPSE ELEKTRIK GRUBU': 'DMC',
        'DMC GARDEN ELEKTRIK GRUBU': 'DMC',
        'DMC ELEKTRIK GRUBU': 'DMC',
        'DMC ELLIPSE GARDEN ELEKTRİK': 'DMC',
        'DMC ELLIPSE ELEKTRİK': 'DMC',
        'DMC GARDEN ELEKTRİK': 'DMC',
        'DMC': 'DMC',
        'KÖKSARAY': 'KÖKSARAY',
        'KOK SARAY': 'KÖKSARAY',
        'OHP': 'OHP',  # Hala normalize ediliyor, sadece opsiyonel
        'TYM': 'TYM',
        'YHP': 'YHP',
        'MMP': 'MMP',
        'RMC': 'RMC',
        'MOS': 'MOS',
        # YENİ DATA CENTR MAPPING'LER EKLENDİ
        'DATA CENTR': 'DATA CENTR',
        'DATA CENTER': 'DATA CENTR',
        'DATA CENTRE': 'DATA CENTR',
        'DATACENTR': 'DATA CENTR',
        'DATACENTER': 'DATA CENTR',
        'DATACENTRE': 'DATA CENTR',
        'DATA_CENTR': 'DATA CENTR',
        'DATA_CENTER': 'DATA CENTR',
        'DATA-CENTER': 'DATA CENTR',
        'DATA-CENTRE': 'DATA CENTR',
        'DATA CENTR ŞANTİYESİ': 'DATA CENTR',
        'DATA CENTER ŞANTİYESİ': 'DATA CENTR'
    }
    
    return mappings.get(site_name, site_name)

# YENİ ŞANTİYE PARSING FONKSİYONU - "TÜMÜ" FİLTRELENDİ
def parse_santiye_list(proje_string):
    """
    YENİ ŞANTİYE PARSING KURALLARI:
    - 'SKP (DAHO) / DMC' → ['SKP', 'DMC']
    - '/' , ',' , '-' , '|' ile ayır
    - Parantez içlerini temizle
    - 'Tümü' → tüm şantiyeler (özel işlem)
    - 'Belli değil' → atla
    - Şantiye isimlerini normalize et
    """
    if not proje_string or pd.isna(proje_string):
        return []
    
    proje_string = str(proje_string).strip()
    
    # Özel durumlar
    if proje_string.upper() == 'TÜMÜ':
        return ['TÜMÜ']
    if proje_string.upper() in ['BELLİ DEĞİL', 'BELİRSİZ', '']:
        return []
    
    # Parantez içlerini temizle: 'SKP (DAHO)' → 'SKP'
    proje_string = re.sub(r'\([^)]*\)', '', proje_string)
    
    # Birden fazla ayırıcı ile böl
    parts = re.split(r'[/,\-\|]', proje_string)
    
    # Temizle, filtrele ve normalize et
    santiyeler = []
    for part in parts:
        part_clean = part.strip()
        if part_clean and part_clean.upper() not in ['BELLİ DEĞİL', 'BELİRSİZ']:
            # Şantiye ismini normalize et
            normalized_site = normalize_site_name(part_clean)
            santiyeler.append(normalized_site)
    
    return santiyeler

# Doğrulama ile gelişmiş Excel yükleme - "TÜMÜ" FİLTRELENDİ
def load_excel_intelligent():
    """Kapsamlı doğrulama ile akıllı Excel dosyası yükleme"""
    global df, rapor_sorumlulari, id_to_name, id_to_projects, id_to_status, id_to_rol
    global ADMINS, IZLEYICILER, TUM_KULLANICILAR, santiye_sorumlulari, santiye_rapor_durumu
    global last_excel_update, excel_file_hash, excel_last_modified
    
    try:
        # Önbellek için dosya hash ve değişiklik zamanını kontrol et
        current_hash = get_file_hash(USERS_FILE)
        current_mtime = os.path.getmtime(USERS_FILE) if os.path.exists(USERS_FILE) else 0
        
        if (current_hash == excel_file_hash and 
            current_mtime == excel_last_modified and 
            df is not None):
            logging.info("✅ Excel önbellekte - Yeniden yüklemeye gerek yok")
            return
        
        # Doğrulama için gerekli kolonları tanımla
        required_columns = ["Telegram ID", "Kullanici Adi Soyadi", "Rol", "Botdaki Statusu", "Proje / Şantiye"]
        
        try:
            # Esnek Excel okuma
            df = safe_read_excel(USERS_FILE, required_columns)
            logging.info("✅ Excel dosyası başarıyla yüklendi")
            
            excel_file_hash = current_hash
            excel_last_modified = current_mtime
            
        except (FileNotFoundError, ValueError) as e:
            logging.error(f"❌ Excel okuma hatası: {e}. Fallback kullanıcı listesi kullanılıyor.")
            df = pd.DataFrame(FALLBACK_USERS)
    
    except Exception as e:
        logging.error(f"❌ Excel yükleme hatası: {e}. Fallback kullanıcı listesi kullanılıyor.")
        df = pd.DataFrame(FALLBACK_USERS)
    
    # ŞANTİYE BAZLI SİSTEM: Güvenli tuple işleme ile Excel verilerini işle
    temp_rapor_sorumlulari = []
    temp_id_to_name = {}
    temp_id_to_projects = {}
    temp_id_to_status = {}
    temp_id_to_rol = {}
    temp_admins = []
    temp_izleyiciler = []
    temp_tum_kullanicilar = []
    temp_santiye_sorumlulari = {}
    processed_names = set()

    for _, r in df.iterrows():
        # Telegram ID parsing
        tid = _to_int_or_none(r.get("Telegram ID"))
        fullname = str(r.get("Kullanici Adi Soyadi") or "").strip()
        
        aktif_pasif = str(r.get("Aktif / Pasif") or "E").strip().upper()
        status = str(r.get("Botdaki Statusu") or "Aktif").strip()
        rol = str(r.get("Rol") or "KULLANICI").strip().upper()

        # Sadece aktif kullanıcıları işle
        if not fullname or aktif_pasif != "E":
            continue

        if tid and fullname:
            # Bilinen ID düzeltmelerini işle
            if tid == 10001573260:
                tid = 1000157326
            if tid == 7097468990:
                tid = 709746899
                
            tid = int(tid)
            temp_id_to_name[tid] = fullname
            temp_id_to_status[tid] = status
            temp_id_to_rol[tid] = rol
            
            temp_tum_kullanicilar.append(tid)
            
            if rol in ["ADMIN", "SÜPER ADMIN", "SUPER ADMIN"]:
                temp_admins.append(tid)
            
            if rol == "İZLEYİCİ":
                temp_izleyiciler.append(tid)
            
            # ŞANTİYE PARSING - "TÜMÜ" FİLTRELENDİ
            raw_projects = str(r.get("Proje / Şantiye") or "")
            projects = parse_santiye_list(raw_projects)
            
            # "TÜMÜ" şantiyesini filtrele - şantiye listesinde görünmesin
            projects = [proje for proje in projects if proje != "TÜMÜ"]
            temp_id_to_projects[tid] = projects
            
            # Şantiye sorumlularını güncelle - "TÜMÜ" hariç
            for proje in projects:
                if proje and proje != "TÜMÜ":  # "TÜMÜ" şantiyesini ekleme
                    if proje not in temp_santiye_sorumlulari:
                        temp_santiye_sorumlulari[proje] = []
                    if tid not in temp_santiye_sorumlulari[proje]:
                        temp_santiye_sorumlulari[proje].append(tid)
            
            # Tüm aktif kullanıcılar rapor sorumlusu listesene eklenir
            if tid and fullname:
                temp_rapor_sorumlulari.append(tid)
                processed_names.add(fullname)

    # Global değişkenleri güncelle
    rapor_sorumlulari = temp_rapor_sorumlulari
    id_to_name = temp_id_to_name
    id_to_projects = temp_id_to_projects
    id_to_status = temp_id_to_status
    id_to_rol = temp_id_to_rol
    ADMINS = temp_admins
    IZLEYICILER = temp_izleyiciler
    TUM_KULLANICILAR = temp_tum_kullanicilar
    santiye_sorumlulari = temp_santiye_sorumlulari
    santiye_rapor_durumu = {}
    
    # Super admin'in admin listesinde olduğundan emin ol
    if SUPER_ADMIN_ID not in ADMINS:
        ADMINS.append(SUPER_ADMIN_ID)
    
    last_excel_update = os.path.getmtime(USERS_FILE) if os.path.exists(USERS_FILE) else 0
    
    # "TÜMÜ" şantiyesi olup olmadığını kontrol et
    tumu_sayisi = sum(1 for projects in temp_id_to_projects.values() if "TÜMÜ" in projects)
    logging.info(f"✅ SİSTEM YÜKLENDİ: {len(rapor_sorumlulari)} aktif kullanıcı, {len(ADMINS)} admin, {len(IZLEYICILER)} izleyici, {len(TUM_KULLANICILAR)} toplam kullanıcı, {len(santiye_sorumlulari)} şantiye, {tumu_sayisi} kullanıcıda 'TÜMÜ' şantiyesi (filtrelendi)")

# Excel yüklemeyi başlat
load_excel_intelligent()

# Google Cloud Storage fonksiyonları
import google.cloud.storage
from google.oauth2 import service_account

def create_google_client():
    try:
        google_key_base64 = os.getenv("GOOGLE_KEY_BASE64")
        if not google_key_base64:
            logging.warning("⚠️ GOOGLE_KEY_BASE64 bulunamadı")
            return None
            
        key_json = base64.b64decode(google_key_base64).decode('utf-8')
        credentials_info = json.loads(key_json)
        
        credentials = service_account.Credentials.from_service_account_info(credentials_info)
        storage_client = google.cloud.storage.Client(
            credentials=credentials,
            project=os.getenv("GOOGLE_PROJECT_ID")
        )
        
        logging.info("✅ Google Cloud Storage client başarıyla oluşturuldu")
        return storage_client
    except Exception as e:
        logging.error(f"❌ Google Cloud Storage client oluşturma hatası: {e}")
        return None

def upload_backup_to_google(filename, remote_path=None):
    try:
        client = create_google_client()
        if not client:
            return False
            
        bucket_name = os.getenv("GOOGLE_BUCKET_NAME")
        if not bucket_name:
            logging.error("❌ GOOGLE_BUCKET_NAME bulunamadı")
            return False
            
        bucket = client.bucket(bucket_name)
        
        if remote_path is None:
            remote_path = f"backups/{os.path.basename(filename)}"
            
        blob = bucket.blob(remote_path)
        
        with open(filename, 'rb') as f:
            blob.upload_from_file(f)
            
        logging.info(f"✅ Dosya Google Cloud Storage'a yüklendi: {remote_path}")
        return True
        
    except Exception as e:
        logging.error(f"❌ Google Cloud Storage yükleme hatası: {e}")
        return False

def download_last_backup(remote_path, local_filename):
    try:
        client = create_google_client()
        if not client:
            return False
            
        bucket_name = os.getenv("GOOGLE_BUCKET_NAME")
        if not bucket_name:
            return False
            
        bucket = client.bucket(bucket_name)
        blob = bucket.blob(remote_path)
        
        blob.download_to_filename(local_filename)
        logging.info(f"✅ Dosya Google Cloud Storage'dan indirildi: {remote_path}")
        return True
        
    except Exception as e:
        logging.error(f"❌ Google Cloud Storage indirme hatası: {e}")
        return False

def list_backups(prefix="backups/"):
    try:
        client = create_google_client()
        if not client:
            return []
            
        bucket_name = os.getenv("GOOGLE_BUCKET_NAME")
        if not bucket_name:
            return []
            
        bucket = client.bucket(bucket_name)
        blobs = bucket.list_blobs(prefix=prefix)
        
        backup_list = []
        for blob in blobs:
            backup_list.append({
                'name': blob.name,
                'size': blob.size,
                'updated': blob.updated
            })
            
        return sorted(backup_list, key=lambda x: x['updated'], reverse=True)
        
    except Exception as e:
        logging.error(f"❌ Google Cloud Storage liste hatası: {e}")
        return []

async def async_upload_to_google(filename, remote_path=None):
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, upload_backup_to_google, filename, remote_path)

async def yedekleme_gorevi(context: ContextTypes.DEFAULT_TYPE):
    try:
        logging.info("💾 Yedekleme işlemi başlatılıyor...")
        
        success_count = 0
        total_count = 0
        
        backup_files = [
            ("Kullanicilar.xlsx", "backups/Kullanicilar.xlsx"),
            ("bot.log", "backups/bot.log")
        ]
        
        for local_file, remote_path in backup_files:
            if os.path.exists(local_file):
                total_count += 1
                if await async_upload_to_google(local_file, remote_path):
                    success_count += 1
            else:
                logging.warning(f"⚠️ Yedeklenecek dosya bulunamadı: {local_file}")
        
        status_msg = f"💾 Gece Yedekleme Raporu\n\n"
        status_msg += f"📅 Tarih: {dt.datetime.now(TZ).strftime('%d.%m.%Y %H:%M')}\n"
        status_msg += f"📁 Dosya: {success_count}/{total_count} başarılı\n"
        
        if success_count == total_count:
            status_msg += "🎉 Tüm yedeklemeler başarılı!"
            logging.info("💾 Gece yedeklemesi tamamlandı: Tüm dosyalar başarıyla yedeklendi")
        else:
            status_msg += f"⚠️ {total_count - success_count} dosya yedeklenemedi"
            logging.warning(f"💾 Gece yedeklemesi kısmen başarılı: {success_count}/{total_count}")
        
        if success_count > 0:
            for admin_id in ADMINS:
                try:
                    await context.bot.send_message(
                        chat_id=admin_id,
                        text=status_msg
                    )
                    logging.info(f"💾 Yedekleme raporu {admin_id} adminine gönderildi")
                except Exception as e:
                    logging.error(f"Yedekleme raporu {admin_id} adminine gönderilemedi: {e}")
                
    except Exception as e:
        logging.error(f"💾 Yedekleme hatası: {e}")

def yedekle_postgres():
    try:
        timestamp = dt.datetime.now(TZ).strftime("%Y-%m-%d_%H-%M")
        dump_file = f"postgres_backup_{timestamp}.dump"
        dump_path = f"/tmp/{dump_file}"

        db_url = os.getenv("DATABASE_URL")
        if not db_url:
            logging.error("❌ DATABASE_URL bulunamadı")
            return False

        try:
            with open(dump_path, 'wb') as f:
                result = subprocess.run(
                    ['pg_dump', '-Fc', db_url],
                    stdout=f,
                    stderr=subprocess.PIPE,
                    check=False
                )
            
            if result.returncode != 0:
                logging.error(f"❌ pg_dump başarısız (code {result.returncode}): {result.stderr.decode()}")
                if os.path.exists(dump_path):
                    os.unlink(dump_path)
                return False
                
        except Exception as e:
            logging.error(f"❌ pg_dump çalıştırma hatası: {e}")
            if os.path.exists(dump_path):
                os.unlink(dump_path)
            return False

        gcs_path = f"backups/sql/{dump_file}"
        success = upload_backup_to_google(dump_path, gcs_path)
        
        if os.path.exists(dump_path):
            os.unlink(dump_path)

        if success:
            logging.info(f"💾 PostgreSQL yedeği alındı ve GCS'ye yüklendi: {dump_file}")
            return True
        else:
            logging.error("❌ PostgreSQL yedeği GCS'ye yüklenemedi")
            return False

    except Exception as e:
        logging.error(f"❌ PostgreSQL yedeği sırasında hata: {e}")
        if 'dump_path' in locals() and os.path.exists(dump_path):
            os.unlink(dump_path)
        return False

async def yedekle_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await super_admin_kontrol(update, context):
        return
    
    await update.message.reply_text("💾 Yedekleme işlemi başlatılıyor...")
    
    try:
        success_count = 0
        backup_files = [
            ("Kullanicilar.xlsx", "backups/Kullanicilar.xlsx"),
            ("bot.log", "backups/bot.log")
        ]
        
        for local_file, remote_path in backup_files:
            if os.path.exists(local_file):
                if await async_upload_to_google(local_file, remote_path):
                    success_count += 1
        
        if success_count == len(backup_files):
            await update.message.reply_text("✅ Tüm yedeklemeler başarıyla tamamlandı!")
        else:
            await update.message.reply_text(f"⚠️ Yedekleme kısmen başarılı: {success_count}/{len(backup_files)} dosya")
            
    except Exception as e:
        await update.message.reply_text(f"❌ Yedekleme hatası: {e}")

def is_media_message(message) -> bool:
    if message.photo:
        return True
    if message.video:
        return True
    if message.audio:
        return True
    if message.voice:
        return True
    if message.animation:
        return True
    if message.video_note:
        return True
    if message.document:
        return True

    if (message.caption and not message.text):
        return True

    return False

# YENİ SİSTEM_PROMPT - "ÇALIŞMA YOK" DÜZELTMESİ GELİŞTİRİLDİ
SYSTEM_PROMPT = """
Sen bir "Rapor Analiz Asistanısın". Görevin, kullanıcıların Telegram üzerinden gönderdiği serbest formatlı günlük personel raporlarını SABİT BİR JSON formatına dönüştürmektir.

ÖNEMLİ KURALLAR:

1. SABİT JSON FORMATI: Her zaman aşağıdaki sabit JSON formatını kullan:

[
  {
    "date": "YYYY-AA-GG",
    "site": "ŞANTİYE_ADI",
    "staff": sayı,
    "calisan": sayı,
    "mobilizasyon": sayı,
    "ambarci": sayı,
    "izinli": sayı,
    "dis_gorev": [
      {"gorev_yeri": "YER_ADI", "sayi": sayı}
    ],
    "dis_gorev_toplam": sayı,
    "genel_toplam": sayı
  }
]

2. ÖNCELİK KURALI - ÇOK ÖNEMLİ:
   - ÖNCE mesajda "GENEL ÖZET" bölümü ara (tüm varyasyonlar: "📝 GENEL ÖZET:", "GENEL ÖZET:", "GENEL ÖZET", "📝 Genel Özet:", "Genel Özet:", "Genel özet", "genel özet", "📝 GENEL OZET:", "GENEL OZET:", "GENEL OZET", "📝 Genel Ozet:", "Genel Ozet:", "Genel ozet", "genel ozet", "📝 genel özet:", "📝 genel ozet:", "📝 Genel özet:", "📝 Genel ozet:", "📝 GENEL ÖZET", "📝 GENEL OZET", "(📝) GENEL ÖZET:", "(📝) Genel Özet:")
   - Eğer GENEL ÖZET bölümü varsa: SADECE GENEL ÖZET bölümündeki sayıları kullan! Detaylı maddeleri TAMAMEN YOK SAY ve parse etme!
   - GENEL ÖZET yoksa veya eksikse, o zaman detaylı maddelerden say

3. YENİ TANIMLAR - KRİTİK:
   - "TAŞERON", "taşeron" → "calisan" kategorisine DAHİL
   - "Yerel Ekipbaşı" → "staff" kategorisine DAHİL
   - "Toplam staff", "staff", "Staff" → "staff"
   - "Toplam imalat", "imalat", "İmalat", "çalışan", "Çalışan" → "calisan"
   - "Toplam mobilizasyon", "mobilizasyon", "Mobilizasyon" → "mobilizasyon"
   - "Toplam ambar", "ambar", "ambarcı", "Ambarcı" → "ambarci"
   - "İzinli", "izinli", "Hasta" → "izinli"
   - "Şantiye dışı görev", "Şantiye dışı", "dış görev", "Dış görev", "Başka şantiye", "Buxoro'ya gitti", "Buxoro", "Başka yere görev" → "dis_gorev"

4. ÇALIŞMA YOK/İŞ YOK RAPORLARI - YENİ KURAL:
   - Mesajda "çalışma yok", "iş yok", "faaliyet yok", "günlük çalışma yok", "bugün çalışma yapılmadı", "aktivite yok", "işçilik yok", "raporlanacak çalışma yok", "çalışma gerçekleştirilmedi", "saha kapalı / faaliyet yapılmadı", "operasyon yok", "gün boş", "bugün iş yok", "çalışma mevcut değil", "planlanan çalışma yok", "saha çalışması yapılmadı", "işlem yapılmamıştır", "görev yok", "aktif iş yok", "rapor yok çalışma yok", "calisma yok", "calışma yok", "çalıșma yok", "çalısma yok", "çalıma yok", "calışma yok", "çalşma yok", "çalışma yoktur", "calişma yok", "çelışma yok", "çalışmayok", "calismayok", "çalşmy yok", "çalılşma yok", "çalışa yok", "çalişma yok", "calıma yok", "çalısma yk", "cal yok", "ç yok", "calyok", "çalışmyok", "çalışm yok", "iş yok", "is yok", "yok çalışma", "bugün yok", "çalışma yk", "çalış. yok", "ç. yok", "işlm yok", "aktif yok" gibi ifadeler varsa:
   - TÜM personel kategorilerini (staff, calisan, mobilizasyon, ambarci, izinli) 0 olarak ayarla!
   - genel_toplam = 0 olarak ayarla!
   - "izinli" kategorisini de 0 olarak ayarla!
   - Çalışma yok raporu, personelsiz şantiye durumu için kullanılır.

5. ÇİFT SAYMA KORUMASI:
   - Asla aynı mesajdan hem GENEL ÖZET hem detay sayma!
   - GENEL ÖZET bulduğunda detayları GÖRMEZDEN GEL!
   - ÖRNEK: Mesajda hem detaylı işler hem de "GENEL ÖZET" varsa, SADECE GENEL ÖZET kullan!

6. YEREL EKİPBAŞI KURALI:
   - "Yerel Ekipbaşı" personel DAİMA "staff" kategorisine DAHİLDİR
   - Raporda "Yerel Ekipbaşı: 5 kişi" görürsen → "staff"a EKLE!
   - ÖRNEK: "Staff: 8, Yerel Ekipbaşı: 5" → staff = 13
   - Yerel Ekipbaşı'yı asla ayrı bir kategori olarak sayma!

7. DIŞ GÖREV KURALI:
   - "dis_gorev_toplam" asla "genel_toplam"a DAHİL EDİLMEZ!
   - Genel toplam = staff + calisan + mobilizasyon + ambarci + izinli
   - Dış görevler sadece bilgi amaçlı "dis_gorev" listesinde gösterilir
   - ÖRNEK: Staff:2 + Çalışan:3 = 5, Dış görev:5 → genel_toplam = 5 (10 değil!)

8. GENEL TOPLAM DOĞRULAMA:
   - Kullanıcı "Genel toplam: X" yazsa bile SEN MATEMATİK KONTROLÜ YAP!
   - Eğer staff+calisan+mobilizasyon+ambarci+izinli ≠ genel_toplam ise
   - O ZAMAN kendi hesapladığın doğru toplamı kullan!
   - ÖRNEK: "Genel toplam: 10" ama staff:2 + çalışan:3 = 5 ise → genel_toplam = 5 kullan!

9. TARİH ALGILAMA:
   - Format: YYYY-AA-GG
   - Örnek: "13.11.2026" → "2026-11-13"
   - Tarih yoksa bugünün tarihini kullan

10. ŞANTİYE NORMALİZASYONU:
    - LOT13, LOT71, SKP, BWC, Piramit, STADYUM, DMC, YHP, TYM, MMP, RMC, PİRAMİT, MOS, DATA CENTR
    - "Lot 13", "lot13", "LOT-13" → "LOT13"
    - "SKP Daho", "📍 SKP Elektrik Grubu", " SKP Elektrik Grubu", "SKP Elektrik Grubu", "SKP", → "SKP"
    - "Piramit Tower", "PİRAMİT TOWER", "PRAMİT", "PIRAMIT", "PİRAMİD", "PIRAMID", "PYRAMIT", "PYRAMID", "PİRAMİT", "PIRAMIT TOWER" → "PİRAMİT"
    - "DMC Ellipse Garden", "DMC ELLIPSE GARDEN", "DMC Ellipse", "DMC Garden", "DMC Ellipse Garden Elektrik Grubu", "DMC ELEKTRIK GRUBU" → "DMC"
    - "YHP" → "YHP"
    - "TYM" → "TYM"
    - "MMP" → "MMP"
    - "RMC" → "RMC"
    - "KOK SARAY" → "KÖKSARAY"
    - "MOS" → "MOS"
    - "Data Center", "Data Centre", "DATA-CENTER", "DATA CENTER ŞANTİYESİ" → "DATA CENTR"

11. PERSONEL KATEGORİLERİ:
    - staff: mühendis, tekniker, formen, ekipbaşı, şef, Türk mühendis, Türk formen, Yerel formen, Yerel Ekipbaşı, Yerel ekipbaşı, Toplam staff, Staff
    - calisan: usta, işçi, yardımcı, operatör, imalat, çalışan, worker, TAŞERON, taşeron, Toplam imalat, İmalat
    - ambarci: ambarcı, depo sorumlusu, malzemeici, ambar, Toplam ambar, Ambarcı
    - mobilizasyon: genel mobilizasyon, saha kontrol, nöbetçi, mobilizasyon takimi, Toplam mobilizasyon, Mobilizasyon
    - izinli: izinli, iş yok, gelmedi, izindeyim, hasta, raporlu, hastalık izni, sıhhat izni, İzinli, Hasta
    - dis_gorev: başka şantiye görev, dış görev, Lot 71 dış görev, Fap dış görev, Şantiye dışı görev, Şantiye dışı, dış görev, Dış görev, Başka şantiye, Buxoro'ya gitti, Buxoro, Başka yere görev, yurt dışı görev, Dış görev, Şantiye dışı

12. HESAPLAMALAR:
    genel_toplam = staff + calisan + mobilizasyon + ambarci + izinli
    dis_gorev_toplam = tüm dış görevlerin toplamı (genel_toplam'a EKLENMEZ!)

13. DİKKAT EDİLECEK NOKTALAR:
    - "Çalışan: 10" → calisan: 10
    - "İzinli: 1" → izinli: 1
    - "Ambarcı: 2" → ambarci: 2
    - "Toplam staff: 1" → staff: 1
    - "Toplam mobilizasyon: 2" → mobilizasyon: 2
    - "Yerel Ekipbaşı: 5 kişi" → staff: 5 (staff'a EKLE!)
    - "TAŞERON: 10 kişi" → calisan: 10
    - "Şantiye dışı görev: 2 kişi" → dis_gorev: [{"gorev_yeri": "ŞANTİYE_DIŞI", "sayi": 2}], dis_gorev_toplam: 2
    - "Buxoro'ya gitti: 2 kişi" → dis_gorev: [{"gorev_yeri": "BUXORO", "sayi": 2}], dis_gorev_toplam: 2
    - "Lot 71 dış görev 8" → dis_gorev: [{"gorev_yeri": "LOT71", "sayi": 8}], dis_gorev_toplam: 8
    - "Genel toplam: 10 kişi" → genel_toplam: 10 (ama MATEMATİK KONTROLÜ yap!)
    - "Çalışma yok", "İş yok", "Hiç personel yok" → staff:0, calisan:0, mobilizasyon:0, ambarci:0, izinli:0, genel_toplam:0

14. ÖZEL DURUM - DMC ÖRNEĞİ:
    Aşağıdaki DMC raporunu analiz ederken:
    • Yerel ekipbaşı: 1 kişi
    • Buxoro'ya gitti: 2 kişi
    ...diğer detaylar...
    📝 GENEL ÖZET:
    • Toplam staff: 1 kişi
    • Toplam imalat: 20 kişi  
    • Toplam mobilizasyon: 2 kişi
    • Şantiye dışı görev: 2 kişi
    • Genel toplam: 25 kişi
    
    ÇÖZÜM: 
    - SADECE GENEL ÖZET kullan!
    - staff: 1 (Yerel ekipbaşı dahil)
    - calisan: 20
    - mobilizasyon: 2  
    - dis_gorev_toplam: 2
    - genel_toplam: 23 (1 + 20 + 2 = 23, kullanıcının 25'i yanlış!)

15. ÖZEL DURUM - ÇALIŞMA YOK RAPORU:
    "06.12.2026 LOT13 çalışma yok" veya "LOT13 bugün iş yok, personel yok"
    ÇÖZÜM:
    - staff: 0
    - calisan: 0
    - mobilizasyon: 0
    - ambarci: 0
    - izinli: 0
    - genel_toplam: 0

16. ÖRNEK ÇIKTI FORMATI:
[
  {
    "date": "2026-11-13",
    "site": "LOT13",
    "staff": 13,
    "calisan": 5,
    "mobilizasyon": 2,
    "ambarci": 1,
    "izinli": 1,
    "dis_gorev": [
      {"gorev_yeri": "LOT71", "sayi": 3},
      {"gorev_yeri": "FAP", "sayi": 2}
    ],
    "dis_gorev_toplam": 5,
    "genel_toplam": 22
  }
]

DİKKAT: 
- Sadece JSON döndür, açıklama yapma!
- Tüm sayıları integer olarak döndür
- Eksik alanları 0 olarak döndür
- dis_gorev her zaman bir liste olmalı, boşsa []
- Her zaman bu sabit JSON formatını kullan!
- GENEL ÖZET BÖLÜMÜ VARSA DETAYLARI YOK SAY!
- genel_toplam = staff + calisan + mobilizasyon + ambarci + izinli (dis_gorev_toplam dahil DEĞİL!)
- Yerel Ekipbaşı her zaman staff kategorisine dahil edilir!
- TAŞERON her zaman calisan kategorisine dahil edilir!
- Kullanıcının genel toplamını KÖRÜ KÖRÜNE KABUL ETME, matematik kontrolü yap!
- ÇALIŞMA YOK raporlarında tüm personel kategorilerini 0 yap!
"""

# Gelişmiş tarih parser fonksiyonları
def enhanced_date_parser(text):
    """Gelişmiş tarih parser - geçici implementasyon"""
    try:
        # Basit tarih parsing implementasyonu
        patterns = [
            r'(\d{1,2})[\.\/\-](\d{1,2})[\.\/\-](\d{4})',
            r'(\d{1,2})[\.\/\-](\d{1,2})[\.\/\-](\d{2})',
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, text)
            for match in matches:
                try:
                    if len(match) == 3:
                        day, month, year = int(match[0]), int(match[1]), int(match[2])
                        if len(str(year)) == 2:
                            year += 2000
                        return dt.datetime(year, month, day).date()
                except ValueError:
                    continue
        return None
    except Exception:
        return None

def get_santiye_sorumlusu(santiye_adi):
    """
    Şantiye adına göre sorumlu kişiyi bul
    """
    try:
        santiye_adi = normalize_site_name(santiye_adi)
        
        # Özel durumlar
        if santiye_adi == "BELİRSİZ":
            return None
            
        # Şantiye sorumluları listesinde ara
        if santiye_adi in santiye_sorumlulari:
            sorumlular = santiye_sorumlulari[santiye_adi]
            if sorumlular:
                # Aktif ilk sorumluyu döndür
                for sorumlu_id in sorumlular:
                    if sorumlu_id in rapor_sorumlulari:
                        return sorumlu_id
                # Eğer hiçbiri aktif değilse ilkini döndür
                return sorumlular[0]
        
        logging.warning(f"⚠️ Şantiye sorumlusu bulunamadı: {santiye_adi}")
        return None
        
    except Exception as e:
        logging.error(f"❌ Şantiye sorumlusu bulma hatası: {e}")
        return None


def extract_max_number(text, patterns):
    """Pattern'lere göre maksimum sayıyı çıkar"""
    max_num = 0
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for match in matches:
            try:
                num = int(match)
                max_num = max(max_num, num)
            except ValueError:
                continue
    return max_num

# Basitleştirilmiş USER_PROMPT_TEMPLATE
USER_PROMPT_TEMPLATE = "<<<RAW_MESSAGE>>>"

client = OpenAI(api_key=OPENAI_API_KEY)

def gpt_analyze(system_prompt, user_prompt):
    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0,
            max_tokens=4000
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        logging.error(f"GPT hatası: {e}")
        return ""

# Gelişmiş GPT analizi ile giriş doğrulama
def gpt_analyze_enhanced(system_prompt, user_prompt):
    """Gelişmiş hata yönetimi ile GPT ile metin analizi"""
    is_valid, cleaned_prompt = validate_user_input(user_prompt, 10000)
    if not is_valid:
        logging.error("GPT'ye geçersiz kullanıcı girişi sağlandı")
        return ""
    
    try:
        client = OpenAI(api_key=OPENAI_API_KEY, timeout=30.0)
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": cleaned_prompt}
            ],
            temperature=0,
            max_tokens=2000,
            timeout=30.0
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        logging.error(f"GPT analiz hatası: {e}")
        return ""

# ÇALIŞMA YOK kontrolü için DÜZELTİLMİŞ fonksiyon
def is_calisma_yok_raporu(text):
    """Metnin "çalışma yok" raporu olup olmadığını kontrol et"""
    if not text:
        return False
    
    # DÜZELTİLDİ: Sadece açıkça "çalışma yok" ifadelerini kontrol et
    calisma_yok_kelimeler = [
        'çalışma yok', 'iş yok', 'faaliyet yok', 'günlük çalışma yok', 
        'bugün çalışma yapılmadı', 'aktivite yok', 'işçilik yok', 
        'raporlanacak çalışma yok', 'çalışma gerçekleştirilmedi', 
        'saha kapalı / faaliyet yapılmadı', 'operasyon yok', 'gün boş', 
        'bugün iş yok', 'çalışma mevcut değil', 'planlanan çalışma yok', 
        'saha çalışması yapılmadı', 'işlem yapılmamıştır', 'görev yok', 
        'aktif iş yok', 'rapor yok çalışma yok', 'calisma yok', 
        'calışma yok', 'çalıșma yok', 'çalısma yok', 'çalıma yok', 
        'calışma yok', 'çalşma yok', 'çalışma yoktur', 'calişma yok', 
        'çelışma yok', 'çalışmayok', 'calismayok', 'çalşmy yok', 
        'çalılşma yok', 'çalışa yok', 'çalişma yok', 'calıma yok', 
        'çalısma yk', 'cal yok', 'ç yok', 'calyok', 'çalışmyok', 
        'çalışm yok', 'iş yok', 'is yok', 'yok çalışma', 'bugün yok', 
        'çalışma yk', 'çalış. yok', 'ç. yok', 'işlm yok', 'aktif yok'
    ]
    
    text_lower = text.lower()
    
    # Tüm kelimeleri kontrol et
    for kelime in calisma_yok_kelimeler:
        if kelime in text_lower:
            # DÜZELTİLDİ: "iş yok" veya "çalışma yok" ifadelerini sadece bağımsız olarak kontrol et
            # Eğer "iş yok" içeriyorsa, "iş yok değil" gibi ifadeleri elemek için regex kullan
            if kelime in ['iş yok', 'is yok', 'çalışma yok', 'calisma yok']:
                # Regex ile tam eşleşme kontrolü
                pattern = r'\b' + re.escape(kelime) + r'\b'
                if re.search(pattern, text_lower):
                    logging.info(f"📝 'Çalışma yok' raporu tespit edildi: '{kelime}'")
                    return True
            else:
                logging.info(f"📝 'Çalışma yok' raporu tespit edildi: '{kelime}'")
                return True
    
    return False

# Doğrulama ile gelişmiş process_incoming_message - DÜZELTİLDİ
def process_incoming_message(raw_text: str, is_group: bool = False):
    """Kapsamlı doğrulama ile gelen mesajı işle - DÜZELTİLDİ: Çalışma yok raporları için geliştirildi"""
    is_valid, cleaned_text = validate_user_input(raw_text)
    if not is_valid:
        return [] if is_group else {"error": "geçersiz_giriş"}
    
    today = dt.date.today()
    max_retries = 2
    retry_delay = 1
    
    # ÖNCE "ÇALIŞMA YOK" KONTROLÜ - DÜZELTİLDİ
    is_calisma_yok = is_calisma_yok_raporu(cleaned_text)
    
    for attempt in range(max_retries):
        try:
            user_prompt = cleaned_text
            content = gpt_analyze_enhanced(SYSTEM_PROMPT, user_prompt)
            
            if not content:
                if attempt < max_retries - 1:
                    time_module.sleep(retry_delay)
                    continue
                return [] if is_group else {"dm_info": "no_report_detected"}
            
            data = safe_json_loads(content)
            if data is None:
                if attempt < max_retries - 1:
                    time_module.sleep(retry_delay)
                    continue
                return [] if is_group else {"dm_info": "no_report_detected"}
            
            if isinstance(data, dict):
                data = [data]
            
            if not isinstance(data, list):
                if attempt < max_retries - 1:
                    time_module.sleep(retry_delay)
                    continue
                return [] if is_group else {"dm_info": "no_report_detected"}
            
            filtered_reports = []
            for report in data:
                if not isinstance(report, dict):
                    continue
                    
                date_str = report.get('date')
                if date_str:
                    try:
                        report_date = dt.datetime.strptime(date_str, '%Y-%m-%d').date()
                        if report_date > today:
                            report['date'] = today.strftime('%Y-%m-%d')
                    except ValueError:
                        report['date'] = today.strftime('%Y-%m-%d')
                else:
                    report['date'] = today.strftime('%Y-%m-%d')
                
                site = report.get('site', 'BELİRSİZ')
                # GPT'DEN GELEN ŞANTİYE İSMİNİ NORMALİZE ET - EKLENDİ
                report['site'] = normalize_site_name(site)
                
                # DÜZELTİLDİ: ÇALIŞMA YOK KONTROLÜ - SADECE AÇIKÇA "ÇALIŞMA YOK" İFADESİ GEÇİYORSA
                if is_calisma_yok:
                    logging.info("📝 'Çalışma yok' raporu tespit edildi - tüm personel kategorileri 0 olarak ayarlanıyor")
                    report['staff'] = 0
                    report['calisan'] = 0
                    report['mobilizasyon'] = 0
                    report['ambarci'] = 0
                    report['izinli'] = 0
                    report['genel_toplam'] = 0
                    # Dış görevler de olmamalı çünkü hiç personel yok
                    report['dis_gorev'] = []
                    report['dis_gorev_toplam'] = 0
                
                for key in ['staff', 'calisan', 'mobilizasyon', 'ambarci', 'izinli', 'dis_gorev_toplam', 'genel_toplam']:
                    value = report.get(key, 0)
                    if not isinstance(value, int):
                        try:
                            report[key] = int(value) if value else 0
                        except (ValueError, TypeError):
                            report[key] = 0
                
                # YENİ: GENEL TOPLAM DOĞRULAMA - Dış görevler dahil edilmez + Tanımlanmamış kategori kontrolü
                calculated_total = (
                    report.get('staff', 0) + 
                    report.get('calisan', 0) + 
                    report.get('mobilizasyon', 0) + 
                    report.get('ambarci', 0) + 
                    report.get('izinli', 0)
                )
                
                # TANIMSIZ KATEGORİ KONTROLÜ - YENİ EKLENDİ
                tanimli_kategoriler_toplami = calculated_total
                tanimsiz_kategori_var = False
                
                # GPT'nin ekstra kategoriler ekleyip eklemediğini kontrol et
                tum_anahtarlar = set(report.keys())
                tanimli_anahtarlar = {'date', 'site', 'staff', 'calisan', 'mobilizasyon', 'ambarci', 'izinli', 'dis_gorev', 'dis_gorev_toplam', 'genel_toplam'}
                ekstra_anahtarlar = tum_anahtarlar - tanimli_anahtarlar
                
                # Ekstra sayısal anahtarları kontrol et (operatör, usta başı vb.)
                for ekstra_anahtar in ekstra_anahtarlar:
                    deger = report.get(ekstra_anahtar, 0)
                    if isinstance(deger, (int, float)) and deger > 0:
                        tanimsiz_kategori_var = True
                        logging.warning(f"⚠️ Tanımlanmamış kategori tespit edildi: {ekstra_anahtar} = {deger}")
                        # Ekstra kategoriyi "calisan"a ekle (varsayılan)
                        report['calisan'] = report.get('calisan', 0) + int(deger)
                        calculated_total += int(deger)
                        logging.info(f"✅ Tanımlanmamış kategori '{ekstra_anahtar}' çalışanlara eklendi: +{deger}")
                
                # Eğer kullanıcının genel toplamı yanlışsa, doğru olanı kullan
                if report.get('genel_toplam', 0) != calculated_total:
                    logging.info(f"🔢 Genel toplam düzeltildi: {report.get('genel_toplam', 0)} → {calculated_total}")
                    if tanimsiz_kategori_var:
                        logging.info(f"📝 Sebep: Tanımlanmamış kategoriler çalışanlara eklendi")
                    report['genel_toplam'] = calculated_total
                
                # ÇALIŞMA YOK raporları da kaydedilmeli
                if report['genel_toplam'] > 0 or report['staff'] > 0 or is_calisma_yok:
                    filtered_reports.append(report)
            
            return filtered_reports
                
        except Exception as e:
            logging.error(f"Mesaj işleme hatası (deneme {attempt + 1}): {e}")
            if attempt < max_retries - 1:
                time_module.sleep(retry_delay)
    
    return [] if is_group else {"dm_info": "no_report_detected"}

# RAPOR KAYIT FONKSİYONU - ŞANTİYE BAZLI SİSTEM
async def raporu_gpt_formatinda_kaydet(user_id, kullanici_adi, orijinal_metin, gpt_rapor, msg, rapor_no=1):
    try:
        site = gpt_rapor.get('site', 'BELİRSİZ')
        date_str = gpt_rapor.get('date')
        
        # GPT'DEN GELEN ŞANTİYE İSMİNİ NORMALİZE ET - EKLENDİ
        site = normalize_site_name(site)
        
        rapor_tarihi = None
        if date_str:
            try:
                rapor_tarihi = dt.datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        if not rapor_tarihi:
            rapor_tarihi = parse_rapor_tarihi(orijinal_metin) or dt.datetime.now(TZ).date()
        
        santiye_sorumlusu_id = get_santiye_sorumlusu(site)
        
        kaydedilecek_user_id = santiye_sorumlusu_id if santiye_sorumlusu_id else user_id
        kaydedilecek_kullanici_adi = id_to_name.get(santiye_sorumlusu_id, kullanici_adi) if santiye_sorumlusu_id else kullanici_adi
        
        staff = gpt_rapor.get('staff', 0)
        calisan = gpt_rapor.get('calisan', 0)
        mobilizasyon = gpt_rapor.get('mobilizasyon', 0)
        ambarci = gpt_rapor.get('ambarci', 0)
        izinli = gpt_rapor.get('izinli', 0)
        dis_gorev_toplam = gpt_rapor.get('dis_gorev_toplam', 0)
        genel_toplam = gpt_rapor.get('genel_toplam', 0)
        
        # YENİ: ÇALIŞMA YOK raporları için kontrol
        if is_calisma_yok_raporu(orijinal_metin):
            logging.info(f"📝 'Çalışma yok' raporu kaydediliyor - Personel: 0")
            # Çalışma yok raporunda tüm personel 0 olmalı
            staff = 0
            calisan = 0
            mobilizasyon = 0
            ambarci = 0
            izinli = 0
            genel_toplam = 0
            # Dış görevler de 0 olmalı
            dis_gorev_toplam = 0
        
        # YENİ: GENEL TOPLAM DOĞRULAMA - Dış görevler dahil edilmez
        calculated_total = staff + calisan + mobilizasyon + ambarci + izinli
        if genel_toplam != calculated_total:
            logging.info(f"🔢 Rapor kaydında genel toplam düzeltildi: {genel_toplam} → {calculated_total}")
            genel_toplam = calculated_total
        
        project_name = site
        if not project_name or project_name == 'BELİRSİZ':
            if santiye_sorumlusu_id:
                user_projects = id_to_projects.get(santiye_sorumlusu_id, [])
            else:
                user_projects = id_to_projects.get(user_id, [])
                
            if user_projects:
                project_name = user_projects[0]
            else:
                project_name = 'BELİRSİZ'
        
        existing_report = await async_fetchone("""
            SELECT id FROM reports 
            WHERE project_name = %s AND report_date = %s
        """, (project_name, rapor_tarihi))
        
        has_existing_report = False
        if existing_report is not None:
            existing_id = safe_get_tuple_value(existing_report, 0)
            if existing_id is not None:
                has_existing_report = True
        
        if has_existing_report:
            logging.warning(f"⚠️ Zaten rapor var: {project_name} - {rapor_tarihi}")
            raise Exception(f"Bu şantiye için bugün zaten rapor gönderilmiş: {project_name}")
        
        # ÇALIŞMA YOK raporlarında rapor tipi "IZIN/ISYOK" olarak kaydedilir
        if izinli > 0 or is_calisma_yok_raporu(orijinal_metin):
            rapor_tipi = "IZIN/ISYOK"
        else:
            rapor_tipi = "RAPOR"
        
        work_description = f"Staff:{staff} Çalışan:{calisan} Mobilizasyon:{mobilizasyon} Ambarcı:{ambarci} İzinli:{izinli}"
        if dis_gorev_toplam > 0:
            work_description += f" DışGörevToplam:{dis_gorev_toplam}"
        
        if santiye_sorumlusu_id and santiye_sorumlusu_id != user_id:
            work_description += f" [Raporu ileten: {kullanici_adi}]"
        
        ai_analysis = {
            "yeni_sabit_format": gpt_rapor,
            "extraction_method": "yeni-sabit-json-format",
            "original_text": orijinal_metin[:500],
            "calculated_total": genel_toplam,
            "personel_dagilimi": {
                "staff": staff,
                "calisan": calisan,
                "mobilizasyon": mobilizasyon,
                "ambarci": ambarci,
                "izinli": izinli
            },
            "rapor_gonderen": {
                "user_id": user_id,
                "kullanici_adi": kullanici_adi
            },
            "santiye_sorumlusu": {
                "user_id": santiye_sorumlusu_id,
                "kullanici_adi": id_to_name.get(santiye_sorumlusu_id, "Belirsiz") if santiye_sorumlusu_id else "Belirsiz"
            } if santiye_sorumlusu_id else None,
            "is_calisma_yok": is_calisma_yok_raporu(orijinal_metin)  # YENİ: Çalışma yok bayrağı
        }
        
        await async_execute("""
            INSERT INTO reports 
            (user_id, project_name, report_date, report_type, person_count, work_description, 
             work_category, personnel_type, delivered_date, is_edited, ai_analysis)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            kaydedilecek_user_id, project_name, rapor_tarihi, rapor_tipi, genel_toplam, 
            work_description[:400], 'detaylı', 'imalat', dt.datetime.now(TZ).date(),
            False, json.dumps(ai_analysis, ensure_ascii=False)
        ))
        
        if santiye_sorumlusu_id and santiye_sorumlusu_id != user_id:
            logging.info(f"✅ ŞANTİYE BAZLI Rapor #{rapor_no} ŞANTİYE SORUMLUSU adına kaydedildi: {kaydedilecek_kullanici_adi} (Raporu ileten: {kullanici_adi}) - {project_name} - {rapor_tarihi}")
        else:
            logging.info(f"✅ ŞANTİYE BAZLI Rapor #{rapor_no} kaydedildi: {user_id} - {project_name} - {rapor_tarihi}")
            
        logging.info(f"📊 Personel Dağılımı: Staff:{staff}, Çalışan:{calisan}, Mobilizasyon:{mobilizasyon}, Ambarcı:{ambarci}, İzinli:{izinli}, DışGörevToplam:{dis_gorev_toplam}, GenelToplam:{genel_toplam}")
        
        maliyet_analiz.kayit_ekle('gpt')
            
    except Exception as e:
        logging.error(f"❌ Şantiye bazlı rapor kaydetme hatası: {e}")
        raise e

async def yeni_gpt_rapor_isleme(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message or update.edited_message
    if not msg:
        return

    user_id = msg.from_user.id
    chat_type = msg.chat.type
    
    is_group = chat_type in ["group", "supergroup"]
    is_dm = chat_type == "private"

    if is_media_message(msg):
        logging.info(f"⛔ Medya mesajı tespit edildi → AI analizi yapılmayacak. User: {user_id}, Chat Type: {chat_type}")
        return

    metin = msg.text or msg.caption
    if not metin:
        return

    if metin.startswith(('/', '.', '!', '\\')):
        return

    try:
        # ÖNCE "ÇALIŞMA YOK" KONTROLÜ
        is_calisma_yok = is_calisma_yok_raporu(metin)
        
        raporlar = process_incoming_message(metin, is_group)
        
        if is_dm and isinstance(raporlar, dict) and raporlar.get('dm_info') == 'no_report_detected':
            # Eğer çalışma yok raporuysa, bunu kabul et
            if is_calisma_yok:
                # Çalışma yok raporu için basit bir JSON oluştur
                today = dt.datetime.now(TZ).date()
                
                # Şantiye adını çıkarmaya çalış
                site = "BELİRSİZ"
                patterns = [
                    r'(LOT13|LOT71|SKP|BWC|STADYUM|DMC|OHP|YHP|TYM|MMP|RMC|PİRAMİT|MOS|KÖKSARAY|DATA CENTR)',
                    r'ŞANTİYE:\s*(\w+)',
                    r'📍\s*ŞANTİYE:\s*(\w+)'
                ]
                
                for pattern in patterns:
                    match = re.search(pattern, metin, re.IGNORECASE)
                    if match:
                        site = normalize_site_name(match.group(1))
                        break
                
                # Çalışma yok raporu için basit bir rapor oluştur
                calisma_yok_rapor = {
                    "date": today.strftime('%Y-%m-%d'),
                    "site": site,
                    "staff": 0,
                    "calisan": 0,
                    "mobilizasyon": 0,
                    "ambarci": 0,
                    "izinli": 0,
                    "dis_gorev": [],
                    "dis_gorev_toplam": 0,
                    "genel_toplam": 0
                }
                
                raporlar = [calisma_yok_rapor]
                logging.info(f"📝 DM'de 'Çalışma yok' raporu tespit edildi: {site}")
            else:
                await msg.reply_text(
                    "❌ Bu mesaj bir rapor olarak algılanmadı.\n\n"
                    "Lütfen şantiye, tarih ve iş bilgilerini içeren bir rapor gönderin.\n"
                    "Örnek: \"01.11.2026 LOT13 2.kat kablo çekimi 5 kişi\"\n\n"
                    "Not: 'Çalışma yok' raporları da kabul edilir: \"OHP bugün çalışma yok\""
                )
                return
        
        if not raporlar or (isinstance(raporlar, list) and len(raporlar) == 0):
            # Çalışma yok raporuysa kabul et
            if is_calisma_yok:
                today = dt.datetime.now(TZ).date()
                
                # Şantiye adını çıkarmaya çalış
                site = "BELİRSİZ"
                patterns = [
                    r'(LOT13|LOT71|SKP|BWC|STADYUM|DMC|OHP|YHP|TYM|MMP|RMC|PİRAMİT|MOS|KÖKSARAY|DATA CENTR)',
                    r'ŞANTİYE:\s*(\w+)',
                    r'📍\s*ŞANTİYE:\s*(\w+)'
                ]
                
                for pattern in patterns:
                    match = re.search(pattern, metin, re.IGNORECASE)
                    if match:
                        site = normalize_site_name(match.group(1))
                        break
                
                # Çalışma yok raporu için basit bir rapor oluştur
                calisma_yok_rapor = {
                    "date": today.strftime('%Y-%m-%d'),
                    "site": site,
                    "staff": 0,
                    "calisan": 0,
                    "mobilizasyon": 0,
                    "ambarci": 0,
                    "izinli": 0,
                    "dis_gorev": [],
                    "dis_gorev_toplam": 0,
                    "genel_toplam": 0
                }
                
                raporlar = [calisma_yok_rapor]
                logging.info(f"📝 'Çalışma yok' raporu tespit edildi (boş raporlar): {site}")
            else:
                logging.info(f"🤖 GPT: Rapor bulunamadı - {user_id} (Chat Type: {chat_type})")
                
                if is_dm:
                    await msg.reply_text(
                        "❌ Rapor bulunamadı.\n\n"
                        "Lütfen şantiye raporunuzu aşağıdaki formatta gönderin:\n"
                        "• Tarih (01.01.2026)\n" 
                        "• Şantiye adı (LOT13, BWC, SKP vb.)\n"
                        "• Yapılan işler\n"
                        "• Personel bilgisi\n\n"
                        "Örnek: \"01.11.2026 LOT13 2.kat kablo çekimi 5 kişi\"\n\n"
                        "Not: 'Çalışma yok' raporları da kabul edilir: \"OHP bugün çalışma yok\""
                    )
                return

        logging.info(f"🤖 GPT: {len(raporlar)} rapor çıkarıldı - {user_id} (Chat Type: {chat_type})")
        
        kullanici_adi = id_to_name.get(user_id, "Kullanıcı")
        
        basarili_kayitlar = 0
        for i, rapor in enumerate(raporlar):
            try:
                await raporu_gpt_formatinda_kaydet(user_id, kullanici_adi, metin, rapor, msg, i+1)
                basarili_kayitlar += 1
            except Exception as e:
                logging.error(f"❌ Rapor {i+1} kaydetme hatası: {e}")
                if is_dm:
                    await msg.reply_text(f"❌ Rapor {i+1} kaydedilemedi: {str(e)}")
        
        if is_dm:
            if basarili_kayitlar == len(raporlar):
                if len(raporlar) == 1:
                    await msg.reply_text("✅ Raporunuz başarıyla işlendi!")
                else:
                    await msg.reply_text(f"✅ {len(raporlar)} rapor başarıyla işlendi!")
            else:
                await msg.reply_text(f"⚠️ {basarili_kayitlar}/{len(raporlar)} rapor işlendi. Bazı raporlar kaydedilemedi.")
        
        logging.info(f"📊 Grup raporu işlendi: {basarili_kayitlar}/{len(raporlar)} başarılı")
            
    except Exception as e:
        logging.error(f"❌ GPT rapor işleme hatası: {e}")
        if is_dm:
            await msg.reply_text("❌ Rapor işlenirken bir hata oluştu. Lütfen daha sonra tekrar deneyin.")

async def excel_durum_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await super_admin_kontrol(update, context):
        return
    
    try:
        mesaj = "📊 EXCEL SİSTEM DURUMU - ŞANTİYE BAZLI SİSTEM\n\n"
        
        if os.path.exists(USERS_FILE):
            file_size = os.path.getsize(USERS_FILE)
            file_mtime = dt.datetime.fromtimestamp(os.path.getmtime(USERS_FILE))
            mesaj += f"✅ Dosya Mevcut: {USERS_FILE}\n"
            mesaj += f"📏 Boyut: {file_size} bytes\n"
            mesaj += f"🕒 Son Değişiklik: {file_mtime.strftime('%d.%m.%Y %H:%M')}\n"
            
            current_hash = get_file_hash(USERS_FILE)
            mesaj += f"🔐 Hash: {current_hash[:8] if current_hash else 'Hesaplanamadı'}\n\n"
        else:
            mesaj += f"❌ Dosya Bulunamadı: {USERS_FILE}\n\n"
            mesaj += "🔄 Fallback sistem aktif\n\n"
        
        mesaj += "💾 ÖNBELLEK DURUMU:\n"
        mesaj += f"• Excel Hash: {excel_file_hash[:8] if excel_file_hash else 'Yok'}\n"
        mesaj += f"• Son Yükleme: {dt.datetime.fromtimestamp(excel_last_modified).strftime('%d.%m.%Y %H:%M') if excel_last_modified else 'Yok'}\n"
        mesaj += f"• DataFrame: {'Mevcut' if df is not None else 'Yok'}\n\n"
        
        mesaj += "📈 ŞANTİYE İSTATİSTİKLERİ:\n"
        mesaj += f"• Aktif Kullanıcı: {len(rapor_sorumlulari)} \n"
        mesaj += f"• Adminler: {len(ADMINS)}\n"
        mesaj += f"• İzleyiciler: {len(IZLEYICILER)}\n"
        mesaj += f"• Toplam Kullanıcı: {len(TUM_KULLANICILAR)}\n"
        mesaj += f"• Şantiyeler: {len(santiye_sorumlulari)} \n\n"
        
        mesaj += "🏗️ AKTİF ŞANTİYELER :\n"
        for santiye in sorted(santiye_sorumlulari.keys())[:10]:
            sorumlu_sayisi = len(santiye_sorumlulari[santiye])
            mesaj += f"• {santiye}: {sorumlu_sayisi} sorumlu\n"
        
        if len(santiye_sorumlulari) > 10:
            mesaj += f"• ... ve {len(santiye_sorumlulari) - 10} diğer şantiye\n"
        
        mesaj += "\n🛡️ GÜVENLİK SİSTEMİ:\n"
        mesaj += f"• Fallback Aktif: {'Evet' if df is not None and any(df['Telegram ID'] == 1000157326) else 'Hayır'}\n"
        mesaj += f"• Super Admin: {SUPER_ADMIN_ID} ({'Aktif' if SUPER_ADMIN_ID in ADMINS else 'Pasif'})\n"
        mesaj += f"• Telegram ID Format: 8-10 digit\n"
        mesaj += f"• Aktif/Pasif Kontrolü: 'E'/'H'\n"
        
        await update.message.reply_text(mesaj)
        
    except Exception as e:
        await update.message.reply_text(f"❌ Durum kontrol hatası: {e}")

async def yeni_uye_karşilama(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        for member in update.message.new_chat_members:
            if member.id == context.bot.id:
                await update.message.reply_text(
                    "🤖 Rapor Botu Aktif!\n\n"
                    "Ben şantiye raporlarınızı otomatik olarak işleyen bir botum.\n"
                    "Günlük çalışma raporlarınızı gönderebilirsiniz.\n\n"
                    "📋 Özellikler:\n"
                    "• Otomatik rapor analizi\n"
                    "• Tarih tanıma\n"
                    "• Personel sayımı\n"
                    "• Şantiye takibi\n\n"
                    "Kolay gelsin! 👷‍♂️"
                )
            else:
                await update.message.reply_text(
                    f"👋 Hoş geldin {member.first_name}!\n\n"
                    f"🤖 Ben şantiye raporlarınızı otomatik işleyen bir botum.\n"
                    f"Günlük çalışma raporlarınızı bu gruba gönderebilirsiniz.\n\n"
                    f"Kolay gelsin! 👷‍♂️"
                )
    except Exception as e:
        logging.error(f"Yeni üye karşılama hatası: {e}")

# Gelişmiş hata yönetimi ile veritabanı başlatma
def init_database():
    """Kapsamlı hata yönetimi ile veritabanını başlat"""
    try:
        _sync_execute_safe("""
            CREATE TABLE IF NOT EXISTS schema_version (
                id INTEGER PRIMARY KEY CHECK (id=1), 
                version INTEGER NOT NULL
            )
        """)
        
        _sync_execute_safe("""
            INSERT INTO schema_version (id, version) 
            SELECT 1, 2
            WHERE NOT EXISTS(SELECT 1 FROM schema_version WHERE id=1)
        """)
        
        _sync_execute_safe("""
            CREATE TABLE IF NOT EXISTS reports (
                id SERIAL PRIMARY KEY,
                user_id BIGINT NOT NULL,
                project_name VARCHAR(200),
                report_date DATE NOT NULL,
                report_type VARCHAR(50) NOT NULL,
                person_count INTEGER DEFAULT 1,
                work_description TEXT,
                work_category VARCHAR(100),
                personnel_type VARCHAR(100),
                delivered_date DATE,
                is_edited BOOLEAN DEFAULT FALSE,
                ai_analysis JSONB,
                message_id BIGINT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        _sync_execute_safe("""
            CREATE TABLE IF NOT EXISTS ai_logs (
                id SERIAL PRIMARY KEY,
                timestamp TEXT,
                user_id INTEGER,
                rapor_metni TEXT,
                ai_cevap TEXT,
                basarili INTEGER,
                hata_mesaji TEXT
            )
        """)
        
        index_queries = [
            "CREATE INDEX IF NOT EXISTS idx_reports_date_user ON reports(report_date, user_id)",
            "CREATE INDEX IF NOT EXISTS idx_reports_project_date ON reports(project_name, report_date)",
            "CREATE INDEX IF NOT EXISTS idx_reports_type_date ON reports(report_type, report_date)",
            "CREATE INDEX IF NOT EXISTS idx_reports_user_date ON reports(user_id, report_date)"
        ]
        
        for query in index_queries:
            try:
                _sync_execute_safe(query)
            except Exception as e:
                logging.warning(f"İndeks oluşturma uyarısı: {e}")
        
        logging.info("✅ Veritabanı şeması başarıyla başlatıldı")
        
    except Exception as e:
        logging.error(f"❌ Veritabanı başlatma hatası: {e}")
        raise

init_database()
init_db_pool()

async def get_santiye_rapor_durumu(bugun):
    """Güvenli tuple işleme ile şantiye rapor durumunu al - OPSİYONEL ŞANTİYELER HARİÇ"""
    try:
        rows = await async_fetchall("""
            SELECT DISTINCT project_name FROM reports 
            WHERE report_date = %s AND project_name IS NOT NULL AND project_name != 'BELİRSİZ'
        """, (bugun,))
        
        if not rows:
            return set()
            
        return set(normalize_site_name(safe_get_tuple_value(row, 0, '')) for row in rows 
                  if safe_get_tuple_value(row, 0, '') and safe_get_tuple_value(row, 0, '') != "TÜMÜ")
    except Exception as e:
        logging.error(f"Şantiye rapor durumu hatası: {e}")
        return set()

async def get_eksik_santiyeler(bugun):
    try:
        # TÜMÜ şantiyesini filtrele, sabit şantiyeleri ekle, OPSİYONEL ŞANTİYELERİ ÇIKAR
        tum_santiyeler = set(santiye for santiye in santiye_sorumlulari.keys() if santiye != "TÜMÜ")
        tum_santiyeler = tum_santiyeler.union(set(SABIT_SANTIYELER))
        
        # OPSİYONEL ŞANTİYELERİ ÇIKAR (OHP gibi)
        tum_santiyeler = tum_santiyeler - set(OPSIYONEL_SANTIYELER)
        
        rapor_veren_santiyeler = await get_santiye_rapor_durumu(bugun)
        eksik_santiyeler = tum_santiyeler - rapor_veren_santiyeler
        
        return {santiye: santiye_sorumlulari.get(santiye, []) for santiye in eksik_santiyeler}
    except Exception as e:
        logging.error(f"Eksik şantiye sorgu hatası: {e}")
        return {}

async def get_santiye_bazli_rapor_durumu(bugun):
    try:
        # TÜMÜ şantiyesini filtrele, SABİT ŞANTİYELERİ EKLE, OPSİYONEL ŞANTİYELERİ ÇIKAR
        tum_santiyeler = set(santiye for santiye in santiye_sorumlulari.keys() if santiye != "TÜMÜ")
        tum_santiyeler = tum_santiyeler.union(set(SABIT_SANTIYELER))
        
        # OPSİYONEL ŞANTİYELERİ ÇIKAR (OHP gibi)
        tum_santiyeler = tum_santiyeler - set(OPSIYONEL_SANTIYELER)
        
        rapor_veren_santiyeler = await get_santiye_rapor_durumu(bugun)
        
        rows = await async_fetchall("""
            SELECT project_name, user_id FROM reports 
            WHERE report_date = %s AND project_name IS NOT NULL AND project_name != 'BELİRSİZ'
        """, (bugun,))
        
        santiye_rapor_verenler = {}
        for row in rows:
            if row and len(row) >= 2:
                project_name = safe_get_tuple_value(row, 0, '')
                # PROJE ADINI NORMALİZE ET
                project_name = normalize_site_name(project_name)
                user_id = safe_get_tuple_value(row, 1, 0)
                if project_name and project_name != "TÜMÜ" and user_id:
                    if project_name not in santiye_rapor_verenler:
                        santiye_rapor_verenler[project_name] = []
                    santiye_rapor_verenler[project_name].append(user_id)
    
        return {
            'tum_santiyeler': tum_santiyeler,
            'rapor_veren_santiyeler': rapor_veren_santiyeler,
            'eksik_santiyeler': tum_santiyeler - rapor_veren_santiyeler,
            'santiye_rapor_verenler': santiye_rapor_verenler
        }
    except Exception as e:
        logging.error(f"Şantiye bazlı rapor durumu hatası: {e}")
        return {'tum_santiyeler': set(), 'rapor_veren_santiyeler': set(), 'eksik_santiyeler': set(), 'santiye_rapor_verenler': {}}

class MaliyetAnaliz:
    def __init__(self):
        self.gpt_count = 0
        self.fallback_count = 0
        
    def kayit_ekle(self, kaynak):
        if kaynak == 'gpt':
            self.gpt_count += 1
        else:
            self.fallback_count += 1
    
    def maliyet_raporu(self):
        toplam = self.gpt_count + self.fallback_count
        if toplam == 0:
            return "📊 Henüz işlem yok"
        
        gpt_orani = (self.gpt_count / toplam) * 100
        maliyet = self.gpt_count * 0.0015
        
        return (
            f"📊 MALİYET ANALİZİ\n\n"
            f"🤖 GPT İşlemleri: {self.gpt_count} (%{gpt_orani:.1f})\n"
            f"🔄 Fallback: {self.fallback_count}\n"
            f"💰 Tahmini Maliyet: ${maliyet:.4f}\n"
            f"🎯 Başarı Oranı: %{gpt_orani:.1f}"
        )
    
    def detayli_ai_raporu(self):
        try:
            result = _sync_fetchone_safe("""
                SELECT 
                    COUNT(*) as toplam,
                    SUM(CASE WHEN basarili = 1 THEN 1 ELSE 0 END) as basarili,
                    SUM(CASE WHEN basarili = 0 THEN 1 ELSE 0 END) as basarisiz,
                    MIN(timestamp) as ilk_tarih,
                    MAX(timestamp) as son_tarih
                FROM ai_logs
            """)
            
            if not result or len(result) < 5 or safe_get_tuple_value(result, 0, 0) is None or safe_get_tuple_value(result, 0, 0) == 0:
                return "🤖 AI Raporu: Henüz AI kullanımı yok"
            
            toplam = safe_get_tuple_value(result, 0, 0)
            basarili = safe_get_tuple_value(result, 1, 0)
            basarisiz = safe_get_tuple_value(result, 2, 0)
            ilk_tarih = safe_get_tuple_value(result, 3, '')
            son_tarih = safe_get_tuple_value(result, 4, '')
            
            rows = _sync_fetchall_safe("""
                SELECT DATE(timestamp::timestamp) as gun, 
                       COUNT(*) as toplam,
                       SUM(CASE WHEN basarili = 1 THEN 1 ELSE 0 END) as basarili
                FROM ai_logs 
                WHERE timestamp::timestamp >= CURRENT_DATE - INTERVAL '7 days'
                GROUP BY DATE(timestamp::timestamp) 
                ORDER BY gun DESC
            """)
            
            rapor = "🤖 DETAYLI AI RAPORU\n\n"
            rapor += f"📈 Genel İstatistikler:\n"
            rapor += f"• Toplam İşlem: {toplam}\n"
            rapor += f"• Başarılı: {basarili} (%{(basarili/toplam*100):.1f})\n"
            rapor += f"• Başarısız: {basarilis}\n"
            rapor += f"• İlk Kullanım: {ilk_tarih[:10] if ilk_tarih else 'Yok'}\n"
            rapor += f"• Son Kullanım: {son_tarih[:10] if son_tarih else 'Yok'}\n\n"
            
            rapor += f"📅 Son 7 Gün:\n"
            for row in rows:
                if row and len(row) >= 3:
                    gun = safe_get_tuple_value(row, 0, '')
                    toplam_gun = safe_get_tuple_value(row, 1, 0)
                    basarili_gun = safe_get_tuple_value(row, 2, 0)
                    oran = (basarili_gun/toplam_gun*100) if toplam_gun > 0 else 0
                    rapor += f"• {gun}: {basarili_gun}/{toplam_gun} (%{oran:.1f})\n"
            
            return rapor
            
        except Exception as e:
            return f"❌ AI raporu oluşturulurken hata: {e}"

maliyet_analiz = MaliyetAnaliz()

def parse_rapor_tarihi(metin):
    try:
        bugun = dt.datetime.now(TZ).date()
        metin_lower = metin.lower()
        
        if 'bugün' in metin_lower or 'bugun' in metin_lower:
            return bugun
        if 'dün' in metin_lower or 'dun' in metin_lower:
            return bugun - dt.timedelta(days=1)
        
        date_patterns = [
            r'(\d{1,2})[\.\/\-](\d{1,2})[\.\/\-](\d{4})',
            r'(\d{1,2})[\.\/\-](\d{1,2})[\.\/\-](\d{2})',
            r'(\d{4})[\.\/\-](\d{1,2})[\.\/\-](\d{1,2})',
            r'(\d{1,2})\s*[/\.\-]\s*(\d{1,2})\s*[/\.\-]\s*(\d{4})',
            r'(\d{1,2})\s*[/\.\-]\s*(\d{1,2})\s*[/\.\-]\s*(\d{2})',
        ]
        
        for pattern in date_patterns:
            matches = re.finditer(pattern, metin)
            for match in matches:
                groups = match.groups()
                if len(groups) == 3:
                    try:
                        if len(groups[2]) == 4:
                            day, month, year = int(groups[0]), int(groups[1]), int(groups[2])
                        elif len(groups[0]) == 4:
                            year, month, day = int(groups[0]), int(groups[1]), int(groups[2])
                        else:
                            day, month, year = int(groups[0]), int(groups[1]), int(groups[2])
                            year += 2000
                        
                        parsed_date = dt.datetime(year, month, day).date()
                        if parsed_date <= bugun:
                            return parsed_date
                    except ValueError:
                        continue
        
        return None
    except Exception:
        return None

def izin_mi(metin):
    metin_lower = metin.lower()
    izin_kelimeler = ['izin', 'rapor yok', 'iş yok', 'çalışma yok', 'tatil', 'hasta', 'izindeyim']
    return any(kelime in metin_lower for kelime in izin_kelimeler)

async def tarih_kontrol_et(rapor_tarihi, user_id):
    bugun = dt.datetime.now(TZ).date()
    
    if not rapor_tarihi:
        return False, "❌ Tarih bulunamadı. Lütfen raporunuzda tarih belirtiniz."
    
    if rapor_tarihi > bugun:
        return False, "❌ Gelecek tarihli rapor. Lütfen bugün veya geçmiş tarih kullanınız."
    
    iki_ay_once = bugun - dt.timedelta(days=60)
    if rapor_tarihi < iki_ay_once:
        return False, "❌ Çok eski tarihli rapor. Lütfen son 2 ay içinde bir tarih kullanınız."
    
    result = await async_fetchone("SELECT EXISTS(SELECT 1 FROM reports WHERE project_name = %s AND report_date = %s)", 
                  (user_id, rapor_tarihi))
    
    exists = safe_get_tuple_value(result, 0, False) if result else False
    if exists:
        return False, "❌ Bu tarih için zaten rapor gönderdiniz."
    
    return True, ""

def parse_tr_date(date_str):
    try:
        normalized_date = date_str.replace('/', '.').replace('-', '.')
        parts = normalized_date.split('.')
        if len(parts) == 3:
            if len(parts[2]) == 4:
                return dt.datetime.strptime(normalized_date, '%d.%m.%Y').date()
            elif len(parts[0]) == 4:
                return dt.datetime.strptime(normalized_date, '%Y.%m.%d').date()
        raise ValueError("Geçersiz tarih formatı")
    except:
        raise ValueError("Geçersiz tarih formatı")

def week_window_to_today():
    end_date = dt.datetime.now(TZ).date()
    start_date = end_date - dt.timedelta(days=6)
    return start_date, end_date

def is_admin(user_id):
    return user_id in ADMINS

def is_super_admin(user_id):
    return user_id == SUPER_ADMIN_ID

def is_izleyici(user_id):
    return user_id in IZLEYICILER

async def admin_kontrol(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if not is_admin(user_id):
        await update.message.reply_text("❌ Bu komut sadece BOT yöneticileri içindir.")
        return False
    return True

async def super_admin_kontrol(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if not is_super_admin(user_id):
        await update.message.reply_text("❌ Bu komut sadece Super Admin içindir.")
        return False
    return True

async def hata_bildirimi(context: ContextTypes.DEFAULT_TYPE, hata_mesaji: str):
    for admin_id in ADMINS:
        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=f"⚠️ Sistem Hatası: {hata_mesaji}"
            )
            await asyncio.sleep(0.1)
        except Exception as e:
            logging.error(f"Hata bildirimi {admin_id} adminine gönderilemedi: {e}")

# Personel özeti fonksiyonu - ŞANTİYE BAZLI - TÜMÜ FİLTRELENDİ - KRİTİK GÜNCELLEME!
async def generate_gelismis_personel_ozeti(target_date):
    """Güvenli tuple işleme ile gelişmiş personel özeti oluştur - KRİTİK GÜNCELLEME: Şantiye Başlığı vs Genel Toplam"""
    try:
        rows = await async_fetchall("""
            SELECT user_id, report_type, project_name, person_count, work_description, ai_analysis
            FROM reports WHERE report_date = %s
        """, (target_date,))
        
        if not rows:
            return f"📭 {target_date.strftime('%d.%m.%Y')} tarihinde rapor bulunamadı."
        
        proje_analizleri = {}
        tum_projeler = set()
        
        genel_staff = 0
        genel_calisan = 0
        genel_mobilizasyon = 0
        genel_ambarci = 0
        genel_izinli = 0
        genel_dis_gorev_toplam = 0
        genel_toplam = 0
        
        for row in rows:
            if len(row) < 6:
                continue
            user_id = safe_get_tuple_value(row, 0, 0)
            rapor_tipi = safe_get_tuple_value(row, 1, '')
            proje_adi = safe_get_tuple_value(row, 2, 'BELİRSİZ')
            kisi_sayisi = safe_get_tuple_value(row, 3, 0)
            yapilan_is = safe_get_tuple_value(row, 4, '')
            ai_analysis = safe_get_tuple_value(row, 5, '{}')
            
            # PROJE ADINI NORMALİZE ET
            proje_adi = normalize_site_name(proje_adi)
            
            if not proje_adi or proje_adi == "TÜMÜ":
                continue
                
            if proje_adi not in proje_analizleri:
                proje_analizleri[proje_adi] = {
                    'toplam': 0,
                    'staff': 0, 'calisan': 0, 'mobilizasyon': 0, 'ambarci': 0, 'izinli': 0, 'dis_gorev_toplam': 0,
                    'santiye_baslik': 0  # YENİ: Şantiye başlık sayısı (dış görevler HARİÇ)
                }
            
            try:
                ai_data = safe_json_loads(ai_analysis)
                yeni_format = ai_data.get('yeni_sabit_format', {})
                personel_dagilimi = ai_data.get('personel_dagilimi', {})
                is_calisma_yok = ai_data.get('is_calisma_yok', False)  # YENİ: Çalışma yok kontrolü
                
                if yeni_format:
                    staff_count = yeni_format.get('staff', 0)
                    calisan_count = yeni_format.get('calisan', 0)
                    mobilizasyon_count = yeni_format.get('mobilizasyon', 0)
                    ambarci_count = yeni_format.get('ambarci', 0)
                    izinli_count = yeni_format.get('izinli', 0)
                    dis_gorev_toplam_count = yeni_format.get('dis_gorev_toplam', 0)
                    
                    proje_analizleri[proje_adi]['staff'] += staff_count
                    proje_analizleri[proje_adi]['calisan'] += calisan_count
                    proje_analizleri[proje_adi]['mobilizasyon'] += mobilizasyon_count
                    proje_analizleri[proje_adi]['ambarci'] += ambarci_count
                    proje_analizleri[proje_adi]['izinli'] += izinli_count
                    proje_analizleri[proje_adi]['dis_gorev_toplam'] += dis_gorev_toplam_count
                    
                    # KRİTİK GÜNCELLEME: Şantiye başlık hesaplaması (dış görevler HARİÇ)
                    santiye_baslik = staff_count + calisan_count + mobilizasyon_count + ambarci_count + izinli_count
                    proje_analizleri[proje_adi]['santiye_baslik'] += santiye_baslik
                    
                    # Toplam = Şantiye başlık + dış görevler
                    proje_analizleri[proje_adi]['toplam'] = santiye_baslik + dis_gorev_toplam_count
                    
                elif personel_dagilimi:
                    staff_count = personel_dagilimi.get('staff', 0)
                    calisan_count = personel_dagilimi.get('calisan', 0)
                    mobilizasyon_count = personel_dagilimi.get('mobilizasyon', 0)
                    ambarci_count = personel_dagilimi.get('ambarci', 0)
                    izinli_count = personel_dagilimi.get('izinli', 0)
                    dis_gorev_toplam_count = personel_dagilimi.get('dis_gorev_toplam', 0)
                    
                    proje_analizleri[proje_adi]['staff'] += staff_count
                    proje_analizleri[proje_adi]['calisan'] += calisan_count
                    proje_analizleri[proje_adi]['mobilizasyon'] += mobilizasyon_count
                    proje_analizleri[proje_adi]['ambarci'] += ambarci_count
                    proje_analizleri[proje_adi]['izinli'] += izinli_count
                    proje_analizleri[proje_adi]['dis_gorev_toplam'] += dis_gorev_toplam_count
                    
                    # KRİTİK GÜNCELLEME: Şantiye başlık hesaplaması (dış görevler HARİÇ)
                    santiye_baslik = staff_count + calisan_count + mobilizasyon_count + ambarci_count + izinli_count
                    proje_analizleri[proje_adi]['santiye_baslik'] += santiye_baslik
                    
                    # Toplam = Şantiye başlık + dış görevler
                    proje_analizleri[proje_adi]['toplam'] = santiye_baslik + dis_gorev_toplam_count
                    
                else:
                    yapilan_is_lower = (yapilan_is or '').lower()
                    
                    if 'staff' in yapilan_is_lower:
                        proje_analizleri[proje_adi]['staff'] += kisi_sayisi
                    elif 'mobilizasyon' in yapilan_is_lower:
                        proje_analizleri[proje_adi]['mobilizasyon'] += kisi_sayisi
                    elif 'ambarci' in yapilan_is_lower or 'ambarcı' in yapilan_is_lower:
                        proje_analizleri[proje_adi]['ambarci'] += kisi_sayisi
                    elif rapor_tipi == "IZIN/ISYOK":
                        proje_analizleri[proje_adi]['izinli'] += kisi_sayisi
                    else:
                        proje_analizleri[proje_adi]['calisan'] += kisi_sayisi
                    
                    # Eski mantık (fallback)
                    proje_analizleri[proje_adi]['toplam'] += kisi_sayisi
                    proje_analizleri[proje_adi]['santiye_baslik'] += kisi_sayisi
                        
            except Exception as e:
                logging.error(f"Personel analiz hatası: {e}")
                yapilan_is_lower = (yapilan_is or '').lower()
                
                if 'staff' in yapilan_is_lower:
                    proje_analizleri[proje_adi]['staff'] += kisi_sayisi
                elif 'mobilizasyon' in yapilan_is_lower:
                    proje_analizleri[proje_adi]['mobilizasyon'] += kisi_sayisi
                elif 'ambarci' in yapilan_is_lower or 'ambarcı' in yapilan_is_lower:
                    proje_analizleri[proje_adi]['ambarci'] += kisi_sayisi
                elif rapor_tipi == "IZIN/ISYOK":
                    proje_analizleri[proje_adi]['izinli'] += kisi_sayisi
                else:
                    proje_analizleri[proje_adi]['calisan'] += kisi_sayisi
                
                # Eski mantık (fallback)
                proje_analizleri[proje_adi]['toplam'] += kisi_sayisi
                proje_analizleri[proje_adi]['santiye_baslik'] += kisi_sayisi
            
            tum_projeler.add(proje_adi)
        
        # KRİTİK GÜNCELLEME: Genel toplam hesaplaması (tüm şantiyelerin toplamı + kendi dış görevleri)
        for proje_adi, analiz in proje_analizleri.items():
            genel_staff += analiz['staff']
            genel_calisan += analiz['calisan']
            genel_mobilizasyon += analiz['mobilizasyon']
            genel_ambarci += analiz['ambarci']
            genel_izinli += analiz['izinli']
            genel_dis_gorev_toplam += analiz['dis_gorev_toplam']
            genel_toplam += analiz['toplam']  # Bu artık şantiye başlık + dış görevler içeriyor
        
        mesaj = f"📊 {target_date.strftime('%d.%m.%Y')} GÜNLÜK PERSONEL ÖZETİ\n\n"
        
        for proje_adi, analiz in sorted(proje_analizleri.items(), key=lambda x: x[1]['toplam'], reverse=True):
            # KRİTİK GÜNCELLEME: Şantiye başlık = santiye_baslik (dış görevler HARİÇ)
            santiye_baslik = analiz['santiye_baslik']
            
            if santiye_baslik > 0 or analiz['dis_gorev_toplam'] > 0:
                emoji = "🏢" if proje_adi == "TYM" else "🏗️"
                mesaj += f"{emoji} {proje_adi}: {santiye_baslik} kişi"
                if analiz['dis_gorev_toplam'] > 0:
                    mesaj += f" (Dış görev: {analiz['dis_gorev_toplam']})"
                mesaj += "\n"
                
                durum_detay = []
                if analiz['staff'] > 0: 
                    durum_detay.append(f"Staff:{analiz['staff']}")
                if analiz['calisan'] > 0: 
                    durum_detay.append(f"Çalışan:{analiz['calisan']}")
                if analiz['mobilizasyon'] > 0: 
                    durum_detay.append(f"Mobilizasyon:{analiz['mobilizasyon']}")
                if analiz['ambarci'] > 0: 
                    durum_detay.append(f"Ambarcı:{analiz['ambarci']}")
                if analiz['izinli'] > 0: 
                    durum_detay.append(f"İzinli:{analiz['izinli']}")
                
                if durum_detay:
                    mesaj += f"   └─ {', '.join(durum_detay)}\n\n"
        
        # KRİTİK GÜNCELLEME: Genel toplam = Σ(tüm şantiyelerin toplamı)
        mesaj += f"📈 GENEL TOPLAM: {genel_toplam} kişi\n"
        
        if genel_toplam > 0:
            mesaj += f"🎯 DAĞILIM:\n"
            if genel_staff > 0:
                mesaj += f"• Staff: {genel_staff} (%{genel_staff/genel_toplam*100:.1f})\n"
            if genel_calisan > 0:
                mesaj += f"• Çalışan: {genel_calisan} (%{genel_calisan/genel_toplam*100:.1f})\n"
            if genel_mobilizasyon > 0:
                mesaj += f"• Mobilizasyon: {genel_mobilizasyon} (%{genel_mobilizasyon/genel_toplam*100:.1f})\n"
            if genel_ambarci > 0:
                mesaj += f"• Ambarcı: {genel_ambarci} (%{genel_ambarci/genel_toplam*100:.1f})\n"
            if genel_izinli > 0:
                mesaj += f"• İzinli: {genel_izinli} (%{genel_izinli/genel_toplam*100:.1f})\n"
            if genel_dis_gorev_toplam > 0:
                mesaj += f"• Dış Görev: {genel_dis_gorev_toplam} (%{genel_dis_gorev_toplam/genel_toplam*100:.1f})\n"
        
        # TÜM SABİT ŞANTİYELERİ DAHİL ET, OPSİYONEL HARİÇ
        tum_santiyeler = set(SABIT_SANTIYELER).union(set(santiye for santiye in santiye_sorumlulari.keys() if santiye != "TÜMÜ"))
        tum_santiyeler = tum_santiyeler - set(OPSIYONEL_SANTIYELER)
        aktif_projeler = set(proje_analizleri.keys())
        eksik_projeler = [s for s in (tum_santiyeler - aktif_projeler) if s not in ["Belli değil", "Tümü"]]
        
        if eksik_projeler:
            mesaj += f"\n❌ EKSİK ŞANTİYELER: {', '.join(sorted(eksik_projeler))}"
        
        return mesaj
    except Exception as e:
        return f"❌ Rapor oluşturulurken hata oluştu: {e}"

# HAFTALIK RAPOR FONKSİYONU - VERİMLİLİK KALDIRILDI
async def generate_haftalik_rapor_mesaji(start_date, end_date):
    try:
        rows = await async_fetchall("""
            SELECT user_id, COUNT(*) as rapor_sayisi
            FROM reports 
            WHERE report_date BETWEEN %s AND %s
            GROUP BY user_id
            ORDER BY rapor_sayisi DESC
        """, (start_date, end_date))
        
        if not rows:
            return f"📭 {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')} arasında rapor bulunamadı."
        
        toplam_rapor = sum([safe_get_tuple_value(x, 1, 0) for x in rows])
        gun_sayisi = (end_date - start_date).days + 1
        
        proje_detay_rows = await async_fetchall("""
            SELECT project_name, ai_analysis
            FROM reports 
            WHERE report_date BETWEEN %s AND %s AND project_name IS NOT NULL AND project_name != 'BELİRSİZ'
        """, (start_date, end_date))
        
        proje_analizleri = {}
        
        for row in proje_detay_rows:
            if len(row) < 2:
                continue
                
            proje_adi = safe_get_tuple_value(row, 0, '')
            ai_analysis = safe_get_tuple_value(row, 1, '{}')
            
            # PROJE ADINI NORMALİZE ET
            proje_adi = normalize_site_name(proje_adi)
            
            # TÜMÜ şantiyesini filtrele
            if not proje_adi or proje_adi == "TÜMÜ":
                continue
                
            if proje_adi not in proje_analizleri:
                proje_analizleri[proje_adi] = {
                    'staff': 0, 'calisan': 0, 'mobilizasyon': 0, 'ambarci': 0, 'izinli': 0, 'dis_gorev_toplam': 0, 
                    'toplam': 0, 'santiye_baslik': 0  # YENİ: Şantiye başlık sayısı
                }
            
            try:
                ai_data = safe_json_loads(ai_analysis)
                yeni_format = ai_data.get('yeni_sabit_format', {})
                personel_dagilimi = ai_data.get('personel_dagilimi', {})
                
                if yeni_format:
                    staff_count = yeni_format.get('staff', 0)
                    calisan_count = yeni_format.get('calisan', 0)
                    mobilizasyon_count = yeni_format.get('mobilizasyon', 0)
                    ambarci_count = yeni_format.get('ambarci', 0)
                    izinli_count = yeni_format.get('izinli', 0)
                    dis_gorev_toplam_count = yeni_format.get('dis_gorev_toplam', 0)
                    
                    proje_analizleri[proje_adi]['staff'] += staff_count
                    proje_analizleri[proje_adi]['calisan'] += calisan_count
                    proje_analizleri[proje_adi]['mobilizasyon'] += mobilizasyon_count
                    proje_analizleri[proje_adi]['ambarci'] += ambarci_count
                    proje_analizleri[proje_adi]['izinli'] += izinli_count
                    proje_analizleri[proje_adi]['dis_gorev_toplam'] += dis_gorev_toplam_count
                    
                    # KRİTİK GÜNCELLEME: Şantiye başlık hesaplaması (dış görevler HARİÇ)
                    santiye_baslik = staff_count + calisan_count + mobilizasyon_count + ambarci_count + izinli_count
                    proje_analizleri[proje_adi]['santiye_baslik'] += santiye_baslik
                    
                    # Toplam = Şantiye başlık + dış görevler
                    proje_analizleri[proje_adi]['toplam'] = santiye_baslik + dis_gorev_toplam_count
                    
                elif personel_dagilimi:
                    staff_count = personel_dagilimi.get('staff', 0)
                    calisan_count = personel_dagilimi.get('calisan', 0)
                    mobilizasyon_count = personel_dagilimi.get('mobilizasyon', 0)
                    ambarci_count = personel_dagilimi.get('ambarci', 0)
                    izinli_count = personel_dagilimi.get('izinli', 0)
                    dis_gorev_toplam_count = personel_dagilimi.get('dis_gorev_toplam', 0)
                    
                    proje_analizleri[proje_adi]['staff'] += staff_count
                    proje_analizleri[proje_adi]['calisan'] += calisan_count
                    proje_analizleri[proje_adi]['mobilizasyon'] += mobilizasyon_count
                    proje_analizleri[proje_adi]['ambarci'] += ambarci_count
                    proje_analizleri[proje_adi]['izinli'] += izinli_count
                    proje_analizleri[proje_adi]['dis_gorev_toplam'] += dis_gorev_toplam_count
                    
                    # KRİTİK GÜNCELLEME: Şantiye başlık hesaplaması (dış görevler HARİÇ)
                    santiye_baslik = staff_count + calisan_count + mobilizasyon_count + ambarci_count + izinli_count
                    proje_analizleri[proje_adi]['santiye_baslik'] += santiye_baslik
                    
                    # Toplam = Şantiye başlık + dış görevler
                    proje_analizleri[proje_adi]['toplam'] = santiye_baslik + dis_gorev_toplam_count
                    
            except Exception as e:
                logging.error(f"Proje analiz hatası: {e}")
                continue
        
        # KRİTİK DÜZELTME: Genel toplamları doğru hesapla - TÜM KATEGORİLERİN TOPLAMI
        genel_staff = sum(proje['staff'] for proje in proje_analizleri.values())
        genel_calisan = sum(proje['calisan'] for proje in proje_analizleri.values())
        genel_mobilizasyon = sum(proje['mobilizasyon'] for proje in proje_analizleri.values())
        genel_ambarci = sum(proje['ambarci'] for proje in proje_analizleri.values())
        genel_izinli = sum(proje['izinli'] for proje in proje_analizleri.values())
        genel_dis_gorev_toplam = sum(proje['dis_gorev_toplam'] for proje in proje_analizleri.values())
        
        # GENEL TOPLAM = Tüm kategorilerin toplamı
        genel_toplam = genel_staff + genel_calisan + genel_mobilizasyon + genel_ambarci + genel_izinli + genel_dis_gorev_toplam
        
        # TÜM SABİT ŞANTİYELERİ DAHİL ET, OPSİYONEL HARİÇ
        tum_santiyeler = set(SABIT_SANTIYELER).union(set(santiye for santiye in santiye_sorumlulari.keys() if santiye != "TÜMÜ"))
        tum_santiyeler = tum_santiyeler - set(OPSIYONEL_SANTIYELER)
        rapor_veren_santiyeler = set(proje_analizleri.keys())
        eksik_santiyeler = [s for s in (tum_santiyeler - rapor_veren_santiyeler) if s not in ["Belli değil", "Tümü"]]
        
        mesaj = f"📈 HAFTALIK ÖZET RAPOR\n"
        mesaj += f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}\n\n"
        
        mesaj += f"📊 GENEL İSTATİSTİKLER:\n"
        mesaj += f"• Toplam Rapor: {toplam_rapor}\n"
        mesaj += f"• Rapor Gönderen: {len(rows)} kişi\n"
        mesaj += f"• İş Günü: {gun_sayisi} gün\n"
        mesaj += f"• Toplam Personel: {genel_toplam} kişi\n\n"
        
        mesaj += f"🏗️ PROJE BAZLI PERSONEL:\n\n"
        
        onemli_projeler = ["SKP", "LOT13", "LOT71", "STADYUM", "BWC", "DMC", "YHP", "TYM", "MMP", "RMC", "PİRAMİT", "MOS", ]
        for proje_adi, analiz in sorted(proje_analizleri.items(), key=lambda x: x[1]['toplam'], reverse=True):
            if proje_adi in onemli_projeler and (analiz['santiye_baslik'] > 0 or analiz['dis_gorev_toplam'] > 0):
                mesaj += f"🏗️ {proje_adi}: {analiz['santiye_baslik']} kişi"
                if analiz['dis_gorev_toplam'] > 0:
                    mesaj += f" (Dış görev: {analiz['dis_gorev_toplam']})"
                mesaj += "\n"
                mesaj += f"   └─ Staff:{analiz['staff']}, Çalışan:{analiz['calisan']}, Mobilizasyon:{analiz['mobilizasyon']}, Ambarcı:{analiz['ambarci']}, İzinli:{analiz['izinli']}\n\n"
        
        for proje_adi, analiz in sorted(proje_analizleri.items(), key=lambda x: x[1]['toplam'], reverse=True):
            if proje_adi not in onemli_projeler and (analiz['santiye_baslik'] > 0 or analiz['dis_gorev_toplam'] > 0):
                emoji = "🏢" if proje_adi == "TYM" else "🏗️"
                mesaj += f"{emoji} {proje_adi}: {analiz['santiye_baslik']} kişi"
                if analiz['dis_gorev_toplam'] > 0:
                    mesaj += f" (Dış görev: {analiz['dis_gorev_toplam']})"
                mesaj += "\n"
                
                detay = []
                if analiz['staff'] > 0: detay.append(f"Staff:{analiz['staff']}")
                if analiz['calisan'] > 0: detay.append(f"Çalışan:{analiz['calisan']}")
                if analiz['mobilizasyon'] > 0: detay.append(f"Mobilizasyon:{analiz['mobilizasyon']}")
                if analiz['ambarci'] > 0: detay.append(f"Ambarcı:{analiz['ambarci']}")
                if analiz['izinli'] > 0: detay.append(f"İzinli:{analiz['izinli']}")
                
                if detay:
                    mesaj += f"   └─ {', '.join(detay)}\n"
        
        # KRİTİK GÜNCELLEME: Genel toplam = Σ(tüm kategorilerin toplamı)
        mesaj += f"\n📈 GENEL TOPLAM: {genel_toplam} kişi\n"
        
        if genel_toplam > 0:
            # DAĞILIM YÜZDELERİNİ DÜZELT - TOPLAM PERSONEL ÜZERİNDEN HESAPLA
            mesaj += f"🎯 DAĞILIM:\n"
            if genel_staff > 0:
                mesaj += f"• Staff: {genel_staff} (%{genel_staff/genel_toplam*100:.1f})\n"
            if genel_calisan > 0:
                mesaj += f"• Çalışan: {genel_calisan} (%{genel_calisan/genel_toplam*100:.1f})\n"
            if genel_mobilizasyon > 0:
                mesaj += f"• Mobilizasyon: {genel_mobilizasyon} (%{genel_mobilizasyon/genel_toplam*100:.1f})\n"
            if genel_ambarci > 0:
                mesaj += f"• Ambarcı: {genel_ambarci} (%{genel_ambarci/genel_toplam*100:.1f})\n"
            if genel_izinli > 0:
                mesaj += f"• İzinli: {genel_izinli} (%{genel_izinli/genel_toplam*100:.1f})\n"
            if genel_dis_gorev_toplam > 0:
                mesaj += f"• Dış Görev: {genel_dis_gorev_toplam} (%{genel_dis_gorev_toplam/genel_toplam*100:.1f})\n"
        
        if eksik_santiyeler:
            mesaj += f"\n❌ EKSİK ŞANTİYELER: {', '.join(sorted(eksik_santiyeler))}"
        
        mesaj += "\n\n📝 Lütfen eksiksiz rapor paylaşımına devam edelim. Teşekkürler."
        
        return mesaj
    except Exception as e:
        return f"❌ Haftalık rapor oluşturulurken hata: {e}"

# AYLIK RAPOR FONKSİYONU - VERİMLİLİK KALDIRILDI
async def generate_aylik_rapor_mesaji(start_date, end_date):
    try:
        rows = await async_fetchall("""
            SELECT user_id, COUNT(*) as rapor_sayisi
            FROM reports 
            WHERE report_date BETWEEN %s AND %s
            GROUP BY user_id
            ORDER BY rapor_sayisi DESC
        """, (start_date, end_date))
        
        if not rows:
            return f"📭 {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')} arasında rapor bulunamadı."
        
        toplam_rapor = sum([safe_get_tuple_value(x, 1, 0) for x in rows])
        gun_sayisi = (end_date - start_date).days + 1
        
        proje_detay_rows = await async_fetchall("""
            SELECT project_name, ai_analysis
            FROM reports 
            WHERE report_date BETWEEN %s AND %s AND project_name IS NOT NULL AND project_name != 'BELİRSİZ'
        """, (start_date, end_date))
        
        proje_analizleri = {}
        
        for row in proje_detay_rows:
            if len(row) < 2:
                continue
                
            proje_adi = safe_get_tuple_value(row, 0, '')
            ai_analysis = safe_get_tuple_value(row, 1, '{}')
            
            # PROJE ADINI NORMALİZE ET
            proje_adi = normalize_site_name(proje_adi)
            
            # TÜMÜ şantiyesini filtrele
            if not proje_adi or proje_adi == "TÜMÜ":
                continue
                
            if proje_adi not in proje_analizleri:
                proje_analizleri[proje_adi] = {
                    'staff': 0, 'calisan': 0, 'mobilizasyon': 0, 'ambarci': 0, 'izinli': 0, 'dis_gorev_toplam': 0, 
                    'toplam': 0, 'santiye_baslik': 0  # YENİ: Şantiye başlık sayısı
                }
            
            try:
                ai_data = safe_json_loads(ai_analysis)
                yeni_format = ai_data.get('yeni_sabit_format', {})
                personel_dagilimi = ai_data.get('personel_dagilimi', {})
                
                if yeni_format:
                    staff_count = yeni_format.get('staff', 0)
                    calisan_count = yeni_format.get('calisan', 0)
                    mobilizasyon_count = yeni_format.get('mobilizasyon', 0)
                    ambarci_count = yeni_format.get('ambarci', 0)
                    izinli_count = yeni_format.get('izinli', 0)
                    dis_gorev_toplam_count = yeni_format.get('dis_gorev_toplam', 0)
                    
                    proje_analizleri[proje_adi]['staff'] += staff_count
                    proje_analizleri[proje_adi]['calisan'] += calisan_count
                    proje_analizleri[proje_adi]['mobilizasyon'] += mobilizasyon_count
                    proje_analizleri[proje_adi]['ambarci'] += ambarci_count
                    proje_analizleri[proje_adi]['izinli'] += izinli_count
                    proje_analizleri[proje_adi]['dis_gorev_toplam'] += dis_gorev_toplam_count
                    
                    # KRİTİK GÜNCELLEME: Şantiye başlık hesaplaması (dış görevler HARİÇ)
                    santiye_baslik = staff_count + calisan_count + mobilizasyon_count + ambarci_count + izinli_count
                    proje_analizleri[proje_adi]['santiye_baslik'] += santiye_baslik
                    
                    # Toplam = Şantiye başlık + dış görevler
                    proje_analizleri[proje_adi]['toplam'] = santiye_baslik + dis_gorev_toplam_count
                    
                elif personel_dagilimi:
                    staff_count = personel_dagilimi.get('staff', 0)
                    calisan_count = personel_dagilimi.get('calisan', 0)
                    mobilizasyon_count = personel_dagilimi.get('mobilizasyon', 0)
                    ambarci_count = personel_dagilimi.get('ambarci', 0)
                    izinli_count = personel_dagilimi.get('izinli', 0)
                    dis_gorev_toplam_count = personel_dagilimi.get('dis_gorev_toplam', 0)
                    
                    proje_analizleri[proje_adi]['staff'] += staff_count
                    proje_analizleri[proje_adi]['calisan'] += calisan_count
                    proje_analizleri[proje_adi]['mobilizasyon'] += mobilizasyon_count
                    proje_analizleri[proje_adi]['ambarci'] += ambarci_count
                    proje_analizleri[proje_adi]['izinli'] += izinli_count
                    proje_analizleri[proje_adi]['dis_gorev_toplam'] += dis_gorev_toplam_count
                    
                    # KRİTİK GÜNCELLEME: Şantiye başlık hesaplaması (dış görevler HARİÇ)
                    santiye_baslik = staff_count + calisan_count + mobilizasyon_count + ambarci_count + izinli_count
                    proje_analizleri[proje_adi]['santiye_baslik'] += santiye_baslik
                    
                    # Toplam = Şantiye başlık + dış görevler
                    proje_analizleri[proje_adi]['toplam'] = santiye_baslik + dis_gorev_toplam_count
                    
            except Exception as e:
                logging.error(f"Proje analiz hatası: {e}")
                continue
        
        # KRİTİK DÜZELTME: Genel toplamları doğru hesapla - TÜM KATEGORİLERİN TOPLAMI
        genel_staff = sum(proje['staff'] for proje in proje_analizleri.values())
        genel_calisan = sum(proje['calisan'] for proje in proje_analizleri.values())
        genel_mobilizasyon = sum(proje['mobilizasyon'] for proje in proje_analizleri.values())
        genel_ambarci = sum(proje['ambarci'] for proje in proje_analizleri.values())
        genel_izinli = sum(proje['izinli'] for proje in proje_analizleri.values())
        genel_dis_gorev_toplam = sum(proje['dis_gorev_toplam'] for proje in proje_analizleri.values())
        
        # GENEL TOPLAM = Tüm kategorilerin toplamı
        genel_toplam = genel_staff + genel_calisan + genel_mobilizasyon + genel_ambarci + genel_izinli + genel_dis_gorev_toplam
        
        # TÜM SABİT ŞANTİYELERİ DAHİL ET, OPSİYONEL HARİÇ
        tum_santiyeler = set(SABIT_SANTIYELER).union(set(santiye for santiye in santiye_sorumlulari.keys() if santiye != "TÜMÜ"))
        tum_santiyeler = tum_santiyeler - set(OPSIYONEL_SANTIYELER)
        rapor_veren_santiyeler = set(proje_analizleri.keys())
        eksik_santiyeler = [s for s in (tum_santiyeler - rapor_veren_santiyeler) if s not in ["Belli değil", "Tümü"]]
        
        mesaj = f"🗓️ AYLIK ÖZET RAPOR\n"
        mesaj += f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}\n\n"
        
        mesaj += f"📈 PERFORMANS ANALİZİ:\n"
        mesaj += f"• Toplam Rapor: {toplam_rapor}\n"
        mesaj += f"• İş Günü: {gun_sayisi} gün\n"
        mesaj += f"• Günlük Ort.: {toplam_rapor/gun_sayisi:.1f} rapor\n"
        mesaj += f"• Toplam Personel: {genel_toplam} kişi\n\n"
        
        mesaj += f"🏗️ PROJE BAZLI PERSONEL:\n\n"
        
        onemli_projeler = ["SKP", "LOT13", "LOT71", "BWC", "DMC", "YHP", "TYM", "MMP", "RMC", "PİRAMİT", "MOS"]
        for proje_adi, analiz in sorted(proje_analizleri.items(), key=lambda x: x[1]['toplam'], reverse=True):
            if proje_adi in onemli_projeler and (analiz['santiye_baslik'] > 0 or analiz['dis_gorev_toplam'] > 0):
                mesaj += f"🏗️ {proje_adi}: {analiz['santiye_baslik']} kişi"
                if analiz['dis_gorev_toplam'] > 0:
                    mesaj += f" (Dış görev: {analiz['dis_gorev_toplam']})"
                mesaj += "\n"
                mesaj += f"   └─ Staff:{analiz['staff']}, Çalışan:{analiz['calisan']}, Mobilizasyon:{analiz['mobilizasyon']}, Ambarcı:{analiz['ambarci']}, İzinli:{analiz['izinli']}\n\n"
        
        for proje_adi, analiz in sorted(proje_analizleri.items(), key=lambda x: x[1]['toplam'], reverse=True):
            if proje_adi not in onemli_projeler and (analiz['santiye_baslik'] > 0 or analiz['dis_gorev_toplam'] > 0):
                emoji = "🏢" if proje_adi == "TYM" else "🏗️"
                mesaj += f"{emoji} {proje_adi}: {analiz['santiye_baslik']} kişi"
                if analiz['dis_gorev_toplam'] > 0:
                    mesaj += f" (Dış görev: {analiz['dis_gorev_toplam']})"
                mesaj += "\n"
                
                detay = []
                if analiz['staff'] > 0: detay.append(f"Staff:{analiz['staff']}")
                if analiz['calisan'] > 0: detay.append(f"Çalışan:{analiz['calisan']}")
                if analiz['mobilizasyon'] > 0: detay.append(f"Mobilizasyon:{analiz['mobilizasyon']}")
                if analiz['ambarci'] > 0: detay.append(f"Ambarcı:{analiz['ambarci']}")
                if analiz['izinli'] > 0: detay.append(f"İzinli:{analiz['izinli']}")
                
                if detay:
                    mesaj += f"   └─ {', '.join(detay)}\n"
        
        # KRİTİK GÜNCELLEME: Genel toplam = Σ(tüm kategorilerin toplamı)
        mesaj += f"\n📈 GENEL TOPLAM: {genel_toplam} kişi\n"
        
        if genel_toplam > 0:
            # DAĞILIM YÜZDELERİNİ DÜZELT - TOPLAM PERSONEL ÜZERİNDEN HESAPLA
            mesaj += f"🎯 DAĞILIM:\n"
            if genel_staff > 0:
                mesaj += f"• Staff: {genel_staff} (%{genel_staff/genel_toplam*100:.1f})\n"
            if genel_calisan > 0:
                mesaj += f"• Çalışan: {genel_calisan} (%{genel_calisan/genel_toplam*100:.1f})\n"
            if genel_mobilizasyon > 0:
                mesaj += f"• Mobilizasyon: {genel_mobilizasyon} (%{genel_mobilizasyon/genel_toplam*100:.1f})\n"
            if genel_ambarci > 0:
                mesaj += f"• Ambarcı: {genel_ambarci} (%{genel_ambarci/genel_toplam*100:.1f})\n"
            if genel_izinli > 0:
                mesaj += f"• İzinli: {genel_izinli} (%{genel_izinli/genel_toplam*100:.1f})\n"
            if genel_dis_gorev_toplam > 0:
                mesaj += f"• Dış Görev: {genel_dis_gorev_toplam} (%{genel_dis_gorev_toplam/genel_toplam*100:.1f})\n"
        
        if eksik_santiyeler:
            mesaj += f"\n❌ EKSİK ŞANTİYELER: {', '.join(sorted(eksik_santiyeler))}"
        
        mesaj += "\n\n📝 Lütfen eksiksiz rapor paylaşımına devam edelim. Teşekkürler."
        
        return mesaj
        
    except Exception as e:
        return f"❌ Aylık rapor oluşturulurken hata: {e}"

async def generate_tarih_araligi_raporu(start_date, end_date):
    try:
        rows = await async_fetchall("""
            SELECT user_id, COUNT(*) as rapor_sayisi
            FROM reports 
            WHERE report_date BETWEEN %s AND %s
            GROUP BY user_id
            ORDER BY rapor_sayisi DESC
        """, (start_date, end_date))
        
        if not rows:
            return f"📭 {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')} arasında rapor bulunamadı."
        
        toplam_rapor = sum([safe_get_tuple_value(x, 1, 0) for x in rows])
        gun_sayisi = (end_date - start_date).days + 1
        
        personel_result = await async_fetchone("""
            SELECT COALESCE(SUM(person_count), 0) as toplam_kisi
            FROM reports 
            WHERE report_date BETWEEN %s AND %s AND report_type = 'RAPOR'
        """, (start_date, end_date))
        
        toplam_personel = safe_get_tuple_value(personel_result, 0, 0)
        
        mesaj = f"📅 TARİH ARALIĞI RAPORU\n"
        mesaj += f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}\n\n"
        
        mesaj += f"📊 GENEL İSTATİSTİKLER:\n"
        mesaj += f"• Toplam Rapor: {toplam_rapor}\n"
        mesaj += f"• Rapor Gönderen: {len(rows)} kişi\n"
        mesaj += f"• Gün Sayısı: {gun_sayisi} gün\n"
        mesaj += f"• Günlük Ort.: {toplam_rapor/gun_sayisi:.1f} rapor\n"
        mesaj += f"• Toplam Personel: {toplam_personel} kişi\n\n"
        
        return mesaj
    except Exception as e:
        return f"❌ Tarih aralığı raporu oluşturulurken hata: {e}"

async def eksikraporlar_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return
    
    try:
        bugun = dt.datetime.now(TZ).date()
        durum = await get_santiye_bazli_rapor_durumu(bugun)
        
        mesaj = f"📊 EKSİK RAPORLAR - {bugun.strftime('%d.%m.%Y')}\n\n"
        
        if not durum['eksik_santiyeler']:
            mesaj += "🎉 Tüm şantiyeler raporlarını iletti! Harika iş!"
        else:
            mesaj += f"❌ Rapor İletilmeyen Şantiyeler ({len(durum['eksik_santiyeler'])}):\n\n"
            
            for santiye in sorted(durum['eksik_santiyeler']):
                if santiye in ["Belli değil", "Tümü"]:
                    continue
                sorumlular = santiye_sorumlulari.get(santiye, [])
                mesaj += f"🏗️ {santiye}\n\n"
        
        if durum['rapor_veren_santiyeler']:
            mesaj += f"✅ Rapor İleten Şantiyeler ({len(durum['rapor_veren_santiyeler'])}):\n"
            for santiye in sorted(durum['rapor_veren_santiyeler']):
                mesaj += f"• {santiye}\n"
        
        await update.message.reply_text(mesaj)
        
    except Exception as e:
        await update.message.reply_text(f"❌ Eksik raporlar kontrol edilirken hata: {e}")

async def istatistik_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return
    
    try:
        bugun = dt.datetime.now(TZ).date()
        bu_hafta_baslangic = bugun - dt.timedelta(days=bugun.weekday())
        bu_ay_baslangic = bugun.replace(day=1)
        
        bugun_rapor = await async_fetchone("SELECT COUNT(*) FROM reports WHERE report_date = %s", (bugun,))
        bugun_rapor_sayisi = safe_get_tuple_value(bugun_rapor, 0, 0)
        
        hafta_rapor = await async_fetchone("""
            SELECT COUNT(*) FROM reports WHERE report_date BETWEEN %s AND %s
        """, (bu_hafta_baslangic, bugun))
        hafta_rapor_sayisi = safe_get_tuple_value(hafta_rapor, 0, 0)
        
        ay_rapor = await async_fetchone("""
            SELECT COUNT(*) FROM reports WHERE report_date BETWEEN %s AND %s
        """, (bu_ay_baslangic, bugun))
        ay_rapor_sayisi = safe_get_tuple_value(ay_rapor, 0, 0)
        
        toplam_rapor = await async_fetchone("SELECT COUNT(*) FROM reports")
        toplam_rapor_sayisi = safe_get_tuple_value(toplam_rapor, 0, 0)
        
        durum = await get_santiye_bazli_rapor_durumu(bugun)
        
        # SADECE AYLIK performans analizi
        aylik_performans = ""
        try:
            # Ayın tamamını analiz et
            ay_bitis = bugun  # Bugüne kadar
            ay_baslangic = bugun.replace(day=1)  # Ayın 1'i
            
            analiz_aylik, gunler_aylik = await analyze_missing_reports(ay_baslangic, ay_bitis)
            if analiz_aylik:
                aylik_performans = await generate_performans_analizi(analiz_aylik, ay_baslangic, ay_bitis, gunler_aylik, "aylik")
        except Exception as e:
            logging.error(f"Aylık performans analizi hatası: {e}")
            aylik_performans = "⚠️ Aylık performans analizi oluşturulamadı."
        
        mesaj = "📊 GENEL İSTATİSTİKLER \n\n"
        
        mesaj += "📅 GÜNLÜK İSTATİSTİKLER:\n"
        mesaj += f"• Bugünkü Rapor: {bugun_rapor_sayisi}\n"
        mesaj += f"• Bu Hafta: {hafta_rapor_sayisi}\n"
        mesaj += f"• Bu Ay: {ay_rapor_sayisi}\n"
        mesaj += f"• Toplam Rapor: {toplam_rapor_sayisi}\n\n"
        
        mesaj += f"🏗️ BUGÜNKÜ ŞANTİYE DURUMU :\n"
        mesaj += f"• Rapor İleten: {len(durum['rapor_veren_santiyeler'])}/{len(durum['tum_santiyeler'])}\n"
        
        toplam_santiye = len(durum['tum_santiyeler'])
        if toplam_santiye > 0:
            basari_orani = (len(durum['rapor_veren_santiyeler']) / toplam_santiye) * 100
            mesaj += f"• Başarı Oranı: %{basari_orani:.1f}\n\n"
        else:
            mesaj += "• Başarı Oranı: %0.0\n\n"
        
        # Aylık performans analizi ekle
        if aylik_performans:
            mesaj += f"{aylik_performans}"
        
        # Mesaj çok uzunsa böl
        if len(mesaj) > 4000:
            part1 = mesaj[:4000]
            part2 = mesaj[4000:]
            await update.message.reply_text(part1)
            await asyncio.sleep(0.5)
            await update.message.reply_text(part2)
        else:
            await update.message.reply_text(mesaj)
        
    except Exception as e:
        await update.message.reply_text(f"❌ İstatistikler oluşturulurken hata: {e}")

# EKSİK RAPOR ANALİZ FONKSİYONLARI
def parse_tr_date(date_str: str) -> dt.date:
    """GG.AA.YYYY formatındaki tarihi parse eder"""
    try:
        day, month, year = map(int, date_str.split('.'))
        return dt.date(year, month, day)
    except (ValueError, AttributeError):
        raise ValueError("Geçersiz tarih formatı. GG.AA.YYYY şeklinde olmalı.")

async def analyze_missing_reports(start_date: dt.date, end_date: dt.date) -> Tuple[Dict, List]:
    """Belirtilen tarih aralığındaki eksik raporları analiz eder - OPSİYONEL ŞANTİYELER HARİÇ"""
    try:
        tum_santiyeler = set(SABIT_SANTIYELER).union(
            set(santiye for santiye in santiye_sorumlulari.keys() if santiye != "TÜMÜ")
        )
        
        # OPSİYONEL ŞANTİYELERİ ÇIKAR (OHP gibi)
        tum_santiyeler = tum_santiyeler - set(OPSIYONEL_SANTIYELER)
        
        current_date = start_date
        gunler = []
        while current_date <= end_date:
            # 7/24 ÇALIŞMA SİSTEMİ: Hafta sonları da dahil ediliyor
            gunler.append(current_date)
            current_date += dt.timedelta(days=1)
        santiye_analiz = {}
        for santiye in tum_santiyeler:
            santiye_analiz[santiye] = {
                'toplam_gun': len(gunler),
                'rapor_verilen_gunler': 0,
                'eksik_gunler': [],
                'rapor_verilen_tarihler': []
            }
            for gun in gunler:
                rapor_var = await async_fetchone("""
                    SELECT EXISTS(
                        SELECT 1 FROM reports 
                        WHERE report_date = %s AND project_name = %s
                    )
                """, (gun, santiye))
                if rapor_var and safe_get_tuple_value(rapor_var, 0, False):
                    santiye_analiz[santiye]['rapor_verilen_gunler'] += 1
                    santiye_analiz[santiye]['rapor_verilen_tarihler'].append(gun)
                else:
                    santiye_analiz[santiye]['eksik_gunler'].append(gun)
        return santiye_analiz, gunler
    except Exception as e:
        logging.error(f"Eksik rapor analiz hatası: {e}")
        return {}, []

async def create_missing_reports_excel(analiz: Dict, start_date: dt.date, end_date: dt.date, gunler: List) -> str:
    """Eksik rapor analizini Excel formatında oluştur - GÜNCELLENMİŞ GÖRÜNÜM VE SIRALAMA"""
    try:
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Eksik Rapor Analizi"
        
        headers = ['Şantiye', 'Toplam Gün', 'Eksik Gün', 'Eksik %'] + [gun.strftime('%d.%m') for gun in gunler]
        
        # DİNAMİK BAŞLIK BİRLEŞTİRME - sütun sayısına göre
        last_column_letter = get_column_letter(len(headers))
        ws.merge_cells(f'A1:{last_column_letter}1')
        ws['A1'] = f"Eksik Rapor Analizi - {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Başlıktan sonra bir boş satır
        ws.row_dimensions[2].height = 15
        
        # Başlık satırı (3. satır)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # İnce kenarlık
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
        
        # YENİ: ŞANTİYELERİ SORUMLUYA GÖRE SIRALA
        santiye_sirali_liste = []
        
        for santiye, a in analiz.items():
            # Şantiye için sorumlu bilgisini al
            sorumlu_adi = ""
            if santiye in santiye_sorumlulari and santiye_sorumlulari[santiye]:
                # İlk sorumluyu al
                sorumlu_id = santiye_sorumlulari[santiye][0]
                sorumlu_adi = id_to_name.get(sorumlu_id, "")
            
            santiye_sirali_liste.append((sorumlu_adi, santiye, a))
        
        # Önce sorumlu adına göre, sonra şantiye adına göre sırala
        santiye_sirali_liste.sort(key=lambda x: (x[0].lower() if x[0] else "", x[1].lower()))
        
        row = 4
        for sorumlu_adi, santiye, a in santiye_sirali_liste:
            # Şantiye adı
            ws.cell(row=row, column=1, value=santiye)
            
            # Sayısal değerler
            ws.cell(row=row, column=2, value=a['toplam_gun'])
            ws.cell(row=row, column=3, value=len(a['eksik_gunler']))
            
            # Yüzde değeri
            eksik_yuzde = (len(a['eksik_gunler']) / a['toplam_gun']) * 100 if a['toplam_gun'] > 0 else 0
            ws.cell(row=row, column=4, value=eksik_yuzde/100)
            ws.cell(row=row, column=4).number_format = '0.00%'
            
            # YENİ: RENKLENDİRME MANTIĞI
            # "Eksik %" değerine göre renk belirle
            eksik_yuzde_hucresi = ws.cell(row=row, column=4)
            
            if eksik_yuzde <= 10:
                # %0-10: SADECE "Eksik %" sütunu yeşil
                eksik_yuzde_hucresi.fill = PatternFill(start_color="63BE7B", end_color="63BE7B", fill_type="solid")
                # "Şantiye", "Toplam Gün", "Eksik Gün" sütunları BEYAZ kalacak (renk yok)
                
            elif eksik_yuzde <= 40:
                # %11-40: Sarı renk TÜM 4 sütuna uygulanacak
                renk_kodu = "FFEB84"
                sütunlar = [1, 2, 3, 4]  # Şantiye, Toplam Gün, Eksik Gün, Eksik %
                for sütun in sütunlar:
                    cell = ws.cell(row=row, column=sütun)
                    cell.fill = PatternFill(start_color=renk_kodu, end_color=renk_kodu, fill_type="solid")
            else:
                # %41-100: Kırmızı renk TÜM 4 sütuna uygulanacak
                renk_kodu = "F8696B"
                sütunlar = [1, 2, 3, 4]  # Şantiye, Toplam Gün, Eksik Gün, Eksik %
                for sütun in sütunlar:
                    cell = ws.cell(row=row, column=sütun)
                    cell.fill = PatternFill(start_color=renk_kodu, end_color=renk_kodu, fill_type="solid")
            
            # Günlük durumlar (✓/✗)
            for col_idx, gun in enumerate(gunler, 5):
                cell = ws.cell(row=row, column=col_idx)
                if gun in a['eksik_gunler']:
                    cell.value = '✗'
                    # Eksik günler kırmızı arka plan
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                else:
                    cell.value = '✓'
                    # Rapor verilen günler yeşil arka plan
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                
                # Kenarlık ve hizalama
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=11)
            
            # Tüm hücrelere kenarlık ve hizalama ekle
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                # Kenarlığı her zaman uygula
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(size=11)
            
            row += 1
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 20  # Şantiye
        ws.column_dimensions['B'].width = 12  # Toplam Gün
        ws.column_dimensions['C'].width = 12  # Eksik Gün
        ws.column_dimensions['D'].width = 12  # Eksik %
        
        # Tarih sütunları için genişlik
        for i in range(len(gunler)):
            col_letter = get_column_letter(5 + i)
            ws.column_dimensions[col_letter].width = 8
        
        # Satır yükseklikleri
        for r in range(3, row + 1):
            ws.row_dimensions[r].height = 25
        
        # Özet sayfası oluştur
        ws_summary = wb.create_sheet("Özet")
        
        # Özet başlığı
        ws_summary.merge_cells('A1:B1')
        ws_summary['A1'] = f"Eksik Rapor Özeti - {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
        ws_summary['A1'].font = Font(bold=True, size=14, color="366092")
        ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Özet bilgileri
        toplam_santiye = len(analiz)
        eksiksiz_santiye = sum(1 for a in analiz.values() if len(a['eksik_gunler']) == 0)
        eksik_santiye = toplam_santiye - eksiksiz_santiye
        toplam_eksik_rapor = sum(len(a['eksik_gunler']) for a in analiz.values())
        toplam_gun = len(gunler)
        
        summary_data = [
            ['📅 Analiz Periyodu', f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"],
            ['🏗️ Toplam Şantiye', toplam_santiye],
            ['✅ Eksiksiz Şantiye', f"{eksiksiz_santiye} (%{eksiksiz_santiye/toplam_santiye*100:.1f})"],
            ['❌ Eksik Raporu Olan', f"{eksik_santiye} (%{eksik_santiye/toplam_santiye*100:.1f})"],
            ['📅 Toplam Gün', toplam_gun],
            ['📊 Toplam EKSİK RAPOR', toplam_eksik_rapor],
            ['🕒 Oluşturulma', dt.datetime.now(TZ).strftime('%d.%m.%Y %H:%M')]
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 3):
            cell_label = ws_summary[f'A{row_idx}']
            cell_value = ws_summary[f'B{row_idx}']
            
            cell_label.value = label
            cell_value.value = value
            
            # Kenarlık
            for cell in [cell_label, cell_value]:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
            
            # Kalın font
            cell_label.font = Font(bold=True, size=11)
            cell_value.font = Font(size=11)
            
            # Satır yüksekliği
            ws_summary.row_dimensions[row_idx].height = 30
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 30
        
        # İlk sayfayı aktif yap
        wb.active = wb["Eksik Rapor Analizi"]
        
        # Dosyayı kaydet
        timestamp = dt.datetime.now(TZ).strftime("%Y%m%d_%H%M%S")
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        
        logging.info(f"✅ Excel raporu oluşturuldu: {temp_file.name}")
        return temp_file.name
        
    except Exception as e:
        logging.error(f"❌ Eksik rapor Excel oluşturma hatası: {e}")
        raise e

def format_missing_reports_message(analiz: Dict, start_date: dt.date, end_date: dt.date, gunler: List) -> str:
    mesaj = f"📋 EKSİK RAPOR DETAY ANALİZİ\n"
    mesaj += f"📅 {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}\n"
    mesaj += "Not: 'Eksik günler:' olarak tarihler yazılmıştır.\n\n"
    
    # DÜZELTİLDİ: Sadece eksik raporu olan şantiyeleri listeye ekle
    eksik_santiyeler = []
    
    for santiye, a in analiz.items():
        if len(a['eksik_gunler']) == 0:  # Eksik raporu yoksa atla
            continue
            
        eksik_yuzde = (len(a['eksik_gunler']) / a['toplam_gun']) * 100 if a['toplam_gun'] > 0 else 0
        eksik_santiyeler.append((santiye, a, eksik_yuzde))
    
    # Eksik şantiyeleri yüzdeye göre sırala
    eksik_santiyeler.sort(key=lambda x: x[2], reverse=True)
    
    if eksik_santiyeler:
        mesaj += f"🔴 EKSİK RAPORU OLAN ŞANTİYELER ({len(eksik_santiyeler)}):\n"
        for santiye, a, yuzde in eksik_santiyeler:
            mesaj += f"• {santiye}: {len(a['eksik_gunler'])}/{a['toplam_gun']} gün (%{yuzde:.1f})\n"
            eksik_gunler_str = ", ".join([gun.strftime('%d') for gun in a['eksik_gunler']])
            mesaj += f"  └─ Eksik günler: {eksik_gunler_str}\n\n"
    else:
        mesaj += "✅ Tüm şantiyeler eksiksiz rapor vermiş!\n\n"
    
    # Özet istatistikler
    toplam_santiye = len(analiz)
    eksiksiz_santiye = sum(1 for a in analiz.values() if len(a['eksik_gunler']) == 0)
    eksik_santiye = toplam_santiye - eksiksiz_santiye
    toplam_eksik_rapor = sum(len(a['eksik_gunler']) for a in analiz.values())
    
    mesaj += f"📊 ÖZET:\n"
    mesaj += f"• Toplam Şantiye: {toplam_santiye}\n"
    mesaj += f"• Eksiksiz Şantiye: {eksiksiz_santiye} (%{eksiksiz_santiye/toplam_santiye*100:.1f})\n"
    mesaj += f"• Eksik Raporu Olan: {eksik_santiye} (%{eksik_santiye/toplam_santiye*100:.1f})\n"
    mesaj += f"• İş Günü: {len(gunler)} gün\n"
    mesaj += f"• Toplam EKSİK RAPOR: {toplam_eksik_rapor}\n"
    
    return mesaj
# REVİZYON 6: PERFORMANS ANALİZİ FONKSİYONU (GÜNCELLENMİŞ - TÜM ŞANTİYELER)
async def generate_performans_analizi(analiz: Dict, start_date: dt.date, end_date: dt.date, gunler: List, period_type="aylik") -> str:
    """AYLIK performans analizi oluştur: En iyi 3 şantiye, en kötü 3 şantiye, hiç göndermeyenler"""
    try:
        # TÜM şantiyeleri analiz et (aktif/pasif ayrımı YOK)
        # Not: OPSİYONEL şantiyeler (DATA CENTR, OHP) hariç tutulur
        
        # OPSİYONEL şantiyeleri filtrele
        filtrelenmis_analiz = {}
        for santiye, a in analiz.items():
            if santiye in OPSIYONEL_SANTIYELER:
                continue  # Opsiyonel şantiyeleri atla
            filtrelenmis_analiz[santiye] = a
        
        if not filtrelenmis_analiz:
            return "⚠️ Analiz edilecek şantiye bulunamadı."
        
        # Performans hesaplama
        performans_listesi = []
        for santiye, a in filtrelenmis_analiz.items():
            eksik_gun = len(a['eksik_gunler'])
            toplam_gun = a['toplam_gun']
            performans_listesi.append((santiye, eksik_gun, toplam_gun))
        
        # Eksik güne göre sırala (en az eksik = en iyi)
        performans_listesi.sort(key=lambda x: x[1])
        
        mesaj = ""
        
        # SADECE AYLIK ANALİZ
        mesaj += f"📅 AYLIK PERFORMANS ANALİZİ\n"
        mesaj += f"📆 {start_date.strftime('%d.%m')}-{end_date.strftime('%d.%m.%Y')} ({len(gunler)} gün)\n\n"
        
        # EN İYİ 3 ŞANTİYE (eksiksiz veya en az eksik)
        eksiksiz_santiyeler = [p for p in performans_listesi if p[1] == 0]
        en_iyi_liste = eksiksiz_santiyeler[:3] if len(eksiksiz_santiyeler) >= 3 else eksiksiz_santiyeler
        
        if en_iyi_liste:
            mesaj += f"🏆 EN İYİ PERFORMANS – TOP {len(en_iyi_liste)}:\n"
            for santiye, eksik, toplam in en_iyi_liste:
                mesaj += f"• {santiye}: {eksik}/{toplam} gün\n"
            mesaj += "\n"
        
        # EN KÖTÜ 3 ŞANTİYE (en çok eksik, ama %100 olmayanlar)
        # Önce %100 olmayanları filtrele
        kotu_adaylar = [p for p in performans_listesi if p[1] > 0 and p[1] < p[2]]  # 0 < eksik < toplam
        if kotu_adaylar:
            # Eksik sayısına göre ters sırala
            kotu_adaylar.sort(key=lambda x: x[1], reverse=True)
            en_kotu_liste = kotu_adaylar[:3]
            
            mesaj += f"⚠️ KÖTÜ PERFORMANSLI ŞANTİYELER – TOP {len(en_kotu_liste)}:\n"
            for santiye, eksik, toplam in en_kotu_liste:
                mesaj += f"• {santiye}: {eksik}/{toplam} gün\n"
            mesaj += "\n"
        
        # HİÇ RAPOR GÖNDERMEYENLER (%100 eksik)
        hic_gondermeyenler = [p for p in performans_listesi if p[1] == p[2] and p[2] > 0]  # eksik == toplam
        
        if hic_gondermeyenler:
            mesaj += f"🔴 HİÇ RAPOR GÖNDERMEYEN ŞANTİYELER ({len(hic_gondermeyenler)}):\n"
            for santiye, eksik, toplam in hic_gondermeyenler:
                mesaj += f"• {santiye}: {eksik}/{toplam} gün - HİÇ RAPOR YOK!\n"
        
        return mesaj  # ÖZET YOK
    except Exception as e:
        logging.error(f"Performans analizi hatası: {e}")
        return "⚠️ Performans analizi oluşturulamadı."

async def eksik_rapor_excel_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return

    if not context.args or len(context.args) != 2:
        await update.message.reply_text(
            "📋 EKSİK RAPOR EXCEL RAPORU\n\n"
            "Kullanım: `/eksik_rapor_excel 01.11.2026 30.11.2026`\n"
            "Belirtilen tarih aralığı için eksik rapor analizi Excel'i oluşturur."
        )
        return

    await update.message.reply_text("⌛ Eksik rapor analizi hazırlanıyor...")

    try:
        tarih1 = context.args[0].replace('/', '.').replace('-', '.')
        tarih2 = context.args[1].replace('/', '.').replace('-', '.')
        
        start_date = parse_tr_date(tarih1)
        end_date = parse_tr_date(tarih2)
        
        if start_date > end_date:
            await update.message.reply_text("❌ Başlangıç tarihi bitiş tarihinden büyük olamaz.")
            return

        analiz, gunler = await analyze_missing_reports(start_date, end_date)
        
        if not analiz:
            await update.message.reply_text("❌ Eksik rapor analizi oluşturulamadı.")
            return

        excel_dosyasi = await create_missing_reports_excel(analiz, start_date, end_date, gunler)
        mesaj = format_missing_reports_message(analiz, start_date, end_date, gunler)

        with open(excel_dosyasi, 'rb') as file:
            await update.message.reply_document(
                document=file,
                filename=f"Eksik_Rapor_Analizi_{start_date.strftime('%d.%m.%Y')}_{end_date.strftime('%d.%m.%Y')}.xlsx",
                caption=f"📊 Eksik Rapor Analizi: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
            )
        
        await update.message.reply_text(mesaj)
        os.unlink(excel_dosyasi)
        
    except Exception as e:
        await update.message.reply_text(f"❌ Eksik rapor analizi hatası: {e}")
        logging.error(f"Eksik rapor analizi hatası: {e}")

async def haftalik_eksik_raporlar_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return

    await update.message.reply_text("⌛ Haftalık eksik rapor analizi hazırlanıyor...")

    try:
        today = dt.datetime.now(TZ).date()
        start_date = today - dt.timedelta(days=6)  # 7 günlük periyot (bugün dahil)
        end_date = today

        analiz, gunler = await analyze_missing_reports(start_date, end_date)
        
        if not analiz:
            await update.message.reply_text("❌ Haftalık eksik rapor analizi oluşturulamadı.")
            return

        excel_dosyasi = await create_missing_reports_excel(analiz, start_date, end_date, gunler)
        mesaj = format_missing_reports_message(analiz, start_date, end_date, gunler)

        with open(excel_dosyasi, 'rb') as file:
            await update.message.reply_document(
                document=file,
                filename=f"Haftalik_Eksik_Rapor_Analizi_{start_date.strftime('%d.%m.%Y')}_{end_date.strftime('%d.%m.%Y')}.xlsx",
                caption=f"📊 Haftalık Eksik Rapor Analizi: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
            )
        
        await update.message.reply_text(mesaj)
        os.unlink(excel_dosyasi)
        
    except Exception as e:
        await update.message.reply_text(f"❌ Haftalık eksik rapor analizi hatası: {e}")
        logging.error(f"Haftalık eksik rapor analizi hatası: {e}")

async def aylik_eksik_raporlar_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return

    await update.message.reply_text("⌛ Aylık eksik rapor analizi hazırlanıyor...")

    try:
        today = dt.datetime.now(TZ).date()
        start_date = today.replace(day=1)
        end_date = today

        analiz, gunler = await analyze_missing_reports(start_date, end_date)
        
        if not analiz:
            await update.message.reply_text("❌ Aylık eksik rapor analizi oluşturulamadı.")
            return

        excel_dosyasi = await create_missing_reports_excel(analiz, start_date, end_date, gunler)
        mesaj = format_missing_reports_message(analiz, start_date, end_date, gunler)

        with open(excel_dosyasi, 'rb') as file:
            await update.message.reply_document(
                document=file,
                filename=f"Aylik_Eksik_Rapor_Analizi_{start_date.strftime('%d.%m.%Y')}_{end_date.strftime('%d.%m.%Y')}.xlsx",
                caption=f"📊 Aylık Eksik Rapor Analizi: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
            )
        
        await update.message.reply_text(mesaj)
        os.unlink(excel_dosyasi)
        
    except Exception as e:
        await update.message.reply_text(f"❌ Aylık eksik rapor analizi hatası: {e}")
        logging.error(f"Aylık eksik rapor analizi hatası: {e}")

# YENİ: HAFTALIK NORMAL RAPOR JOB FONKSİYONU
async def haftalik_normal_rapor_job(context: ContextTypes.DEFAULT_TYPE):
    """Her Pazar 09:00'da haftalık normal raporu gruba gönder"""
    try:
        today = dt.datetime.now(TZ).date()
        now_time = dt.datetime.now(TZ).time()
        
        # Sadece Pazar günü ve saat 09:00'da çalıştır
        if today.weekday() != 6:  # 0=Pazartesi, 6=Pazar
            return
        if not (8 <= now_time.hour <= 9):  # Saat 09:00 civarında
            return
            
        # Haftalık rapor tarih aralığı: Geçmiş 7 gün (bugün dahil değil)
        # Örnek: Pazar 09:00 gönderimi için Pazartesi 00:00 - Pazar 00:00 (7 gün)
        end_date = today - dt.timedelta(days=1)  # Dün (Cumartesi)
        start_date = end_date - dt.timedelta(days=6)  # 7 gün önce (Pazartesi)
        
        logging.info(f"📊 Haftalık normal rapor tetiklendi: {start_date} - {end_date}")
        
        mesaj = await generate_haftalik_rapor_mesaji(start_date, end_date)
        
        if GROUP_ID:
            try:
                await context.bot.send_message(chat_id=GROUP_ID, text=mesaj)
                logging.info(f"📊 Haftalık normal rapor gruba gönderildi: {start_date} - {end_date}")
            except Exception as e:
                logging.error(f"📊 Gruba haftalık normal rapor gönderilemedi: {e}")
        else:
            logging.error("📊 GROUP_ID ayarlanmamış, haftalık normal rapor gönderilemedi")
    except Exception as e:
        logging.error(f"📊 Haftalık normal rapor job hatası: {e}")

# YENİ: GÜNCELLENMİŞ HAFTALIK EKSİK RAPOR JOB FONKSİYONU
async def haftalik_eksik_rapor_job(context: ContextTypes.DEFAULT_TYPE):
    """Her Pazar 10:00'da haftalık eksik raporu gruba gönder"""
    try:
        today = dt.datetime.now(TZ).date()
        now_time = dt.datetime.now(TZ).time()
        
        # Sadece Pazar günü ve saat 10:00'da çalıştır
        if today.weekday() != 6:  # 0=Pazartesi, 6=Pazar
            return
        if not (9 <= now_time.hour <= 10):  # Saat 10:00 civarında
            return
        
        # Haftalık eksik rapor tarih aralığı: Haftalık normal raporla BİREBİR AYNI
        end_date = today - dt.timedelta(days=1)  # Dün (Cumartesi)
        start_date = end_date - dt.timedelta(days=6)  # 7 gün önce (Pazartesi)
        
        logging.info(f"📊 Haftalık eksik rapor tetiklendi: {start_date} - {end_date}")
        
        analiz, gunler = await analyze_missing_reports(start_date, end_date)
        
        if not analiz:
            logging.info("📊 Haftalık eksik rapor analizi: analiz yapılamadı.")
            return

        excel_dosyasi = await create_missing_reports_excel(analiz, start_date, end_date, gunler)
        mesaj = format_missing_reports_message(analiz, start_date, end_date, gunler)

        if GROUP_ID:
            try:
                with open(excel_dosyasi, 'rb') as file:
                    await context.bot.send_document(
                        chat_id=GROUP_ID,
                        document=file,
                        filename=f"Haftalik_Eksik_Rapor_Analizi_{start_date.strftime('%d.%m.%Y')}_{end_date.strftime('%d.%m.%Y')}.xlsx",
                        caption=f"📊 HAFTALIK EKSİK RAPOR ANALİZİ\n{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
                    )
                await context.bot.send_message(chat_id=GROUP_ID, text=mesaj)
                logging.info(f"📊 Haftalık eksik rapor analizi gruba gönderildi: {start_date} - {end_date}")
            except Exception as e:
                logging.error(f"📊 Gruba haftalık eksik rapor gönderilemedi: {e}")
        else:
            logging.error("📊 GROUP_ID ayarlanmamış, haftalık eksik rapor analizi gönderilemedi")

        os.unlink(excel_dosyasi)
    except Exception as e:
        logging.error(f"📊 Haftalık eksik rapor job hatası: {e}")

# YENİ: AYLIK NORMAL RAPOR JOB FONKSİYONU
async def aylik_normal_rapor_job(context: ContextTypes.DEFAULT_TYPE):
    """Her ayın 1'inde 08:30'da aylık normal raporu gruba gönder"""
    try:
        today = dt.datetime.now(TZ).date()
        now_time = dt.datetime.now(TZ).time()
        
        # Sadece ayın 1'inde ve saat 08:30'da çalıştır
        if today.day != 1:
            return
        if not (8 <= now_time.hour <= 9):  # Saat 08:30 civarında
            return
        
        # Aylık rapor tarih aralığı: Bir önceki takvim ayının TAMAMI
        # Örnek: 01.12.2026 08:30 gönderimi için 01.11.2026 - 30.11.2026
        end_date = today.replace(day=1) - dt.timedelta(days=1)  # Önceki ayın son günü
        start_date = end_date.replace(day=1)  # Önceki ayın 1'i
        
        logging.info(f"🗓️ Aylık normal rapor tetiklendi: {start_date} - {end_date}")
        
        mesaj = await generate_aylik_rapor_mesaji(start_date, end_date)
        
        if GROUP_ID:
            try:
                await context.bot.send_message(chat_id=GROUP_ID, text=mesaj)
                logging.info(f"🗓️ Aylık normal rapor gruba gönderildi: {start_date} - {end_date}")
            except Exception as e:
                logging.error(f"🗓️ Gruba aylık normal rapor gönderilemedi: {e}")
        else:
            logging.error("🗓️ GROUP_ID ayarlanmamış, aylık normal rapor gönderilemedi")
    except Exception as e:
        logging.error(f"🗓️ Aylık normal rapor job hatası: {e}")

# YENİ: GÜNCELLENMİŞ AYLIK EKSİK RAPOR JOB FONKSİYONU
async def aylik_eksik_rapor_job(context: ContextTypes.DEFAULT_TYPE):
    """Her ayın 1'inde 08:45'de aylık eksik raporu gruba gönder"""
    try:
        today = dt.datetime.now(TZ).date()
        now_time = dt.datetime.now(TZ).time()
        
        # Sadece ayın 1'inde ve saat 08:45'de çalıştır
        if today.day != 1:
            return
        if not (8 <= now_time.hour <= 9):  # Saat 08:45 civarında
            return
        
        # Aylık eksik rapor tarih aralığı: Aylık normal raporla BİREBİR AYNI
        end_date = today.replace(day=1) - dt.timedelta(days=1)  # Önceki ayın son günü
        start_date = end_date.replace(day=1)  # Önceki ayın 1'i
        
        logging.info(f"🗓️ Aylık eksik rapor tetiklendi: {start_date} - {end_date}")
        
        analiz, gunler = await analyze_missing_reports(start_date, end_date)
        
        if not analiz:
            logging.info("🗓️ Aylık eksik rapor analizi: analiz yapılamadı.")
            return

        excel_dosyasi = await create_missing_reports_excel(analiz, start_date, end_date, gunler)
        mesaj = format_missing_reports_message(analiz, start_date, end_date, gunler)

        if GROUP_ID:
            try:
                with open(excel_dosyasi, 'rb') as file:
                    await context.bot.send_document(
                        chat_id=GROUP_ID,
                        document=file,
                        filename=f"Aylik_Eksik_Rapor_Analizi_{start_date.strftime('%d.%m.%Y')}_{end_date.strftime('%d.%m.%Y')}.xlsx",
                        caption=f"🗓️ AYLIK EKSİK RAPOR ANALİZİ\n{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
                    )
                await context.bot.send_message(chat_id=GROUP_ID, text=mesaj)
                logging.info(f"🗓️ Aylık eksik rapor analizi gruba gönderildi: {start_date} - {end_date}")
            except Exception as e:
                logging.error(f"🗓️ Gruba aylık eksik rapor gönderilemedi: {e}")
        else:
            logging.error("🗓️ GROUP_ID ayarlanmamış, aylık eksik rapor analizi gönderilemedi")

        os.unlink(excel_dosyasi)
    except Exception as e:
        logging.error(f"🗓️ Aylık eksik rapor job hatası: {e}")

async def start_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🤖 Rapor Botu Aktif! \n\n"
        "Komutlar için `/info` yazın.\n\n"
        "📋 Temel Kullanım:\n"
        "• Rapor göndermek için direkt mesaj yazın\n"
        "• `/info` - Tüm komutları görüntüle\n"
        "• `/hakkinda` - Bot hakkında bilgi"
    )

async def info_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    user_name = update.message.from_user.first_name
    
    if is_admin(user_id):
        info_text = (
            f"🤖 Yapay Zeka Destekli Rapor Botu - ŞANTİYE BAZLI SİSTEM\n\n"
            f"👋 Hoş geldiniz {user_name}!\n\n"
            f"📋 Tüm Kullanıcılar İçin:\n"
            f"• Rapor göndermek için direkt mesaj yazın\n"
            f"`/start` - Botu başlat\n"
            f"`/info` - Komut bilgisi\n"
            f"`/hakkinda` - Bot hakkında\n\n"
            f"🛡️ Admin Komutları:\n"
            f"`/bugun` - Bugünün özeti (Admin)\n"
            f"`/dun` - Dünün özeti (Admin)\n"
            f"`/eksikraporlar` - Eksik raporları listele (Admin)\n"
            f"`/istatistik` - Genel istatistikler (Admin)\n"
            f"`/haftalik_rapor` - Haftalık rapor (Admin)\n"
            f"`/aylik_rapor` - Aylık rapor (Admin)\n"
            f"`/tariharaligi` - Tarih aralığı raporu (Admin)\n"
            f"`/haftalik_istatistik` - Haftalık istatistik (Admin)\n"
            f"`/aylik_istatistik` - Aylık istatistik (Admin)\n"
            f"`/excel_tariharaligi` - Excel raporu (Admin)\n"
            f"`/maliyet` - Maliyet analizi (Admin)\n"
            f"`/ai_rapor` - Detaylı AI raporu (Admin)\n"
            f"`/kullanicilar` - Tüm kullanıcı listesi (Admin)\n"
            f"`/santiyeler` - Şantiye listesi (Admin)\n"
            f"`/santiye_durum` - Şantiye rapor durumu (Admin)\n"
            f"`/eksik_rapor_excel` - Eksik rapor Excel analizi (Admin)\n"
            f"`/haftalik_eksik_raporlar` - Haftalık eksik rapor analizi (Admin)\n"
            f"`/aylik_eksik_raporlar` - Aylık eksik rapor analizi (Admin)\n\n"
            f"⚡ Super Admin Komutları:\n"
            f"`/reload` - Excel dosyasını yenile (Super Admin)\n"
            f"`/yedekle` - Manuel yedekleme (Super Admin)\n"
            f"`/chatid` - Chat ID göster (Super Admin)\n"
            f"`/excel_durum` - Excel sistem durumu (Super Admin)\n"
            f"`/reset_database` - Veritabanını sıfırla (Super Admin)\n"
            f"`/fix_sequences` - Sequence'leri düzelt (Super Admin)\n\n"
            f"🔒 Not: Komutlar yetkinize göre çalışacaktır."
        )
    else:
        info_text = (
            f"🤖 Yapay Zeka Destekli Rapor Botu\n\n"
            f"👋 Hoş geldiniz {user_name}!\n\n"
            f"📋 Kullanıcı Komutları:\n"
            f"• Rapor göndermek için direkt mesaj yazın\n"
            f"`/start` - Botu başlat\n"
            f"`/info` - Komut bilgisi\n"
            f"`/hakkinda` - Bot hakkında\n\n"
            f"🔒 Admin komutları sadece yetkililer içindir."
        )
    
    await update.message.reply_text(info_text)

async def hakkinda_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    hakkinda_text = (
        "🤖 Rapor Botu Hakkında \n\n"
        "Geliştirici: Atamurat Kamalov\n"
        "Versiyon: 4.7.7\n"
        "Özellikler:\n\n"
        "• Her sabah 09:00'da dünkü personel icmalini Eren Boz'a gönderir\n"
        "• GPT-4 ile akıllı rapor analizi: otomatik parsing ve personel dağılımı\n"
        "• Şantiye bazlı sistem: 13+ şantiye takibi (OHP opsiyonel)\n"
        "• Otomatik hatırlatmalar: 12:30, 15:00, 17:30'da grup bildirimleri\n"
        "• Eksik rapor analizi: Excel ve detaylı raporlama\n"
        "• Haftalık rapor: Pazar 09:00 (Pazar 00:00 - Cumartesi 23:59)\n"
        "• Haftalık eksik rapor: Pazar 10:00 (aynı periyot)\n"
        "• Aylık rapor: Ayın 1'i 08:30 (önceki ayın tamamı)\n"
        "• Aylık eksik rapor: Ayın 1'i 08:45 (aynı periyot)\n"
        "• Google Cloud Storage yedekleme: Otomatik günlük yedekler\n"
        "• Gerçek zamanlı Excel takibi: Kullanıcı/şantiye güncellemeleri\n\n"
        "Daha detaylı bilgi için /info yazın."
    )
    await update.message.reply_text(hakkinda_text)

async def chatid_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await super_admin_kontrol(update, context):
        return
    
    chat_id = update.message.chat_id
    user_id = update.message.from_user.id
    
    await update.message.reply_text(
        f"📋 Chat ID Bilgileri:\n\n"
        f"👤 Kullanıcı ID: `{user_id}`\n"
        f"💬 Chat ID: `{chat_id}`\n"
        f"👥 Grup ID: `{GROUP_ID}`"
    )

async def bugun_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return
    
    target_date = dt.datetime.now(TZ).date()
    await update.message.chat.send_action(action="typing")
    rapor_mesaji = await generate_gelismis_personel_ozeti(target_date)
    await update.message.reply_text(rapor_mesaji)

async def dun_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return
    
    target_date = dt.datetime.now(TZ).date() - dt.timedelta(days=1)
    await update.message.chat.send_action(action="typing")
    rapor_mesaji = await generate_gelismis_personel_ozeti(target_date)
    await update.message.reply_text(rapor_mesaji)

async def haftalik_rapor_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return
    
    await update.message.chat.send_action(action="typing")
    
    today = dt.datetime.now(TZ).date()
    start_date = today - dt.timedelta(days=6)  # 7 günlük periyot (bugün dahil)
    end_date = today
    
    mesaj = await generate_haftalik_rapor_mesaji(start_date, end_date)
    await update.message.reply_text(mesaj)

async def aylik_rapor_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return
    
    await update.message.chat.send_action(action="typing")
    
    today = dt.datetime.now(TZ).date()
    start_date = today.replace(day=1)
    end_date = today
    
    mesaj = await generate_aylik_rapor_mesaji(start_date, end_date)
    await update.message.reply_text(mesaj)

async def haftalik_istatistik_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return
    
    await update.message.chat.send_action(action="typing")
    
    today = dt.datetime.now(TZ).date()
    start_date = today - dt.timedelta(days=6)  # 7 günlük periyot (bugün dahil)
    end_date = today
    
    mesaj = await generate_haftalik_rapor_mesaji(start_date, end_date)
    await update.message.reply_text(mesaj)

async def aylik_istatistik_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return
    
    await update.message.chat.send_action(action="typing")
    
    today = dt.datetime.now(TZ).date()
    start_date = today.replace(day=1)
    end_date = today
    
    mesaj = await generate_aylik_rapor_mesaji(start_date, end_date)
    await update.message.reply_text(mesaj)

async def tariharaligi_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return
    
    if not context.args or len(context.args) != 2:
        await update.message.reply_text(
            "📅 Tarih Aralığı Kullanımı:\n\n"
            "`/tariharaligi 01.11.2026 15.11.2026`\n"
            "Belirtilen tarih aralığı için detaylı rapor oluşturur."
        )
        return
    
    await update.message.chat.send_action(action="typing")
    
    try:
        start_date = parse_tr_date(context.args[0])
        end_date = parse_tr_date(context.args[1])
        
        if start_date > end_date:
            await update.message.reply_text("❌ Başlangıç tarihi bitiş tarihinden büyük olamaz.")
            return
        
        mesaj = await generate_tarih_araligi_raporu(start_date, end_date)
        
        await update.message.reply_text(mesaj)
        
    except Exception as e:
        await update.message.reply_text("❌ Tarih formatı hatalı. GG.AA.YYYY şeklinde girin.")

async def excel_tariharaligi_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return

    if not context.args or len(context.args) != 2:
        await update.message.reply_text(
            "📅 Excel Tarih Aralığı Raporu\n\n"
            "Kullanım: `/excel_tariharaligi 01.11.2026 15.11.2026`\n"
            "Belirtilen tarih aralığı için Excel raporu oluşturur."
        )
        return

    await update.message.reply_text("⌛ Excel raporu hazırlanıyor...")

    try:
        tarih1 = context.args[0].replace('/', '.').replace('-', '.')
        tarih2 = context.args[1].replace('/', '.').replace('-', '.')
        
        start_date = parse_tr_date(tarih1)
        end_date = parse_tr_date(tarih2)
        
        if start_date > end_date:
            await update.message.reply_text("❌ Başlangıç tarihi bitiş tarihinden büyük olamaz.")
            return

        mesaj = await generate_tarih_araligi_raporu(start_date, end_date)
        excel_dosyasi = await create_excel_report(start_date, end_date, 
                                                 f"Tarih_Araligi_{start_date.strftime('%d.%m.%Y')}_{end_date.strftime('%d.%m.%Y')}")

        await update.message.reply_text(mesaj)
        
        with open(excel_dosyasi, 'rb') as file:
            await update.message.reply_document(
                document=file,
                filename=f"Rapor_{start_date.strftime('%d.%m.%Y')}_{end_date.strftime('%d.%m.%Y')}.xlsx",
                caption=f"📊 Tarih Aralığı Raporu: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
            )
        
        os.unlink(excel_dosyasi)
        
    except Exception as e:
        await update.message.reply_text("❌ Tarih formatı hatalı. GG.AA.YYYY şeklinde girin.")
        logging.error(f"Excel tarih aralığı rapor hatası: {e}")

async def kullanicilar_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return
    
    mesaj = "👥 TÜM KULLANICI LİSTESİ - ŞANTİYE BAZLI\n\n"
    
    mesaj += f"📋 Aktif Kullanıcılar ({len(rapor_sorumlulari)} - Aktif/Pasif='E'):\n"
    
    # Proje bazlı kullanıcı sayıları - TÜMÜ hariç
    proje_kullanici_sayilari = {}
    for tid in rapor_sorumlulari:
        projeler = id_to_projects.get(tid, [])
        # TÜMÜ şantiyesini filtrele
        projeler = [proje for proje in projeler if proje != "TÜMÜ"]
        for proje in projeler:
            if proje not in proje_kullanici_sayilari:
                proje_kullanici_sayilari[proje] = 0
            proje_kullanici_sayilari[proje] += 1
    
    for proje, sayi in sorted(proje_kullanici_sayilari.items()):
        mesaj += f"• {proje}: {sayi} kullanıcı\n"
    
    mesaj += f"\n🛡️ Adminler: {len(ADMINS)}\n"
    mesaj += f"👀 İzleyiciler: {len(IZLEYICILER)}\n"
    mesaj += f"🏗️ Toplam Şantiye: {len(santiye_sorumlulari)} \n"
    
    await update.message.reply_text(mesaj)

async def santiyeler_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return
    
    mesaj = "🏗️ ŞANTİYE LİSTESİ\n\n"
    
    # TÜMÜ şantiyesini filtrele, SABİT ŞANTİYELERİ EKLE
    filtered_santiyeler = {santiye: sorumlular for santiye, sorumlular in santiye_sorumlulari.items() if santiye != "TÜMÜ"}
    
    # Sabit şantiyeleri ekle
    for santiye in SABIT_SANTIYELER:
        if santiye not in filtered_santiyeler:
            filtered_santiyeler[santiye] = []
    
    # OPSİYONEL ŞANTİYELERİ EKLE (OHP gibi) ama opsiyonel olduğunu belirt
    for santiye in OPSIYONEL_SANTIYELER:
        if santiye not in filtered_santiyeler:
            filtered_santiyeler[santiye] = []
    
    for santiye in sorted(filtered_santiyeler.keys()):
        # Opsiyonel şantiyeler için özel işaret
        if santiye in OPSIYONEL_SANTIYELER:
            mesaj += f"• {santiye} (Opsiyonel - rapor gönderilirse işlenir)\n"
        else:
            mesaj += f"• {santiye}\n"
    
    mesaj += f"\n📊 Toplam {len(filtered_santiyeler)} şantiye ({len(OPSIYONEL_SANTIYELER)} opsiyonel)"
    
    await update.message.reply_text(mesaj)

async def santiye_durum_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return
    
    bugun = dt.datetime.now(TZ).date()
    durum = await get_santiye_bazli_rapor_durumu(bugun)
    
    mesaj = f"📊 Şantiye Rapor Durumu - {bugun.strftime('%d.%m.%Y')} \n\n"
    
    mesaj += f"✅ Rapor İleten Şantiyeler ({len(durum['rapor_veren_santiyeler'])}):\n"
    for santiye in sorted(durum['rapor_veren_santiyeler']):
        mesaj += f"• {santiye}\n"
    
    mesaj += f"\n❌ Rapor İletilmeyen Şantiyeler ({len(durum['eksik_santiyeler'])}):\n"
    for santiye in sorted(durum['eksik_santiyeler']):
        if santiye in ["Belli değil", "TÜMÜ"]:
            continue
        mesaj += f"• {santiye}\n"
    
    # OPSİYONEL ŞANTİYELER HAKKINDA NOT
    if OPSIYONEL_SANTIYELER:
        mesaj += f"\nℹ️ Opsiyonel Şantiyeler (OHP): Rapor gönderilirse işlenir, gönderilmezse eksik sayılmaz\n"
    
    mesaj += f"\n📈 Özet: {len(durum['rapor_veren_santiyeler'])}/{len(durum['tum_santiyeler'])} şantiye rapor iletmiş"
    
    await update.message.reply_text(mesaj)

async def maliyet_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return
    
    rapor = maliyet_analiz.maliyet_raporu()
    await update.message.reply_text(rapor)

async def ai_rapor_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await admin_kontrol(update, context):
        return
    
    await update.message.chat.send_action(action="typing")
    
    rapor = maliyet_analiz.detayli_ai_raporu()
    await update.message.reply_text(rapor)

async def reload_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await super_admin_kontrol(update, context):
        return
    
    global excel_file_hash, excel_last_modified
    excel_file_hash = None
    excel_last_modified = 0
    
    load_excel_intelligent()
    await update.message.reply_text("✅ Excel dosyası ZORUNLU yeniden yüklendi! (Önbellek temizlendi)")

async def reset_database_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await super_admin_kontrol(update, context):
        return
    
    await update.message.reply_text("🔄 Veritabanı sıfırlanıyor... Bu işlem biraz zaman alabilir.")
    
    try:
        _sync_execute_safe("DROP SCHEMA public CASCADE")
        _sync_execute_safe("CREATE SCHEMA public")
        
        init_database()
        init_db_pool()
        
        await update.message.reply_text("✅ Veritabanı başarıyla sıfırlandı! Tüm tablolar yeniden oluşturuldu.")
        
    except Exception as e:
        logging.error(f"❌ Veritabanı sıfırlama hatası: {e}")
        await update.message.reply_text(f"❌ Veritabanı sıfırlama hatası: {e}")

async def fix_sequences_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await super_admin_kontrol(update, context):
        return
    
    await update.message.reply_text("🔄 Sequence'ler düzeltiliyor...")
    
    try:
        reports_result = await async_fetchone("SELECT COALESCE(MAX(id), 0) FROM reports")
        reports_max_id = safe_get_tuple_value(reports_result, 0, 0)
        new_reports_seq = max(reports_max_id + 1, 1)
        
        await async_execute(f"ALTER SEQUENCE reports_id_seq RESTART WITH {new_reports_seq}")
        
        ai_logs_result = await async_fetchone("SELECT COALESCE(MAX(id), 0) FROM ai_logs")
        ai_logs_max_id = safe_get_tuple_value(ai_logs_result, 0, 0)
        new_ai_logs_seq = max(ai_logs_max_id + 1, 1)
        
        await async_execute(f"ALTER SEQUENCE ai_logs_id_seq RESTART WITH {new_ai_logs_seq}")
        
        await update.message.reply_text(
            f"✅ Sequence'ler başarıyla düzeltildi!\n\n"
            f"📊 Reports: {new_reports_seq}\n"
            f"🤖 AI Logs: {new_ai_logs_seq}"
        )
        
    except Exception as e:
        logging.error(f"❌ Sequence düzeltme hatası: {e}")
        await update.message.reply_text(f"❌ Sequence düzeltme hatası: {e}")

# YENİ: ŞANTİYE-SORUMLU SABİT EŞLEŞMELERİ (Excel çıktısı için)
SABIT_SANTIYE_SORUMLULARI = {
    "BWC": "Yusuf Özçelik",
    "SKP": "Yusuf Mutlu", 
    "DMC": "Umut Ali",
    "PİRAMİT": "Onur Çetin",
    "LOT13": "Adnan Keleş",
    "LOT71": "Adnan Keleş", 
    "STADYUM": "Adnan Keleş",
    "MMP": "Orhan Ceylan",
    "MOS": "Orhan Ceylan",
    "RMC": "Orhan Ceylan",
    "TYM": "Orhan Ceylan",
    "YHP": "Orhan Ceylan",
    "KÖKSARAY": "Erdoğan Karamısır",
    # "OHP" ve "DATA CENTR": "Opsiyonel - Rapor gelirse işlenir, gelmezse eksik sayılmaz"
}

# YENİ: DİNAMİK EXCEL RAPORU OLUŞTURMA FONKSİYONU BELIRLI TARIH ARALIGI ICIN
async def create_excel_report(start_date, end_date, rapor_baslik):
    """Güncellendi: Günler satırlarda, önce GENEL TOPLAM sütunları, sonra şantiyeler (her biri 7 sütun)"""
    try:
        # 1. Tarih aralığındaki tüm günleri listele
        gunler = []
        current_date = start_date
        while current_date <= end_date:
            gunler.append(current_date)
            current_date += dt.timedelta(days=1)
        
        # 2. Şantiyeleri SABIT_SANTIYE_SORUMLULARI'ndan al (verilen sırayla)
        santiye_sirasi = [
            "BWC", "SKP", "DMC", "KÖKSARAY", "LOT13", "LOT71",
            "STADYUM", "MMP", "MOS", "PİRAMİT", "RMC", "TYM", "YHP",
        ]
        
        # 3. Veritabanından raporları al
        rows = await async_fetchall("""
            SELECT project_name, report_date, ai_analysis
            FROM reports
            WHERE report_date BETWEEN %s AND %s
            ORDER BY report_date
        """, (start_date, end_date))
        
        # 4. Raporları işle ve sözlükte sakla
        rapor_dict = {}
        for row in rows:
            if len(row) < 3:
                continue
            proje_adi = safe_get_tuple_value(row, 0, '')
            tarih = safe_get_tuple_value(row, 1, None)
            ai_analysis = safe_get_tuple_value(row, 2, '{}')
            if not tarih:
                continue
            if isinstance(tarih, dt.datetime):
                tarih = tarih.date()
            proje_adi = normalize_site_name(proje_adi)
            for sabit_santiye in SABIT_SANTIYE_SORUMLULARI.keys():
                if sabit_santiye in proje_adi or proje_adi in sabit_santiye:
                    proje_adi = sabit_santiye
                    break
            if not proje_adi or proje_adi not in SABIT_SANTIYE_SORUMLULARI:
                continue
            if proje_adi not in rapor_dict:
                rapor_dict[proje_adi] = {}
            try:
                ai_data = safe_json_loads(ai_analysis)
                yeni_format = ai_data.get('yeni_sabit_format', {})
                personel_dagilimi = ai_data.get('personel_dagilimi', {})
                if yeni_format:
                    staff = yeni_format.get('staff', 0)
                    calisan = yeni_format.get('calisan', 0)
                    ambarci = yeni_format.get('ambarci', 0)
                    mobilizasyon = yeni_format.get('mobilizasyon', 0)
                    izinli = yeni_format.get('izinli', 0)
                    dis_gorev = yeni_format.get('dis_gorev_toplam', 0)
                elif personel_dagilimi:
                    staff = personel_dagilimi.get('staff', 0)
                    calisan = personel_dagilimi.get('calisan', 0)
                    ambarci = personel_dagilimi.get('ambarci', 0)
                    mobilizasyon = personel_dagilimi.get('mobilizasyon', 0)
                    izinli = personel_dagilimi.get('izinli', 0)
                    dis_gorev = personel_dagilimi.get('dis_gorev_toplam', 0)
                else:
                    staff = calisan = ambarci = mobilizasyon = izinli = dis_gorev = 0
            except Exception:
                staff = calisan = ambarci = mobilizasyon = izinli = dis_gorev = 0
            has_data = staff > 0 or calisan > 0 or ambarci > 0 or mobilizasyon > 0 or izinli > 0 or dis_gorev > 0
            rapor_dict[proje_adi][tarih] = {
                'staff': staff,
                'calisan': calisan,
                'ambarci': ambarci,
                'mobilizasyon': mobilizasyon,
                'izinli': izinli,
                'dis_gorev': dis_gorev,
                'has_data': has_data
            }
        
        # 5. Excel oluşturma
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Raporlar"
        
        # Stiller
        baslik_font = Font(bold=True, color="366092", size=14, name="Calibri")
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        subheader_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        normal_font = Font(size=11, name="Calibri")
        bold_font = Font(bold=True, size=11, name="Calibri")
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Mavi
        toplam_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")  # Açık mavi
        calisma_yok_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Sarı
        eksik_rapor_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Açık kırmızı
        data_fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")  # Açık yeşil
        
        thin_border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center')
        number_format = '# ##0'
        
        # Sütun tanımları
        COL_TARIH = 1                                # A
        COL_GENEL_START = 2                          # B (GENEL TOPLAM'ın ilk sütunu)
        # GENEL TOPLAM 7 sütun kaplar: B'den H'ye
        GENEL_SUTUN_ADEDI = 7
        # İlk şantiye sütunu: COL_GENEL_START + GENEL_SUTUN_ADEDI = 2+7 = 9 → I sütunu
        COL_SANTIYE_START = COL_GENEL_START + GENEL_SUTUN_ADEDI  # I
        
        # Toplam sütun sayısı = 1 (tarih) + 7 (genel) + (şantiye sayısı * 7)
        toplam_sutun_sayisi = 1 + GENEL_SUTUN_ADEDI + len(santiye_sirasi) * 7
        
        # 1. SATIR: ANA BAŞLIK (tüm sütunlar birleşik)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=toplam_sutun_sayisi)
        title_cell = ws.cell(row=1, column=1, value=f"Tarih Aralığı Raporu: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}")
        title_cell.font = baslik_font
        title_cell.alignment = center_align
        
        # 2. SATIR: ÜST BAŞLIKLAR
        # A2: TARIH
        a2 = ws.cell(row=2, column=COL_TARIH, value="TARIH")
        a2.font = header_font
        a2.fill = header_fill
        a2.alignment = center_align
        a2.border = thin_border
        
        # B2:H2: GENEL TOPLAM (7 sütun birleşik)
        ws.merge_cells(start_row=2, start_column=COL_GENEL_START, end_row=2, end_column=COL_GENEL_START+GENEL_SUTUN_ADEDI-1)
        genel_baslik = ws.cell(row=2, column=COL_GENEL_START, value="GENEL TOPLAM")
        genel_baslik.font = header_font
        genel_baslik.fill = header_fill
        genel_baslik.alignment = center_align
        genel_baslik.border = thin_border
        
        # Her şantiye için 7 sütunluk başlık (I2'den itibaren)
        col_idx = COL_SANTIYE_START
        for santiye in santiye_sirasi:
            end_col = col_idx + 6
            ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=end_col)
            cell = ws.cell(row=2, column=col_idx, value=santiye)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            col_idx = end_col + 1
        
        # 3. SATIR: ALT BAŞLIKLAR (kategoriler)
        kategoriler = ["Staff", "Çalışan", "Ambarcı", "Mobilizasyon", "İzinli", "Dış Görev", "Toplam"]
        
        # GENEL TOPLAM kategorileri (B3:H3)
        for i, kategori in enumerate(kategoriler):
            col = COL_GENEL_START + i
            cell = ws.cell(row=3, column=col, value=kategori)
            cell.font = subheader_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Her şantiye için kategoriler
        col_idx = COL_SANTIYE_START
        for santiye in santiye_sirasi:
            for i, kategori in enumerate(kategoriler):
                cell = ws.cell(row=3, column=col_idx + i, value=kategori)
                cell.font = subheader_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            col_idx += 7
        
        # 4. SATIR ve sonrası: GÜN SATIRLARI
        row_idx = 4
        for gun in gunler:
            # A sütunu: Tarih
            tarih_cell = ws.cell(row=row_idx, column=COL_TARIH, value=gun.strftime('%d.%m.%Y'))
            tarih_cell.alignment = left_align
            tarih_cell.font = normal_font
            tarih_cell.border = thin_border
            
            # Önce GENEL TOPLAM için tüm şantiyelerin o günkü değerlerini topla
            genel_toplam = {k: 0 for k in kategoriler[:-1]}  # "Toplam" hariç, onu sonra hesapla
            for santiye in santiye_sirasi:
                gun_rapor = rapor_dict.get(santiye, {}).get(gun, None)
                if gun_rapor and gun_rapor['has_data']:
                    genel_toplam['Staff'] += gun_rapor['staff']
                    genel_toplam['Çalışan'] += gun_rapor['calisan']
                    genel_toplam['Ambarcı'] += gun_rapor['ambarci']
                    genel_toplam['Mobilizasyon'] += gun_rapor['mobilizasyon']
                    genel_toplam['İzinli'] += gun_rapor['izinli']
                    genel_toplam['Dış Görev'] += gun_rapor['dis_gorev']
            
            # GENEL TOPLAM sütunlarını doldur (B:H)
            for i, key in enumerate(['Staff', 'Çalışan', 'Ambarcı', 'Mobilizasyon', 'İzinli', 'Dış Görev']):
                deger = genel_toplam[key]
                cell = ws.cell(row=row_idx, column=COL_GENEL_START + i, value=deger if deger != 0 else "")
                if deger != 0:
                    cell.number_format = number_format
                    cell.fill = data_fill
                cell.alignment = center_align
                cell.font = normal_font
                cell.border = thin_border
            
            # GENEL TOPLAM - Toplam sütunu (H sütunu)
            toplam_deger = sum(genel_toplam[k] for k in ['Staff','Çalışan','Ambarcı','Mobilizasyon','İzinli'])
            toplam_cell = ws.cell(row=row_idx, column=COL_GENEL_START + 6, value=toplam_deger if toplam_deger != 0 else "")
            if toplam_deger != 0:
                toplam_cell.number_format = number_format
                toplam_cell.fill = toplam_fill
            toplam_cell.alignment = center_align
            toplam_cell.font = bold_font
            toplam_cell.border = thin_border
            
            # Şimdi her şantiye için sütunları doldur
            col_idx = COL_SANTIYE_START
            for santiye in santiye_sirasi:
                gun_rapor = rapor_dict.get(santiye, {}).get(gun, None)
                if gun_rapor:
                    if not gun_rapor['has_data']:
                        # Çalışma Yok – 7 sütun birleşik, sarı arka plan
                        ws.merge_cells(start_row=row_idx, start_column=col_idx, end_row=row_idx, end_column=col_idx+6)
                        cell = ws.cell(row=row_idx, column=col_idx, value="Çalışma Yok")
                        cell.alignment = center_align
                        cell.fill = calisma_yok_fill
                        cell.font = bold_font
                        cell.border = thin_border
                    else:
                        degerler = [
                            gun_rapor['staff'],
                            gun_rapor['calisan'],
                            gun_rapor['ambarci'],
                            gun_rapor['mobilizasyon'],
                            gun_rapor['izinli'],
                            gun_rapor['dis_gorev']
                        ]
                        for j, deger in enumerate(degerler):
                            cell = ws.cell(row=row_idx, column=col_idx + j, value=deger if deger != 0 else "")
                            if deger != 0:
                                cell.number_format = number_format
                                cell.fill = data_fill
                            cell.alignment = center_align
                            cell.font = normal_font
                            cell.border = thin_border
                        # Şantiye toplamı (ilk 5 kategori)
                        santiye_toplam = sum(degerler[:5])
                        toplam_cell = ws.cell(row=row_idx, column=col_idx + 6, value=santiye_toplam if santiye_toplam != 0 else "")
                        if santiye_toplam != 0:
                            toplam_cell.number_format = number_format
                            toplam_cell.fill = toplam_fill
                        toplam_cell.alignment = center_align
                        toplam_cell.font = bold_font
                        toplam_cell.border = thin_border
                else:
                    # Rapor yok – 7 sütun birleşik, kırmızı ✗
                    ws.merge_cells(start_row=row_idx, start_column=col_idx, end_row=row_idx, end_column=col_idx+6)
                    cell = ws.cell(row=row_idx, column=col_idx, value="✗")
                    cell.alignment = center_align
                    cell.fill = eksik_rapor_fill
                    cell.font = bold_font
                    cell.border = thin_border
                col_idx += 7
            row_idx += 1
        
        # TOPLAM SATIRI (her sütun için tüm günlerin toplamı)
        toplam_satir = row_idx
        ws.cell(row=toplam_satir, column=COL_TARIH, value="TOPLAM").font = bold_font
        ws.cell(row=toplam_satir, column=COL_TARIH).border = thin_border
        
        # GENEL TOPLAM sütunları için toplam (B'den H'ye)
        for col in range(COL_GENEL_START, COL_GENEL_START + GENEL_SUTUN_ADEDI):
            baslangic_satir = 4
            bitis_satir = toplam_satir - 1
            hucre_aralik = f"{get_column_letter(col)}{baslangic_satir}:{get_column_letter(col)}{bitis_satir}"
            formül = f"=SUM({hucre_aralik})"
            cell = ws.cell(row=toplam_satir, column=col, value=formül)
            cell.number_format = number_format
            cell.alignment = center_align
            cell.font = normal_font
            cell.border = thin_border
        
        # Şantiye sütunları için toplam
        col_idx = COL_SANTIYE_START
        for santiye in santiye_sirasi:
            for j in range(7):
                col = col_idx + j
                baslangic_satir = 4
                bitis_satir = toplam_satir - 1
                hucre_aralik = f"{get_column_letter(col)}{baslangic_satir}:{get_column_letter(col)}{bitis_satir}"
                formül = f"=SUM({hucre_aralik})"
                cell = ws.cell(row=toplam_satir, column=col, value=formül)
                cell.number_format = number_format
                cell.alignment = center_align
                cell.font = normal_font
                cell.border = thin_border
            col_idx += 7
        
        # EKSİK RAPOR SATIRI (her sütun için eksik rapor sayısı)
        eksik_satir = toplam_satir + 1
        ws.cell(row=eksik_satir, column=COL_TARIH, value="Eksik Rapor").font = bold_font
        ws.cell(row=eksik_satir, column=COL_TARIH).border = thin_border
        
        # GENEL TOPLAM sütunları için eksik rapor sayısı (✗ sayısı)
        # Genel Toplam sütunlarında ✗ olmaz (çünkü onlar toplam), ama tutarlılık için boş bırakabiliriz veya formül koymayız.
        # Görselde genel toplamda eksik rapor satırı yok gibi, ama biz yine de boş bırakalım.
        for col in range(COL_GENEL_START, COL_GENEL_START + GENEL_SUTUN_ADEDI):
            ws.cell(row=eksik_satir, column=col).border = thin_border
        
        # Şantiye sütunları için eksik rapor sayısı (her sütun için ayrı)
        col_idx = COL_SANTIYE_START
        for santiye in santiye_sirasi:
            # Her şantiye için 7 sütun, ama eksik rapor sadece "Staff" sütununda ✗ var, diğerleri boş olabilir.
            # Ancak biz her sütun için ayrı ayrı ✗ sayısını hesaplayabiliriz (sadece Staff sütunu için değil, tüm sütunlar aynı durumda)
            # Daha basit: Her şantiye bloğunun ilk sütunundaki ✗ sayısını al ve 7 sütuna yay (birleştir)
            start_col = col_idx
            end_col = col_idx + 6
            baslangic_satir = 4
            bitis_satir = toplam_satir - 1
            hucre_aralik = f"{get_column_letter(start_col)}{baslangic_satir}:{get_column_letter(start_col)}{bitis_satir}"
            formül = f'=COUNTIF({hucre_aralik},"✗")'
            ws.merge_cells(start_row=eksik_satir, start_column=start_col, end_row=eksik_satir, end_column=end_col)
            cell = ws.cell(row=eksik_satir, column=start_col, value=formül)
            cell.number_format = number_format
            cell.alignment = center_align
            cell.fill = eksik_rapor_fill
            cell.font = bold_font
            cell.border = thin_border
            col_idx += 7
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 15  # TARIH
        # GENEL TOPLAM sütunları (B-H)
        for col in range(COL_GENEL_START, COL_GENEL_START + GENEL_SUTUN_ADEDI):
            ws.column_dimensions[get_column_letter(col)].width = 10
        # Şantiye sütunları
        for i in range(len(santiye_sirasi) * 7):
            col_letter = get_column_letter(COL_SANTIYE_START + i)
            ws.column_dimensions[col_letter].width = 10
        
        # Satır yükseklikleri
        for r in range(1, eksik_satir + 1):
            ws.row_dimensions[r].height = 25
        
        # Otomatik filtre
        ws.auto_filter.ref = f"A3:{get_column_letter(toplam_sutun_sayisi)}{eksik_satir}"
        
        # Birleştirilmiş hücrelerin kenar çizgilerini tamamla
        for merged_range in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = merged_range.min_col, merged_range.min_row, merged_range.max_col, merged_range.max_row
            # Üst kenar
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=min_row, column=col)
                current_border = cell.border
                if current_border.top.style is None:
                    cell.border = Border(
                        left=current_border.left,
                        right=current_border.right,
                        top=Side(style='thin', color="000000"),
                        bottom=current_border.bottom
                    )
            # Alt kenar
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=max_row, column=col)
                current_border = cell.border
                if current_border.bottom.style is None:
                    cell.border = Border(
                        left=current_border.left,
                        right=current_border.right,
                        top=current_border.top,
                        bottom=Side(style='thin', color="000000")
                    )
            # Sol kenar
            for row in range(min_row, max_row + 1):
                cell = ws.cell(row=row, column=min_col)
                current_border = cell.border
                if current_border.left.style is None:
                    cell.border = Border(
                        left=Side(style='thin', color="000000"),
                        right=current_border.right,
                        top=current_border.top,
                        bottom=current_border.bottom
                    )
            # Sağ kenar
            for row in range(min_row, max_row + 1):
                cell = ws.cell(row=row, column=max_col)
                current_border = cell.border
                if current_border.right.style is None:
                    cell.border = Border(
                        left=current_border.left,
                        right=Side(style='thin', color="000000"),
                        top=current_border.top,
                        bottom=current_border.bottom
                    )
        
        # Dosyayı kaydet
        timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        return temp_file.name
        
    except Exception as e:
        logging.error(f"❌ Excel raporu oluşturma hatası: {e}")
        import traceback
        logging.error(traceback.format_exc())
        raise e

# YENİ: GÜNCELLENMİŞ ZAMANLAMA SİSTEMİ
def schedule_jobs(app):
    jq = app.job_queue
    
    logging.info("⏰ GÜNCELLENMİŞ ZAMANLAMA SİSTEMİ AYARLANIYOR...")
    logging.info(f"🔍 GROUP_ID değeri: {GROUP_ID}")
    
    if not GROUP_ID:
        logging.error("❌ GROUP_ID ayarlanmamış! Hatırlatma mesajları gönderilemeyecek.")
    else:
        logging.info(f"✅ GROUP_ID ayarlandı: {GROUP_ID}")
    
    # Mevcut job'ları ayarla
    jq.run_repeating(auto_watch_excel, interval=60, first=10)
    jq.run_daily(gunluk_rapor_ozeti, time=dt.time(9, 0, tzinfo=TZ))
    
    # YENİ ZAMANLAMALAR
    hatirlatma_job = jq.run_daily(hatirlatma_mesaji, time=dt.time(12, 30, tzinfo=TZ))
    ilk_kontrol_job = jq.run_daily(ilk_rapor_kontrol, time=dt.time(15, 0, tzinfo=TZ))
    son_kontrol_job = jq.run_daily(son_rapor_kontrol, time=dt.time(17, 30, tzinfo=TZ))
    
    # YENİ: HAFTALIK NORMAL RAPOR - HER PAZAR 09:00
    jq.run_daily(haftalik_normal_rapor_job, time=dt.time(9, 0, tzinfo=TZ), days=(6,))  # 6 = Pazar
    
    # YENİ: HAFTALIK EKSİK RAPOR - HER PAZAR 10:00
    jq.run_daily(haftalik_eksik_rapor_job, time=dt.time(10, 0, tzinfo=TZ), days=(6,))  # 6 = Pazar
    
    # YENİ: AYLIK NORMAL RAPOR - HER AYIN 1'İ 08:30
    jq.run_daily(aylik_normal_rapor_job, time=dt.time(8, 30, tzinfo=TZ))
    
    # YENİ: AYLIK EKSİK RAPOR - HER AYIN 1'İ 08:45
    jq.run_daily(aylik_eksik_rapor_job, time=dt.time(8, 45, tzinfo=TZ))
    
    jq.run_daily(yedekleme_gorevi, time=dt.time(23, 0, tzinfo=TZ))
    jq.run_daily(lambda context: asyncio.create_task(async_yedekle_postgres()), time=dt.time(23, 10, tzinfo=TZ))
    
    logging.info("⏰ Tüm zamanlamalar ayarlandı ✅")
    logging.info("   - Haftalık normal rapor: Pazar 09:00")
    logging.info("   - Haftalık eksik rapor: Pazar 10:00")
    logging.info("   - Aylık normal rapor: Ayın 1'i 08:30")
    logging.info("   - Aylık eksik rapor: Ayın 1'i 08:45")

# YENİ: DÜZELTİLMİŞ HAFTALIK RAPOR FONKSİYONU (geriye uyumluluk için)
async def haftalik_grup_raporu_duzeltilmis(context: ContextTypes.DEFAULT_TYPE):
    """DÜZELTİLDİ: Cumartesi 17:35'te Pazartesi 00:00'dan Cumartesi 17:35'e kadar olan raporları içerir"""
    try:
        today = dt.datetime.now(TZ).date()
        now_time = dt.datetime.now(TZ).time()
        
        logging.info(f"📅 Haftalık rapor tetiklendi: Bugün = {today}, Saat = {now_time}")
        
        # Haftalık rapor tarih aralığını hesapla
        # Pazartesi 00:00'dan bugün (Cumartesi) 17:35'e kadar
        start_date = today - dt.timedelta(days=today.weekday())  # Pazartesi
        end_date = today  # Bugün (Cumartesi)
        
        logging.info(f"📊 Haftalık rapor tarih aralığı: {start_date} - {end_date}")
        
        # Haftalık raport oluştur
        mesaj = await generate_haftalik_rapor_mesaji(start_date, end_date)
        
        # Grup ID kontrolü
        if GROUP_ID:
            try:
                await context.bot.send_message(chat_id=GROUP_ID, text=mesaj)
                logging.info(f"📊 Haftalık grup raporu gönderildi: {start_date} - {end_date}")
            except Exception as e:
                logging.error(f"📊 Haftalık grup raporu gönderilemedi: {e}")
        else:
            logging.error("📊 GROUP_ID ayarlanmamış, haftalık rapor gönderilemedi")
        
        # Adminlere de gönder
        for admin_id in ADMINS:
            try:
                await context.bot.send_message(chat_id=admin_id, text=mesaj)
                logging.info(f"📊 Haftalık rapor {admin_id} adminine gönderildi")
                await asyncio.sleep(0.5)
            except Exception as e:
                if "Chat not found" not in str(e):
                    logging.error(f"📊 {admin_id} adminine haftalık rapor gönderilemedi: {e}")
        
    except Exception as e:
        logging.error(f"📊 Haftalık grup raporu hatası: {e}")
        await hata_bildirimi(context, f"Haftalık grup raporu hatası: {e}")

# YENİ: ASYNC POSTGRES YEDEKLEME
async def async_yedekle_postgres():
    """Async Postgres yedekleme"""
    loop = asyncio.get_running_loop()
    await loop.run_in_executor(None, yedekle_postgres)

# YENİ: AYLIK RAPOR KONTROL FONKSİYONU (geriye uyumluluk için)
async def aylik_grup_raporu_kontrol(context: ContextTypes.DEFAULT_TYPE):
    """Ayın 1'inde aylık rapor gönder"""
    try:
        today = dt.datetime.now(TZ).date()
        if today.day == 1:  # Ayın 1'inde çalıştır
            # Önceki ayın raporunu oluştur
            start_date = today.replace(day=1) - dt.timedelta(days=1)
            start_date = start_date.replace(day=1)
            end_date = today.replace(day=1) - dt.timedelta(days=1)
            
            await aylik_grup_raporu_tarihli(context, start_date, end_date)
    except Exception as e:
        logging.error(f"🗓️ Aylık rapor kontrol hatası: {e}")

# YENİ: TARİHLİ AYLIK RAPOR (geriye uyumluluk için)
async def aylik_grup_raporu_tarihli(context: ContextTypes.DEFAULT_TYPE, start_date, end_date):
    """Belirli tarih aralığı için aylık rapor gönder"""
    try:
        mesaj = await generate_aylik_rapor_mesaji(start_date, end_date)
        
        if GROUP_ID:
            try:
                await context.bot.send_message(chat_id=GROUP_ID, text=mesaj)
                logging.info(f"🗓️ Aylık grup raporu gönderildi: {start_date} - {end_date}")
            except Exception as e:
                logging.error(f"🗓️ Aylık grup raporu gönderilemedi: {e}")
        
        for admin_id in ADMINS:
            try:
                await context.bot.send_message(chat_id=admin_id, text=mesaj)
                logging.info(f"🗓️ Aylık rapor {admin_id} adminine gönderildi")
                await asyncio.sleep(0.5)
            except Exception as e:
                if "Chat not found" not in str(e):
                    logging.error(f"🗓️ {admin_id} adminine aylık rapor gönderilemedi: {e}")
        
    except Exception as e:
        logging.error(f"🗓️ Aylık grup raporu hatası: {e}")

async def auto_watch_excel(context: ContextTypes.DEFAULT_TYPE):
    try:
        load_excel_intelligent()
    except Exception as e:
        logging.error(f"Excel otomatik izleme hatası: {e}")

async def gunluk_rapor_ozeti(context: ContextTypes.DEFAULT_TYPE):
    try:
        dun = (dt.datetime.now(TZ) - dt.timedelta(days=1)).date()
        rapor_mesaji = await generate_gelismis_personel_ozeti(dun)
        
        # DÜZELTİLDİ: Hem Eren Boz'a hem de sana (Super Admin) gönder
        hedef_kullanicilar = [709746899, 1000157326]  # Eren Boz ve Atamurat Kamalov
        
        for user_id in hedef_kullanicilar:
            try:
                await context.bot.send_message(chat_id=user_id, text=rapor_mesaji)
                logging.info(f"🕘 09:00 özeti {user_id} kullanıcısına gönderildi")
                await asyncio.sleep(0.5)
            except Exception as e:
                logging.error(f"🕘 {user_id} kullanıcısına özet gönderilemedi: {e}")
                
    except Exception as e:
        logging.error(f"🕘 09:00 rapor hatası: {e}")
        await hata_bildirimi(context, f"09:00 rapor hatası: {e}")

async def hatirlatma_mesaji(context: ContextTypes.DEFAULT_TYPE):
    try:
        logging.info("12:30 hatırlatma mesajı tetiklendi")
        bugun = dt.datetime.now(TZ).date()
        durum = await get_santiye_bazli_rapor_durumu(bugun)
        
        if GROUP_ID:
            if not durum['eksik_santiyeler']:
                mesaj = "✅ Bugün için tüm şantiyelerden raporlar alınmış.\n\n"
                mesaj += "📝 Not: Eksik rapor bulunmamaktadır. Düzenli paylaşımlarınız için teşekkürler. 🙏"
            else:
                mesaj = "❌ Eksik raporlar var:\n"
                for santiye in sorted(durum['eksik_santiyeler']):
                    # OPSİYONEL ŞANTİYELERİ ATLA (OHP gibi)
                    if santiye in OPSIYONEL_SANTIYELER:
                        continue
                    
                    # Şantiye için kullanıcı adlarını al
                    usernames = SANTIYE_USERNAME_MAPPING.get(santiye, [])
                    if usernames:
                        # Kullanıcı adlarını @ ile birleştir
                        username_str = " @" + ", @".join(usernames)
                        mesaj += f"• {santiye} ({username_str} )\n"
                    else:
                        mesaj += f"• {santiye}\n"
                
                # Eğer opsiyonel şantiyeler hariç tüm şantiyeler rapor verdiyse
                eksik_santiyeler_filtreli = [s for s in durum['eksik_santiyeler'] if s not in OPSIYONEL_SANTIYELER]
                if not eksik_santiyeler_filtreli:
                    mesaj = "✅ Bugün için tüm şantiyelerden raporlar alınmış.\n\n"
                    mesaj += "📝 Not: Eksik rapor bulunmamaktadır. Düzenli paylaşımlarınız için teşekkürler. 🙏"
                else:
                    # SABİT NOT EKLENİYOR (eksik rapor varsa)
                    mesaj += "\n\n📝 Not: Şantiyenin dili verdiği rapordur; raporu olmayan iş tamamlanmış sayılmaz. ⚠️\nLütfen günlük raporlarınızı zamanında iletiniz."
            
            try:
                await context.bot.send_message(chat_id=GROUP_ID, text=mesaj)
                logging.info(f"🟡 12:30 hatırlatma mesajı gruba gönderildi: {GROUP_ID}")
            except Exception as e:
                logging.error(f"🟡 Gruba hatırlatma mesajı gönderilemedi: {e}")
        else:
            logging.error("🟡 GROUP_ID ayarlanmamış, hatırlatma mesajı gönderilemedi")
            
    except Exception as e:
        logging.error(f"Hatırlatma mesajı hatası: {e}")

async def ilk_rapor_kontrol(context: ContextTypes.DEFAULT_TYPE):
    try:
        bugun = dt.datetime.now(TZ).date()
        durum = await get_santiye_bazli_rapor_durumu(bugun)
        
        mesaj = "🕒 15:00 Şantiye Rapor Durumu\n\n"
        
        if durum['rapor_veren_santiyeler']:
            mesaj += f"✅ Rapor iletilen şantiyeler ({len(durum['rapor_veren_santiyeler'])}):\n"
            for santiye in sorted(durum['rapor_veren_santiyeler']):
                mesaj += f"• {santiye}\n"
            mesaj += "\n"
        else:
            mesaj += "✅ Rapor iletilen şantiyeler (0):\n\n"
        
        # OPSİYONEL ŞANTİYELER HARİÇ EKSİK ŞANTİYELER
        eksik_santiyeler_filtreli = [s for s in sorted(durum['eksik_santiyeler']) if s not in OPSIYONEL_SANTIYELER and s not in ["Belli değil", "Tümü"]]
        
        if eksik_santiyeler_filtreli:
            mesaj += f"❌ Rapor iletilmeyen şantiyeler ({len(eksik_santiyeler_filtreli)}):\n"
            for santiye in eksik_santiyeler_filtreli:
                mesaj += f"• {santiye}\n"
            
            # EKSİK RAPOR VARSA MEVCUT NOT
            mesaj += "\n\n📝 Not: Yapılan işin raporunu vermek, işi yapmak kadar önemlidir. ⚠️\nEksik olan raporları lütfen iletiniz."
        else:
            mesaj += "❌ Rapor iletilmeyen şantiyeler (0):\n"
            mesaj += "🎉 Tüm şantiyeler raporlarını iletti!\n\n"
            # EKSİK RAPOR YOKSA YENİ NOT
            mesaj += "📝 Not: Eksik rapor bulunmamaktadır. Düzenli paylaşımlarınız için teşekkürler. 🙏"
        
        if GROUP_ID:
            try:
                await context.bot.send_message(chat_id=GROUP_ID, text=mesaj)
                logging.info(f"🟠 15:00 şantiye kontrol mesajı gruba gönderildi: {GROUP_ID}")
            except Exception as e:
                logging.error(f"🟠 Gruba şantiye kontrol mesajı gönderilemedi: {e}")
        else:
            logging.error("🟠 GROUP_ID ayarlanmamış, şantiye kontrol mesajı gönderilemedi")
        
    except Exception as e:
        logging.error(f"🟠 Şantiye rapor kontrol hatası: {e}")
        await hata_bildirimi(context, f"Şantiye rapor kontrol hatası: {e}")

async def son_rapor_kontrol(context: ContextTypes.DEFAULT_TYPE):
    """🔴 17:30 - Gün sonu şantiye bazlı rapor analizi - GRUBA GÖNDER"""
    try:
        bugun = dt.datetime.now(TZ).date()
        durum = await get_santiye_bazli_rapor_durumu(bugun)
        
        result = await async_fetchone("SELECT COUNT(*) FROM reports WHERE report_date = %s", (bugun,))
        toplam_rapor = result[0] if result else 0
        
        mesaj = "🕠 Gün Sonu Şantiye Rapor Analizi\n\n"
        
        # OPSİYONEL ŞANTİYELER HARİÇ EKSİK ŞANTİYELER
        eksik_santiyeler_filtreli = [s for s in sorted(durum['eksik_santiyeler']) if s not in OPSIYONEL_SANTIYELER]
        
        if eksik_santiyeler_filtreli:
            mesaj += f"❌ Rapor İletilmeyen Şantiyeler ({len(eksik_santiyeler_filtreli)}):\n"
            for santiye in eksik_santiyeler_filtreli:
                mesaj += f"• {santiye}\n"
            
            mesaj += f"\n📊 Bugün toplam {toplam_rapor} rapor alındı."
            mesaj += f"\n🏗️ {len(durum['rapor_veren_santiyeler'])}/{len(durum['tum_santiyeler'])} şantiye rapor iletmiş durumda."
            
            # EKSİK RAPOR VARSA MEVCUT NOT
            mesaj += "\n\n📝 Not:\nYapılan işin raporunu vermek, saha yönetiminin en kritik adımıdır. 📊\nBunca çabaya rağmen rapor iletmeyen şantiyeler, lütfen rapor düzenine özen göstersin. 🙏\nUnutmayın: İşi yapmak cesarettir, raporlamak ise disiplindir. ⚠️"
        else:
            mesaj += "❌ Rapor İletilmeyen Şantiyeler (0):\n"
            mesaj += "🎉 Tüm şantiyeler raporlarını iletti!\n"
            mesaj += f"\n📊 Bugün toplam {toplam_rapor} rapor alındı."
            mesaj += f"\n🏗️ {len(durum['rapor_veren_santiyeler'])}/{len(durum['tum_santiyeler'])} şantiye rapor iletmiş durumda.\n\n"
            
            # EKSİK RAPOR YOKSA YENİ NOT
            mesaj += "📝 Not: Eksik rapor bulunmamaktadır. Düzenli paylaşımlarınız için teşekkürler. 🙏"
        
        # DÜZELTİLDİ: GRUBA GÖNDER
        if GROUP_ID:
            try:
                await context.bot.send_message(chat_id=GROUP_ID, text=mesaj)
                logging.info(f"🔴 17:30 gün sonu analizi gruba gönderildi: {GROUP_ID}")
            except Exception as e:
                logging.error(f"🔴 Gruba gün sonu analizi gönderilemedi: {e}")
        else:
            logging.error("🔴 GROUP_ID ayarlanmamış, gün sonu analizi gönderilemedi")
        
    except Exception as e:
        logging.error(f"🔴 Şantiye son rapor kontrol hatası: {e}")
        await hata_bildirimi(context, f"Şantiye son rapor kontrol hatası: {e}")

async def haftalik_grup_raporu(context: ContextTypes.DEFAULT_TYPE):
    """Eski haftalık rapor fonksiyonu - geriye uyumluluk için (artık kullanılmayacak)"""
    try:
        today = dt.datetime.now(TZ).date()
        
        # DÜZELTİLDİ: ÖNCEKİ HAFTANIN RAPORUNU HAZIRLA (Bugünden 7 gün geriye)
        end_date = today - dt.timedelta(days=1)  # Dün dahil
        start_date = end_date - dt.timedelta(days=6)  # 6 gün geri (7 günlük periyot)
        
        mesaj = await generate_haftalik_rapor_mesaji(start_date, end_date)
        
        if GROUP_ID:
            try:
                await context.bot.send_message(chat_id=GROUP_ID, text=mesaj)
                logging.info(f"📊 Haftalık grup raporu gönderildi: {start_date} - {end_date}")
            except Exception as e:
                logging.error(f"📊 Haftalık grup raporu gönderilemedi: {e}")
        
        for admin_id in ADMINS:
            try:
                await context.bot.send_message(chat_id=admin_id, text=mesaj)
                logging.info(f"📊 Haftalık rapor {admin_id} adminine gönderildi")
                await asyncio.sleep(0.5)
            except Exception as e:
                if "Chat not found" not in str(e):
                    logging.error(f"📊 {admin_id} adminine haftalık rapor gönderilemedi: {e}")
        
    except Exception as e:
        logging.error(f"📊 Haftalık grup raporu hatası: {e}")
        await hata_bildirimi(context, f"Haftalık grup raporu hatası: {e}")

async def aylik_grup_raporu(context: ContextTypes.DEFAULT_TYPE):
    """Mevcut aylık rapor fonksiyonu - geriye uyumluluk için"""
    try:
        today = dt.datetime.now(TZ).date()
        start_date = today.replace(day=1)
        end_date = today
        
        mesaj = await generate_aylik_rapor_mesaji(start_date, end_date)
        
        if GROUP_ID:
            try:
                await context.bot.send_message(chat_id=GROUP_ID, text=mesaj)
                logging.info(f"🗓️ Aylık grup raporu gönderildi: {start_date} - {end_date}")
            except Exception as e:
                logging.error(f"🗓️ Aylık grup raporu gönderilemedi: {e}")
        
        for admin_id in ADMINS:
            try:
                await context.bot.send_message(chat_id=admin_id, text=mesaj)
                logging.info(f"🗓️ Aylık rapor {admin_id} adminine gönderildi")
                await asyncio.sleep(0.5)
            except Exception as e:
                if "Chat not found" not in str(e):
                    logging.error(f"🗓️ {admin_id} adminine aylık rapor gönderilemedi: {e}")
        
    except Exception as e:
        logging.error(f"🗓️ Aylık grup raporu hatası: {e}")
        await hata_bildirimi(context, f"Aylık grup raporu hatası: {e}")

async def bot_baslatici_mesaji(context: ContextTypes.DEFAULT_TYPE):
    try:
        mesaj = "🤖 Rapor Kontrol Botu Aktif!\n\nKontrol bende ⚡️\nKolay gelsin 👷‍♂️"
        
        for admin_id in ADMINS:
            try:
                await context.bot.send_message(chat_id=admin_id, text=mesaj)
                logging.info(f"Başlangıç mesajı {admin_id} adminine gönderildi")
                await asyncio.sleep(0.5)
            except Exception as e:
                if "Chat not found" not in str(e):
                    logging.error(f"Başlangıç mesajı {admin_id} adminine gönderilemedi: {e}")
        
    except Exception as e:
        logging.error(f"Bot başlatıcı mesaj hatası: {e}")

async def post_init(application: Application):
    # Temel komutlar (tüm kullanıcılar için)
    basic_commands = [
        BotCommand("start", "Botu başlat"),
        BotCommand("info", "Komut bilgisi"),
        BotCommand("hakkinda", "Bot hakkında bilgi"),
    ]
    
    # Admin komutları (sadece adminler için)
    admin_commands = basic_commands + [
        BotCommand("bugun", "Bugünün özeti (Admin)"),
        BotCommand("dun", "Dünün özeti (Admin)"),
        BotCommand("eksikraporlar", "Eksik raporları listele (Admin)"),
        BotCommand("istatistik", "Genel istatistikler (Admin)"),
        BotCommand("haftalik_rapor", "Haftalık rapor (Admin)"),
        BotCommand("aylik_rapor", "Aylık rapor (Admin)"),
        BotCommand("tariharaligi", "Tarih aralığı raporu mesaj halinde (Admin)"),
        BotCommand("haftalik_istatistik", "Haftalık istatistik (Admin)"),
        BotCommand("aylik_istatistik", "Aylık istatistik (Admin)"),
        BotCommand("excel_tariharaligi", "Excel tarih aralığı raporu (Admin)"),
        BotCommand("maliyet", "Maliyet analizi (Admin)"),
        BotCommand("ai_rapor", "Detaylı AI raporu (Admin)"),
        BotCommand("kullanicilar", "Tüm kullanıcı listesi (Admin)"),
        BotCommand("santiyeler", "Şantiye listesi (Admin)"),
        BotCommand("santiye_durum", "Şantiye rapor durumu (Admin)"),
        BotCommand("eksik_rapor_excel", "Eksik rapor Excel analizi (Admin)"),
        BotCommand("haftalik_eksik_raporlar", "Haftalık eksik rapor analizi (Admin)"),
        BotCommand("aylik_eksik_raporlar", "Aylık eksik rapor analizi (Admin)"),
    ]
    
    # Super Admin komutları (sadece super admin için)
    super_admin_commands = admin_commands + [
        BotCommand("reload", "Excel yenile (Super Admin)"),
        BotCommand("yedekle", "Manuel yedekleme (Super Admin)"),
        BotCommand("chatid", "Chat ID göster (Super Admin)"),
        BotCommand("excel_durum", "Excel sistem durumu (Super Admin)"),
        BotCommand("reset_database", "Veritabanını sıfırla (Super Admin)"),
        BotCommand("fix_sequences", "Sequence'leri düzelt (Super Admin)"),
    ]
    
    # Varsayılan komutları temel komutlar olarak ayarla.
    await application.bot.set_my_commands(basic_commands)
    
    # Eğer özel sohbetler için komut ayarlama desteği varsa, tüm özel sohbetler için temel komutları ayarla.
    if HAS_PRIVATE_SCOPE:
        try:
            from telegram import BotCommandScopeAllPrivateChats
            await application.bot.set_my_commands(basic_commands, scope=BotCommandScopeAllPrivateChats())
            logging.info("Özel sohbetler için temel komutlar ayarlandı.")
        except Exception as e:
            logging.error(f"Özel sohbetler için komutlar ayarlanamadı: {e}")
    
    # Her bir admin kullanıcısı için admin komutlarını ayarla.
    for admin_id in ADMINS:
        try:
            from telegram import BotCommandScopeChat
            scope = BotCommandScopeChat(chat_id=admin_id)
            # Eğer admin aynı zamanda super admin ise, super admin komutlarını ayarla.
            if admin_id == SUPER_ADMIN_ID:
                await application.bot.set_my_commands(super_admin_commands, scope=scope)
            else:
                await application.bot.set_my_commands(admin_commands, scope=scope)
            logging.info(f"Admin komutları {admin_id} kullanıcısı için ayarlandı.")
        except Exception as e:
            logging.error(f"Admin {admin_id} için komutlar ayarlanamadı: {e}")
    
    await bot_baslatici_mesaji(application)

def main():
    try:
        logging.info("🚀 Bot başlatılıyor...")
        
        # Önce veritabanı bağlantılarını test et
        init_db_pool()
        init_database()
        
        app = Application.builder().token(BOT_TOKEN).post_init(post_init).build()
        
        # Handler'ları ekle
        app.add_handler(CommandHandler("start", start_cmd))
        app.add_handler(CommandHandler("info", info_cmd))
        app.add_handler(CommandHandler("hakkinda", hakkinda_cmd))
        
        app.add_handler(CommandHandler("bugun", bugun_cmd))
        app.add_handler(CommandHandler("dun", dun_cmd))
        app.add_handler(CommandHandler("eksikraporlar", eksikraporlar_cmd))
        app.add_handler(CommandHandler("istatistik", istatistik_cmd))
        app.add_handler(CommandHandler("haftalik_rapor", haftalik_rapor_cmd))
        app.add_handler(CommandHandler("aylik_rapor", aylik_rapor_cmd))
        app.add_handler(CommandHandler("haftalik_istatistik", haftalik_istatistik_cmd))
        app.add_handler(CommandHandler("aylik_istatistik", aylik_istatistik_cmd))
        app.add_handler(CommandHandler("tariharaligi", tariharaligi_cmd))
        app.add_handler(CommandHandler("excel_tariharaligi", excel_tariharaligi_cmd))
        app.add_handler(CommandHandler("kullanicilar", kullanicilar_cmd))
        app.add_handler(CommandHandler("santiyeler", santiyeler_cmd))
        app.add_handler(CommandHandler("santiye_durum", santiye_durum_cmd))
        app.add_handler(CommandHandler("maliyet", maliyet_cmd))
        app.add_handler(CommandHandler("ai_rapor", ai_rapor_cmd))
        app.add_handler(CommandHandler("reload", reload_cmd))
        app.add_handler(CommandHandler("yedekle", yedekle_cmd))
        app.add_handler(CommandHandler("chatid", chatid_cmd))
        app.add_handler(CommandHandler("excel_durum", excel_durum_cmd))
        app.add_handler(CommandHandler("reset_database", reset_database_cmd))
        app.add_handler(CommandHandler("fix_sequences", fix_sequences_cmd))
        app.add_handler(CommandHandler("eksik_rapor_excel", eksik_rapor_excel_cmd))
        app.add_handler(CommandHandler("haftalik_eksik_raporlar", haftalik_eksik_raporlar_cmd))
        app.add_handler(CommandHandler("aylik_eksik_raporlar", aylik_eksik_raporlar_cmd))
        
        # Rapor işleme handler'ı
        app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, yeni_gpt_rapor_isleme))
        
        # Yeni üye karşılama
        app.add_handler(MessageHandler(filters.StatusUpdate.NEW_CHAT_MEMBERS, yeni_uye_karşilama))
        
        # Zamanlama job'larını ayarla
        schedule_jobs(app)
        
        # Railway için uygulamayı başlat
        app.run_polling(allowed_updates=Update.ALL_TYPES)
        
    except Exception as e:
        logging.error(f"❌ Bot başlatma hatası: {e}")

if __name__ == '__main__':
    main()